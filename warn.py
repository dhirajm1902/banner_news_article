"""
WARN Act Unified Scraper
========================
Add new states by implementing a scrape_<state>() function and
registering it in SCRAPERS at the bottom of the file.

Each scraper must return a pd.DataFrame with at least these columns
(add state-specific columns freely — they'll be preserved):
    state, company, city, notice_date, layoff_date,
    employees_affected, closure_type, notes
"""

import asyncio
import re
import sys
import time
import logging
import requests
import pandas as pd
from datetime import date, datetime
from pathlib import Path
from io import StringIO

from bs4 import BeautifulSoup
from playwright.async_api import async_playwright

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(name)s | %(message)s")
log = logging.getLogger("warn_scraper")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

def _run_async(coro):
    """
    Run an async coroutine safely from both plain Python scripts and
    Jupyter notebooks (which already have a running event loop).
    """
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = None

    if loop and loop.is_running():
        import nest_asyncio
        nest_asyncio.apply()
        return loop.run_until_complete(coro)
    else:
        if sys.platform == "win32":
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
        return asyncio.run(coro)


OUTPUT_COLS = [
    "state", "company", "city", "notice_date", "layoff_date",
    "employees_affected", "closure_type", "notes",
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _filter_from_2025(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows whose notice_date is before 2025. Rows with unparseable dates are kept."""
    if df.empty or "notice_date" not in df.columns:
        return df
    parsed = pd.to_datetime(df["notice_date"], errors="coerce")
    keep = parsed.isna() | (parsed.dt.year >= 2025)
    return df[keep].reset_index(drop=True)


def _scrape_layoffdata_gsheets(state: str, gsheets_id: str) -> pd.DataFrame:
    """Download a layoffdata.com Google Sheets XLSX and map to OUTPUT_COLS."""
    import io, time
    url = (
        f"https://docs.google.com/spreadsheets/d/{gsheets_id}"
        f"/export?format=xlsx&id={gsheets_id}"
    )
    log.info(f"  {state}: fetching layoffdata.com Google Sheets XLSX")
    sess = requests.Session()
    sess.headers["User-Agent"] = HEADERS["User-Agent"]
    for attempt in range(3):
        try:
            r = sess.get(url, timeout=40)
            r.raise_for_status()
            df = pd.read_excel(io.BytesIO(r.content), dtype=str)
            break
        except Exception as exc:
            if attempt < 2:
                log.warning(f"  {state}: attempt {attempt+1} failed ({exc}), retrying in 30s")
                time.sleep(30)
            else:
                log.error(f"  {state}: all attempts failed: {exc}")
                return _normalise(pd.DataFrame(), state)

    df.columns = [str(c).strip() for c in df.columns]
    df = df[[c for c in df.columns if not c.startswith("Unnamed")]]

    # Combine Closure/Layoff + Temporary/Permanent into closure_type
    cl_col = next((c for c in df.columns if c.lower() in ("closure/layoff", "layoff/closure")), None)
    tp_col = next((c for c in df.columns if c.lower() == "temporary/permanent"), None)
    if cl_col or tp_col:
        parts = []
        for col in [cl_col, tp_col]:
            if col:
                parts.append(df[col].fillna("").str.strip())
        df["closure_type"] = parts[0] if len(parts) == 1 else (
            parts[0] + " - " + parts[1]
        ).str.strip(" -")
        for col in [cl_col, tp_col]:
            if col:
                df.drop(columns=[col], inplace=True)

    df.rename(columns={
        "Company":           "company",
        "City":              "city",
        "Number of Workers": "employees_affected",
        "WARN Received Date": "notice_date",
        "Received Date":     "notice_date",
        "Effective Date":    "layoff_date",
    }, inplace=True)

    # Merge Notes + Region/County into notes (first non-empty wins)
    note_parts = []
    for col in ("Notes", "Region", "County"):
        if col in df.columns:
            note_parts.append(df[col].fillna("").astype(str).str.strip())
            df.drop(columns=[col], inplace=True)
    if note_parts:
        merged = note_parts[0]
        for p in note_parts[1:]:
            merged = merged.where(merged != "", p)
        df["notes"] = merged

    log.info(f"  {state}: {len(df)} rows (pre-filter)")
    return _normalise(df, state)


def _normalise(df: pd.DataFrame, state: str) -> pd.DataFrame:
    """Ensure every required column exists, state is set, and rows are 2025+."""
    df = df.copy()
    # Merge any duplicate column names that arise from multi-column rename maps
    if df.columns.duplicated().any():
        duped = df.columns[df.columns.duplicated(keep=False)].unique()
        for col in duped:
            vals = df.loc[:, df.columns == col]
            merged = vals.apply(
                lambda r: " | ".join(
                    v for v in r.astype(str).str.strip()
                    if v and v not in ("nan", "None", "")
                ),
                axis=1,
            )
            df = df.loc[:, df.columns != col].copy()
            df[col] = merged
    df["state"] = state
    for col in OUTPUT_COLS:
        if col not in df.columns:
            df[col] = ""
    extras = [c for c in df.columns if c not in OUTPUT_COLS]
    return _filter_from_2025(df[OUTPUT_COLS + extras])


# ── State scrapers ────────────────────────────────────────────────────────────

async def _scrape_alabama_async() -> list[dict]:
    data = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()
        await page.goto("https://workforce.alabama.gov/warn-list/")
        await page.wait_for_selector('tr.fw-warn-list__items[data-year="2026"]', timeout=15000)

        rows = await page.query_selector_all(
            'tr.fw-warn-list__items[data-year="2025"], tr.fw-warn-list__items[data-year="2026"]'
        )
        for row in rows:
            status = await row.query_selector('td[data-label="Closing or Layoff"]')
            if not status:
                continue
            status_text = (await status.inner_text()).strip().lower()
            if status_text != "closure":
                continue

            async def _txt(label):
                el = await row.query_selector(f'td[data-label="{label}"]')
                return (await el.inner_text()).strip() if el else ""

            data.append({
                "company":            await _txt("Company"),
                "city":               await _txt("City"),
                "notice_date":        await _txt("Initial Report Date"),
                "layoff_date":        await _txt("Planned Starting Date"),
                "employees_affected": await _txt("Planned # of Affected Employees"),
                "closure_type":       "closure",
                "notes":              "",
            })
        await browser.close()
    return data


def scrape_alabama() -> pd.DataFrame:
    """
    Scrapes Alabama WARN list (2026, closure events only) via async Playwright.
    URL: https://workforce.alabama.gov/warn-list/
    """
    log.info("Scraping Alabama...")
    data = _run_async(_scrape_alabama_async())
    df = pd.DataFrame(data)
    log.info(f"  Alabama: {len(df)} rows")
    return _normalise(df, "Alabama")


def scrape_alaska() -> pd.DataFrame:
    """
    Scrapes Alaska WARN notices via requests + BeautifulSoup.
    URL: https://jobs.alaska.gov/rr/WARN_notices.htm
    """
    log.info("Scraping Alaska...")
    resp = requests.get("https://jobs.alaska.gov/rr/WARN_notices.htm", headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "html.parser")

    data = []
    for row in soup.find_all("tr")[1:]:   # [1:] skips the single header row
        cols = row.find_all("td")
        if len(cols) < 6:
            continue
        data.append({
            "company":            cols[0].get_text(strip=True),
            "city":               cols[1].get_text(strip=True),
            "notice_date":        cols[2].get_text(strip=True),
            "layoff_date":        cols[3].get_text(strip=True),
            "employees_affected": cols[4].get_text(strip=True),
            "closure_type":       "",
            "notes":              cols[5].get_text(strip=True),
        })

    df = pd.DataFrame(data)
    log.info(f"  Alaska: {len(df)} rows")
    return _normalise(df, "Alaska")


async def _scrape_dc_async() -> list[dict]:
    """
    Parse DC WARN table by scraping <tr>/<td> directly via Playwright,
    bypassing pd.read_html entirely.
    Fetches both current year and previous year pages to get full data.
    """
    current_year = date.today().year
    URLS = [
        f"https://does.dc.gov/page/industry-closings-and-layoffs-warn-notifications-{current_year}",
        f"https://does.dc.gov/page/industry-closings-and-layoffs-warn-notifications-{current_year - 1}",
    ]
    SELECTOR = ".field-name-body table, .field-items table, article table"

    col_map = {
        "organization name": "company",
        "company name":      "company",
        "company":           "company",
        "notice date":       "notice_date",
        "date":              "notice_date",
        "effective layoff date": "layoff_date",
        "layoff date":       "layoff_date",
        "number toemployees affected": "employees_affected",
        "number of employees affected": "employees_affected",
        "employees affected": "employees_affected",
        "# employees":       "employees_affected",
        "affected":          "employees_affected",
        "code type":         "closure_type",
        "type":              "closure_type",
        "location":          "city",
        "city":              "city",
    }

    all_records = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        for url in URLS:
            try:
                page = await browser.new_page()
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                await page.wait_for_selector(SELECTOR, timeout=20000)

                rows_data = await page.eval_on_selector(
                    SELECTOR,
                    """tbl => {
                        const rows = Array.from(tbl.querySelectorAll('tr'));
                        return rows.map(r =>
                            Array.from(r.querySelectorAll('th, td'))
                                .map(c => c.innerText.trim())
                        );
                    }"""
                )
                await page.close()

                if not rows_data:
                    continue

                header_idx = next(
                    (i for i, r in enumerate(rows_data) if sum(bool(c) for c in r) >= 3),
                    0
                )
                headers = [h.strip() for h in rows_data[header_idx]]
                data_rows = rows_data[header_idx + 1:]
                norm_headers = [col_map.get(h.lower(), h.lower()) for h in headers]

                for row in data_rows:
                    if not any(row):
                        continue
                    padded = row + [""] * (len(norm_headers) - len(row))
                    all_records.append(dict(zip(norm_headers, padded)))

            except Exception as exc:
                log.warning(f"  DC: failed to scrape {url} — {exc}")

        await browser.close()

    return all_records


def scrape_dc() -> pd.DataFrame:
    """
    Scrapes Washington DC WARN notifications via async Playwright.
    URL: https://does.dc.gov/page/industry-closings-and-layoffs-warn-notifications-2025
    """
    log.info("Scraping DC...")
    records = _run_async(_scrape_dc_async())
    raw = pd.DataFrame(records)
    raw.dropna(how="all", inplace=True)
    raw = raw[raw.apply(lambda r: r.astype(str).str.strip().ne("").any(), axis=1)]

    log.info(f"  DC: {len(raw)} rows")
    return _normalise(raw, "DC")


def scrape_washington() -> pd.DataFrame:
    """
    Scrapes Washington State WARN notices via ASP.NET postback pagination.
    URL: https://fortress.wa.gov/esd/file/WARN/Public/SearchWARN.aspx
    """
    log.info("Scraping Washington...")
    URL = "https://fortress.wa.gov/esd/file/WARN/Public/SearchWARN.aspx"
    session = requests.Session()

    def _get_payload(soup):
        return {i.get("name"): i.get("value", "") for i in soup.select("input") if i.get("name")}

    def _extract_rows(soup):
        rows = []
        for r in soup.select("#ucPSW_gvMain tr"):
            cols = [c.get_text(strip=True) for c in r.select("td")]
            if len(cols) >= 7 and not cols[0].isdigit() and cols[0] not in ("Company", ""):
                rows.append(cols[:7])
        return rows

    res = session.get(URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(res.text, "html.parser")
    payload = _get_payload(soup)
    payload["ucPSW$btnSearchCompany"] = "Search"

    res = session.post(URL, data=payload, timeout=15)
    soup = BeautifulSoup(res.text, "html.parser")

    data = []
    page = 1

    while True:
        log.info(f"  Washington page {page}")
        rows = _extract_rows(soup)
        if not rows:
            log.info("  No rows found, stopping.")
            break
        data.extend(rows)

        pager_links = soup.select("#ucPSW_gvMain a")
        page_numbers = [a.get_text(strip=True) for a in pager_links if a.get_text(strip=True).isdigit()]
        next_page = page + 1

        if str(next_page) not in page_numbers:
            log.info(f"  No more pages after page {page}.")
            break

        payload = _get_payload(soup)
        payload["__EVENTTARGET"] = "ucPSW$gvMain"
        payload["__EVENTARGUMENT"] = f"Page${next_page}"
        payload.pop("ucPSW$btnSearchCompany", None)

        res = session.post(URL, data=payload, timeout=15)
        soup = BeautifulSoup(res.text, "html.parser")
        page += 1

    raw_cols = ["company", "city", "layoff_date", "employees_affected", "closure_type", "type_detail", "notice_date"]
    df = pd.DataFrame(data, columns=raw_cols)
    log.info(f"  Washington: {len(df)} rows")
    return _normalise(df, "Washington")


# ── Washington standalone runner (mirrors original script behaviour) ──────────

def run_washington_standalone(
    output_dir: str = ".",
    max_pages: int | None = None,
) -> pd.DataFrame:
    """
    Reproduces the original standalone Washington script exactly:
    scrapes all pages (or up to max_pages), prints progress, saves a
    dated CSV, and returns the raw DataFrame with the original column names.
    """
    URL = "https://fortress.wa.gov/esd/file/WARN/Public/SearchWARN.aspx"
    session = requests.Session()

    def _get_payload(soup):
        return {i.get("name"): i.get("value", "") for i in soup.select("input") if i.get("name")}

    def _extract_rows(soup):
        rows = []
        for r in soup.select("#ucPSW_gvMain tr"):
            cols = [c.get_text(strip=True) for c in r.select("td")]
            if len(cols) >= 7 and not cols[0].isdigit() and cols[0] not in ("Company", ""):
                rows.append(cols[:7])
        return rows

    res = session.get(URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(res.text, "html.parser")
    payload = _get_payload(soup)
    payload["ucPSW$btnSearchCompany"] = "Search"

    res = session.post(URL, data=payload, timeout=15)
    soup = BeautifulSoup(res.text, "html.parser")

    data = []
    page = 1

    while True:
        print(f"Scraping page: {page}")

        if max_pages and page > max_pages:
            print(f"Reached maximum pages ({max_pages}), stopping.")
            break

        rows = _extract_rows(soup)
        if not rows:
            print("No rows found, stopping.")
            break
        data.extend(rows)

        pager_links = soup.select("#ucPSW_gvMain a")
        page_numbers = [a.get_text(strip=True) for a in pager_links if a.get_text(strip=True).isdigit()]
        next_page = page + 1

        if str(next_page) not in page_numbers:
            print(f"No more pages after page {page}.")
            break

        payload = _get_payload(soup)
        payload["__EVENTTARGET"] = "ucPSW$gvMain"
        payload["__EVENTARGUMENT"] = f"Page${next_page}"
        payload.pop("ucPSW$btnSearchCompany", None)

        res = session.post(URL, data=payload, timeout=15)
        soup = BeautifulSoup(res.text, "html.parser")
        page += 1

    columns = ["Company", "Location", "Layoff Start Date", "# Workers", "Closure/Layoff", "Type", "Received Date"]
    df = pd.DataFrame(data, columns=columns)

    print(df.head())
    print("Total rows:", len(df))

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    csv_path = out / f"warn_layoffs_{date.today().strftime('%Y-%m-%d')}.csv"
    df.to_csv(csv_path, index=False)
    print(f"Saved to {csv_path}")

    return df


def scrape_maryland() -> pd.DataFrame:
    """
    Scrapes Maryland WARN/ESA table via requests + pd.read_html.
    URL: https://labor.maryland.gov/employment/warn.shtml
    """
    log.info("Scraping Maryland...")
    resp = requests.get("https://labor.maryland.gov/employment/warn.shtml", headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    table = soup.find("table")
    if not table:
        raise RuntimeError("Maryland: no table found on page.")

    raw = pd.read_html(StringIO(str(table)), header=0)[0]
    raw.columns = [c.strip().replace("\n", " ") for c in raw.columns]
    raw.dropna(how="all", inplace=True)
    raw = raw.loc[:, raw.columns.notna()]

    col_map = {
        "Company Name":           "company",
        "Company":                "company",
        "City":                   "city",
        "Location":               "city",
        "Notice Date":            "notice_date",
        "Effective Date":         "layoff_date",
        "Layoff Date":            "layoff_date",
        "Number of Employees":    "employees_affected",
        "Employees Affected":     "employees_affected",
        "Type":                   "closure_type",
        "Notes":                  "notes",
    }
    raw.rename(columns={k: v for k, v in col_map.items() if k in raw.columns}, inplace=True)

    log.info(f"  Maryland: {len(raw)} rows")
    return _normalise(raw, "Maryland")


# ── Vermont ───────────────────────────────────────────────────────────────────

_VT_RESULTS_URL = (
    "https://www.vermontjoblink.com/search/warn_lookups"
    "?commit=Search"
    "&q%5Bemployer_name_cont%5D="
    "&q%5Bmain_contact_contact_info_addresses_full_location_city_matches%5D="
    "&q%5Bnotice_eq%5D=true"
    "&q%5Bnotice_on_gteq%5D="
    "&q%5Bnotice_on_lteq%5D="
    "&q%5Bservice_delivery_area_id_eq%5D="
    "&q%5Bzipcode_code_start%5D="
)

_VT_COLUMNS = ["Employer", "City", "ZIP", "LWIB Area", "Notice Date", "WARN Type"]


def _vt_parse_rows(soup: BeautifulSoup) -> list[dict]:
    import re
    table = soup.find("table", {"id": re.compile(r"^a11y_table_")})
    if not table:
        return []
    tbody = table.find("tbody")
    if not tbody:
        return []

    rows = []
    for tr in tbody.find_all("tr", recursive=False):
        tds = tr.find_all("td", recursive=False)
        if not tds:
            continue
        cells = [re.sub(r"\s+", " ", td.get_text(" ", strip=True)).strip() for td in tds]
        while len(cells) < len(_VT_COLUMNS):
            cells.append("")
        row = dict(zip(_VT_COLUMNS, cells[: len(_VT_COLUMNS)]))
        if row["Employer"]:
            rows.append(row)
    return rows


def _vt_find_next_url(soup: BeautifulSoup, current_url: str) -> str | None:
    import re
    from urllib.parse import urljoin

    nav = soup.find("div", class_=re.compile(r"pagination", re.I)) or \
          soup.find("nav", attrs={"aria-label": re.compile(r"Page controls", re.I)})
    if not nav:
        return None

    a = nav.find("a", attrs={"rel": "next"}) or \
        nav.find("a", class_=re.compile(r"next_page", re.I))
    if a and a.get("href"):
        return urljoin(current_url, a["href"])

    for link in nav.find_all("a", href=True):
        if "next" in link.get_text(" ", strip=True).lower():
            return urljoin(current_url, link["href"])
    return None


def scrape_vermont(max_pages: int | None = None) -> pd.DataFrame:
    """
    Scrapes Vermont WARN notices via requests + BeautifulSoup with pagination.
    URL: https://www.vermontjoblink.com/search/warn_lookups/new
    """
    log.info("Scraping Vermont...")
    session = requests.Session()
    session.headers.update(HEADERS)

    resp = session.get(_VT_RESULTS_URL, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    current_url = resp.url

    all_rows: list[dict] = []
    seen: set[tuple] = set()
    page_num = 1

    while True:
        rows = _vt_parse_rows(soup)
        for r in rows:
            key = tuple(r.get(col, "") for col in _VT_COLUMNS)
            if key not in seen:
                seen.add(key)
                all_rows.append(r)

        log.info(f"  Vermont page {page_num}: {len(rows)} rows | total: {len(all_rows)}")

        if max_pages and page_num >= max_pages:
            break

        next_url = _vt_find_next_url(soup, current_url)
        if not next_url:
            break

        time.sleep(1.0)
        resp = session.get(next_url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        current_url = resp.url
        page_num += 1

    raw = pd.DataFrame(all_rows)
    if raw.empty:
        raw = pd.DataFrame(columns=_VT_COLUMNS)

    # Map Vermont columns → standard schema
    raw.rename(columns={
        "Employer":    "company",
        "City":        "city",
        "Notice Date": "notice_date",
        "WARN Type":   "closure_type",
        "ZIP":         "zip",
        "LWIB Area":   "lwib_area",
    }, inplace=True)

    raw["layoff_date"] = ""
    raw["employees_affected"] = ""
    raw["notes"] = ""

    log.info(f"  Vermont: {len(raw)} rows total")
    return _normalise(raw, "Vermont")


# ── Texas ─────────────────────────────────────────────────────────────────────

_TX_CSV_URL = (
    "https://data.texas.gov/api/views/8w53-c4f6/rows.csv?accessType=DOWNLOAD"
)


def scrape_texas() -> pd.DataFrame:
    """
    Downloads the full Texas WARN dataset directly from the Socrata open-data API.
    URL: https://data.texas.gov/dataset/Worker-Adjustment-and-Retraining-Notification-WARN/8w53-c4f6
    """
    log.info("Scraping Texas (direct CSV download)...")
    resp = requests.get(_TX_CSV_URL, headers=HEADERS, timeout=120)
    resp.raise_for_status()

    raw = pd.read_csv(StringIO(resp.text), low_memory=False)
    raw.columns = [c.strip() for c in raw.columns]
    log.info(f"  Texas raw columns: {raw.columns.tolist()}")

    # Actual Texas Socrata CSV uses ALL_CAPS_UNDERSCORE column names.
    # Map both the real names and common title-case variants as fallback.
    col_map = {
        # company — actual Socrata name
        "JOB_SITE_NAME":            "company",
        # city
        "CITY_NAME":                "city",
        "City":                     "city",
        # notice date
        "NOTICE_DATE":              "notice_date",
        "Notice Date":              "notice_date",
        # layoff / effective date
        "LayOff_Date":              "layoff_date",
        "LAYOFF_DATE":              "layoff_date",
        "Layoff Date":              "layoff_date",
        "Effective Date":           "layoff_date",
        # employees
        "TOTAL_LAYOFF_NUMBER":      "employees_affected",
        "# Employees Affected":     "employees_affected",
        "Employees Affected":       "employees_affected",
        "Number of Affected Workers": "employees_affected",
        # closure type (Texas dataset has no dedicated type column; leave blank)
        "Type of Layoff/Closure":   "closure_type",
        "Type of Layoff":           "closure_type",
        "Closure/Layoff":           "closure_type",
        # notes
        "Notes":                    "notes",
        "Comments":                 "notes",
    }
    raw.rename(columns={k: v for k, v in col_map.items() if k in raw.columns}, inplace=True)

    log.info(f"  Texas: {len(raw)} rows")
    return _normalise(raw, "Texas")


# ── Virginia ──────────────────────────────────────────────────────────────────

async def _scrape_virginia_async(max_pages: int | None = None) -> list[dict]:
    import re

    URL = "https://virginiaworks.gov/im-an-employer/retain-and-grow/warn-notices/"
    REQUEST_DELAY_MS = 800

    def _split_company_address(first_td_html: str) -> tuple[str, str]:
        """Split the first cell into company name and address."""
        from bs4 import BeautifulSoup as BS
        td = BS(first_td_html, "html.parser")
        text = td.get_text("\n", strip=True)
        parts = [re.sub(r"\s+", " ", p).strip() for p in text.split("\n") if p.strip()]
        if not parts:
            return "", ""
        return parts[0], ", ".join(parts[1:]) if len(parts) > 1 else ""

    def _parse_table(html: str) -> list[dict]:
        from bs4 import BeautifulSoup as BS
        soup = BS(html, "html.parser")
        table = soup.find("table", {"id": "warn-notice-table"})
        if not table:
            return []
        tbody = table.find("tbody")
        if not tbody:
            return []

        rows = []
        for tr in tbody.find_all("tr", recursive=False):
            tds = tr.find_all("td", recursive=False)
            if not tds:
                continue

            company, address = _split_company_address(str(tds[0]))
            other = [re.sub(r"\s+", " ", td.get_text(" ", strip=True)).strip() for td in tds[1:]]
            while len(other) < 7:
                other.append("")

            row = {
                "company":            company,
                "address":            address,
                "notice_date":        other[0],
                "layoff_date":        other[1],
                "employees_affected": other[2],
                "city":               other[3],
                "contact_person":     other[4],
                "closure_type":       other[5],
                "collective_bargaining_unit": other[6],
            }
            if row["company"]:
                rows.append(row)
        return rows

    all_rows: list[dict] = []
    seen: set[tuple] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width": 1400, "height": 1200})

        log.info(f"  Virginia: loading {URL}")
        await page.goto(URL, wait_until="networkidle", timeout=60000)
        await page.wait_for_selector("#warn-notice-table tbody tr", timeout=60000)

        page_num = 1
        while True:
            html = await page.content()
            rows = _parse_table(html)

            added = 0
            for r in rows:
                key = (r["company"], r["notice_date"], r["layoff_date"])
                if key not in seen:
                    seen.add(key)
                    all_rows.append(r)
                    added += 1

            log.info(f"  Virginia page {page_num}: {len(rows)} rows | new: {added} | total: {len(all_rows)}")

            if max_pages and page_num >= max_pages:
                break

            next_btn = page.locator('button.dt-paging-button.next[aria-label="Next"]')
            if await next_btn.count() == 0:
                break
            aria_disabled = await next_btn.first.get_attribute("aria-disabled")
            if aria_disabled == "true":
                break

            prev_first = rows[0]["company"] if rows else ""
            await next_btn.first.click()
            await page.wait_for_timeout(REQUEST_DELAY_MS)
            await page.wait_for_load_state("networkidle")

            try:
                await page.wait_for_function(
                    """(prev) => {
                        const cell = document.querySelector('#warn-notice-table tbody tr td');
                        return cell && cell.textContent.trim() !== prev;
                    }""",
                    arg=prev_first,
                    timeout=15000,
                )
            except Exception:
                pass

            page_num += 1

        await browser.close()

    return all_rows


def scrape_virginia(max_pages: int | None = None) -> pd.DataFrame:
    """
    Scrapes Virginia WARN notices via async Playwright (DataTables pagination).
    URL: https://virginiaworks.gov/im-an-employer/retain-and-grow/warn-notices/
    """
    log.info("Scraping Virginia...")
    data = _run_async(_scrape_virginia_async(max_pages=max_pages))
    df = pd.DataFrame(data) if data else pd.DataFrame()
    log.info(f"  Virginia: {len(df)} rows")
    return _normalise(df, "Virginia")


# ── Ohio ─────────────────────────────────────────────────────────────────────
# Uses Selenium (lazy imports so the rest of warn.py works without selenium installed).

_OH_URLS = {
    "2025": (
        "https://jfs.ohio.gov/job-services-and-unemployment/"
        "job-services/job-programs-and-services/submit-a-warn-notice/"
        "2025-Public-Notice-of-Layoffs-and-Closures"
    ),
    "2026": (
        "https://jfs.ohio.gov/job-services-and-unemployment/"
        "job-services/job-programs-and-services/submit-a-warn-notice/"
        "current-public-notices-of-layoffs-and-closures-sa"
    ),
}
_OH_TABLE_SELS = (
    "table#js-table-visualization",
    "table.dataTable",
    "table",
)
_OH_ROW_SELS  = tuple(f"{s} tbody tr" for s in _OH_TABLE_SELS)
_OH_NEXT_SELS = (
    "button.dt-paging-button.next",
    "a.paginate_button.next",
    "button[aria-label='Next']",
    "a[aria-label='Next']",
    "#js-table-visualization_next",
)


def _oh_build_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1600,1200")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"},
    )
    return driver


def _oh_count_rows(driver) -> int:
    from selenium.webdriver.common.by import By
    for sel in _OH_ROW_SELS:
        try:
            n = len(driver.find_elements(By.CSS_SELECTOR, sel))
            if n:
                return n
        except Exception:
            pass
    return 0


def _oh_switch_iframe(driver) -> bool:
    from selenium.webdriver.common.by import By
    driver.switch_to.default_content()
    for frame in driver.find_elements(By.TAG_NAME, "iframe"):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            if driver.find_elements(By.CSS_SELECTOR, "table"):
                return True
        except Exception:
            continue
    driver.switch_to.default_content()
    return False


def _oh_find_table(driver, timeout=30):
    from selenium.webdriver.common.by import By
    end = time.time() + timeout
    while time.time() < end:
        for sel in _OH_TABLE_SELS:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    if el.is_displayed():
                        return el
                if els and sel == "table#js-table-visualization":
                    return els[0]
            except Exception:
                pass
        time.sleep(0.3)
    return None


def _oh_set_page_length(driver) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC

    for sel in ("select#dt-length-0", "select[name$='_length']", "select[aria-controls]"):
        try:
            el = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
            s = Select(el)
            values = [o.get_attribute("value") for o in s.options]
            for target in ("-1", "100"):
                if target in values:
                    s.select_by_value(target)
                    time.sleep(2.5)
                    return
            s.select_by_index(len(s.options) - 1)
            time.sleep(2.5)
            return
        except Exception:
            continue


def _oh_rows_from_embedded_json(driver, url: str) -> list[dict]:
    import json as _json
    src = driver.page_source or ""
    start = src.find('{"data":')
    if start < 0:
        return []
    depth, in_str, esc, end = 0, False, False, None
    for i in range(start, len(src)):
        ch = src[i]
        if in_str:
            esc = (ch == "\\") and not esc
            if not esc and ch == '"':
                in_str = False
        elif ch == '"':
            in_str = True
        elif ch == '{':
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    if not end:
        return []
    try:
        payload = _json.loads(src[start:end])
    except Exception:
        return []
    data = payload.get("data") or []
    if len(data) < 3:
        return []
    rows = []
    for rec in data[2:]:
        if not isinstance(rec, list) or len(rec) < 9:
            continue
        company = str(rec[0] or "").strip()
        if not company:
            continue
        phone = str(rec[7] or "").strip()
        union = str(rec[8] or "").strip()
        notes_parts = [f"Phone: {phone}" if phone else "", f"Union: {union}" if union else ""]
        rows.append({
            "company":            company,
            "notice_date":        str(rec[1] or "").strip(),
            "city":               str(rec[3] or "").strip(),
            "closure_type":       str(rec[4] or "").strip(),
            "employees_affected": str(rec[5] or "").strip(),
            "layoff_date":        str(rec[6] or "").strip(),
            "notes":              " | ".join(p for p in notes_parts if p),
        })
    return rows


def _oh_rows_from_page(driver, url: str) -> list[dict]:
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException

    trs = []
    for sel in _OH_ROW_SELS:
        try:
            trs = driver.find_elements(By.CSS_SELECTOR, sel)
            if trs:
                break
        except Exception:
            pass

    rows = []
    for tr in trs:
        try:
            tds = tr.find_elements(By.TAG_NAME, "td")
            if len(tds) < 2:
                continue
            company = (tds[0].text or "").strip()
            if not company or company.lower() in ("company", "employer"):
                continue

            pdf_url = ""
            try:
                a = tds[0].find_element(By.CSS_SELECTOR, "a[href]")
                pdf_url = (a.get_attribute("href") or "").strip()
            except NoSuchElementException:
                pass

            def cell(i):
                return (tds[i].text or "").strip() if i < len(tds) else ""

            phone, union = cell(6), cell(7)
            notes_parts = [
                f"Phone: {phone}" if phone else "",
                f"Union: {union}" if union else "",
                f"PDF: {pdf_url}" if pdf_url else "",
            ]
            rows.append({
                "company":            company,
                "notice_date":        cell(1),
                "city":               cell(2),
                "closure_type":       cell(3),
                "employees_affected": cell(4),
                "layoff_date":        cell(5),
                "notes":              " | ".join(p for p in notes_parts if p),
            })
        except StaleElementReferenceException:
            continue
        except Exception:
            continue
    return rows


def _oh_click_next(driver) -> bool:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait

    before = _oh_count_rows(driver)
    for sel in _OH_NEXT_SELS:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, sel)
            btn = next((b for b in btns if b.is_displayed()), None)
            if not btn:
                continue
            cls = (btn.get_attribute("class") or "").lower()
            if "disabled" in cls or btn.get_attribute("aria-disabled") == "true":
                continue
            driver.execute_script("arguments[0].click();", btn)
            try:
                WebDriverWait(driver, 15).until(lambda d: _oh_count_rows(d) != before)
            except Exception:
                time.sleep(2.5)
            time.sleep(1.0)
            return True
        except Exception:
            continue
    return False


def _oh_scrape_year(driver, year: str, url: str) -> list[dict]:
    from selenium.webdriver.support.ui import WebDriverWait

    log.info(f"  Ohio [{year}] loading {url}")
    driver.switch_to.default_content()
    driver.get(url)
    try:
        WebDriverWait(driver, 45).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except Exception:
        pass
    time.sleep(5.0 if year == "2025" else 3.0)

    _oh_switch_iframe(driver)
    table = _oh_find_table(driver, timeout=30)
    if not table:
        driver.refresh()
        time.sleep(5.0)
        _oh_switch_iframe(driver)
        table = _oh_find_table(driver, timeout=30)

    if not table:
        fallback = _oh_rows_from_embedded_json(driver, url)
        if fallback:
            log.info(f"  Ohio [{year}] {len(fallback)} rows from embedded JSON fallback")
            return fallback
        log.warning(f"  Ohio [{year}] no table found, skipping")
        return []

    _oh_set_page_length(driver)
    time.sleep(1.0)

    all_rows, seen, page_no = [], set(), 1
    while True:
        for r in _oh_rows_from_page(driver, url):
            key = (r["company"], r["notice_date"], r["city"])
            if key not in seen:
                seen.add(key)
                all_rows.append(r)
        log.info(f"  Ohio [{year}] page {page_no}: {len(all_rows)} total")
        if not _oh_click_next(driver):
            break
        page_no += 1
        if page_no > 100:
            log.warning(f"  Ohio [{year}] safety-stopped at 100 pages")
            break

    return all_rows


def scrape_ohio() -> pd.DataFrame:
    """
    Scrapes Ohio WARN notices for 2025 and 2026 via Selenium.
    2025: https://jfs.ohio.gov/.../2025-Public-Notice-of-Layoffs-and-Closures
    2026: https://jfs.ohio.gov/.../current-public-notices-of-layoffs-and-closures-sa
    """
    log.info("Scraping Ohio...")
    driver = None
    try:
        driver = _oh_build_driver()
        rows: list[dict] = []
        for year, url in _OH_URLS.items():
            rows.extend(_oh_scrape_year(driver, year, url))
    except Exception as e:
        log.error(f"  Ohio scraping failed: {e}")
        rows = []
    finally:
        if driver:
            driver.quit()

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    log.info(f"  Ohio: {len(df)} rows")
    return _normalise(df, "Ohio")


# ── Oregon ───────────────────────────────────────────────────────────────────
# Pure requests + BeautifulSoup — no Selenium needed.
# Flow: GET form → grab hidden tokens → POST form → parse download link → GET xlsx.

_OR_BASE_URL     = "https://ccwd.hecc.oregon.gov"
_OR_DOWNLOAD_URL = f"{_OR_BASE_URL}/Layoff/WARN/Download"
_OR_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Referer": _OR_DOWNLOAD_URL,
}


def _or_get_form_tokens(session: requests.Session) -> dict:
    r = session.get(_OR_DOWNLOAD_URL, headers=_OR_HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    return {
        inp.get("name"): inp.get("value", "")
        for inp in soup.find_all("input", {"type": "hidden"})
        if inp.get("name")
    }


def _or_submit_form(session: requests.Session, tokens: dict) -> BeautifulSoup:
    form_data = {**tokens, "WARNFormat": "xlsx", "WARNSort": "NoticeDate"}
    r = session.post(
        _OR_DOWNLOAD_URL,
        data=form_data,
        headers={**_OR_HEADERS, "Content-Type": "application/x-www-form-urlencoded"},
        timeout=60,
        allow_redirects=True,
    )
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")


def _or_find_download_link(soup: BeautifulSoup) -> str | None:
    import re as _re
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if _re.search(r"/Layoff/Reports/WARN.*\.(xlsx|xls|csv)", href, _re.I):
            return (_OR_BASE_URL + href) if href.startswith("/") else href
    m = _re.search(r'href="(/Layoff/Reports/WARN[^"]+\.(xlsx|xls|csv))"', str(soup), _re.I)
    if m:
        return _OR_BASE_URL + m.group(1)
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "WARN" in href.upper() and _re.search(r"\.(xlsx|xls|csv)$", href, _re.I):
            return (_OR_BASE_URL + href) if href.startswith("/") else href
    return None


def scrape_oregon() -> pd.DataFrame:
    """
    Downloads Oregon WARN data via form POST (no Selenium required).
    URL: https://ccwd.hecc.oregon.gov/Layoff/WARN/Download
    """
    import re as _re
    from io import BytesIO

    log.info("Scraping Oregon...")
    session = requests.Session()
    try:
        tokens  = _or_get_form_tokens(session)
        soup    = _or_submit_form(session, tokens)
        dl_link = _or_find_download_link(soup)
        if not dl_link:
            log.warning("  Oregon: download link not found in response")
            return _normalise(pd.DataFrame(), "Oregon")

        log.info(f"  Oregon: downloading {dl_link}")
        r = session.get(dl_link, headers=_OR_HEADERS, timeout=120)
        r.raise_for_status()

        ext_m = _re.search(r"\.(xlsx|xls|csv)$", dl_link, _re.I)
        ext   = ext_m.group(0).lower() if ext_m else ".xlsx"
        raw   = (pd.read_excel(BytesIO(r.content), engine="openpyxl")
                 if ext in (".xlsx", ".xls")
                 else pd.read_csv(BytesIO(r.content)))

    except Exception as e:
        log.error(f"  Oregon scraping failed: {e}")
        return _normalise(pd.DataFrame(), "Oregon")

    raw.columns = [str(c).strip() for c in raw.columns]
    log.info(f"  Oregon raw columns: {raw.columns.tolist()}")

    col_map = {
        "Employer":            "company",
        "Company":             "company",
        "Company Name":        "company",
        "Firm Name":           "company",
        "City":                "city",
        "Location":            "city",
        "Notice Date":         "notice_date",
        "Received Date":       "notice_date",
        "Layoff Date":         "layoff_date",
        "Effective Date":      "layoff_date",
        "Employees Affected":  "employees_affected",
        "Number of Employees": "employees_affected",
        "Workers Affected":    "employees_affected",
        "# Employees":         "employees_affected",
        "Type":                "closure_type",
        "Closure Type":        "closure_type",
        "Layoff/Closure":      "closure_type",
        "Notice Type":         "closure_type",
        "Notes":               "notes",
        "Comments":            "notes",
    }
    raw.rename(columns={k: v for k, v in col_map.items() if k in raw.columns}, inplace=True)

    log.info(f"  Oregon: {len(raw)} rows")
    return _normalise(raw, "Oregon")


# ── Oklahoma ─────────────────────────────────────────────────────────────────
_OK_URL = "https://www.employoklahoma.gov/Participants/s/warnnotices"


async def _ok_scrape_async():
    from bs4 import BeautifulSoup as _BS
    all_rows = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        await page.goto(_OK_URL, wait_until="networkidle", timeout=60000)
        await page.wait_for_timeout(3000)

        page_no = 1
        while True:
            content = await page.content()
            soup = _BS(content, "html.parser")
            trs = soup.find_all("tr", attrs={"data-row-number": True})
            if not trs:
                break
            for tr in trs:
                row: dict[str, str] = {}
                for cell in tr.find_all(attrs={"data-label": True}):
                    label = (cell.get("data-label") or "").strip()
                    value = (cell.get("data-cell-value") or cell.get_text(" ", strip=True)).strip()
                    if label:
                        row[label] = value
                if row.get("Employer"):
                    all_rows.append(row)

            log.info(f"  Oklahoma page {page_no}: {len(trs)} rows")

            # Check if Next is enabled
            next_btn = await page.query_selector("button:has-text('Next')")
            if not next_btn:
                break
            disabled = await next_btn.get_attribute("aria-disabled")
            if (disabled or "").lower() == "true":
                break

            prev_first = trs[0].get("data-row-number") if trs else None
            await next_btn.click()
            # Wait until row numbers change
            for _ in range(20):
                await page.wait_for_timeout(1000)
                new_content = await page.content()
                new_soup = _BS(new_content, "html.parser")
                new_trs = new_soup.find_all("tr", attrs={"data-row-number": True})
                if new_trs and new_trs[0].get("data-row-number") != prev_first:
                    break
            page_no += 1
            if page_no > 100:
                break

        await browser.close()
    return all_rows


def scrape_oklahoma() -> pd.DataFrame:
    log.info("Scraping Oklahoma...")
    try:
        all_rows = _run_async(_ok_scrape_async())
    except Exception as exc:
        log.error(f"  Oklahoma scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Oklahoma")

    raw = pd.DataFrame(all_rows) if all_rows else pd.DataFrame()
    if not raw.empty:
        raw.rename(columns={
            "Employer": "company", "City": "city",
            "Notice Date": "notice_date", "Notice Type": "closure_type",
            "Local Workforce Board": "notes",
        }, inplace=True)
    log.info(f"  Oklahoma: {len(raw)} rows")
    return _normalise(raw, "Oklahoma")


# ── Pennsylvania ──────────────────────────────────────────────────────────────
# JS-rendered nested accordions: Year → Month → Company → panel text.
# Uses Selenium (lazy imports).

_PA_URL = (
    "https://www.pa.gov/agencies/dli/programs-services/"
    "workforce-development-home/warn-requirements/warn-notices"
)
_PA_YEARS = ["2025", "2026"]


def _pa_build_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    svc = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=svc, options=opts)


def _pa_safe_click(driver, element) -> bool:
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", element)
        return True
    except Exception:
        return False


def _pa_expand_if_collapsed(driver, btn) -> bool:
    from selenium.common.exceptions import StaleElementReferenceException
    try:
        if (btn.get_attribute("aria-expanded") or "false").lower() == "false":
            _pa_safe_click(driver, btn)
            time.sleep(0.5)
        return True
    except StaleElementReferenceException:
        return False


def _pa_parse_panel_text(raw_text: str) -> dict:
    lines = [l.strip() for l in raw_text.strip().splitlines() if l.strip()]
    result = {"address": "", "county": "", "employees_affected": "",
              "notice_date": "", "closure_type": ""}
    address_lines = []
    for line in lines:
        lu = line.upper()
        if "COUNTY:" in lu:
            result["county"] = re.sub(r"COUNTY\s*:\s*", "", line, flags=re.I).strip()
        elif "AFFECTED" in lu:
            result["employees_affected"] = re.sub(r"#?\s*AFFECTED\s*:\s*", "", line, flags=re.I).strip()
        elif "EFFECTIVE DATE" in lu:
            result["notice_date"] = re.sub(r"EFFECTIVE DATE\s*:\s*", "", line, flags=re.I).strip()
        elif "CLOSURE OR LAYOFF" in lu or "CLOSING OR LAYOFF" in lu:
            result["closure_type"] = re.sub(
                r"(CLOSURE OR LAYOFF|CLOSING OR LAYOFF)\s*:\s*", "", line, flags=re.I
            ).strip()
        elif not any(k in lu for k in ["COUNTY", "AFFECTED", "EFFECTIVE", "CLOSURE", "CLOSING"]):
            if not result["county"]:
                address_lines.append(line)
    result["address"] = " ".join(address_lines).strip()
    return result


def _pa_scrape_year(driver, year: str) -> list[dict]:
    from selenium.webdriver.common.by import By
    from selenium.common.exceptions import NoSuchElementException

    all_rows = []
    year_headings = driver.find_elements(
        By.XPATH,
        f"//h2[contains(@class,'cmp-accordion__main-heading') and normalize-space(text())='{year}']"
    )
    if not year_headings:
        log.warning(f"  Pennsylvania [{year}]: year heading not found")
        return all_rows

    year_section = driver.execute_script("""
        var headings = document.querySelectorAll(
            'h2.cmp-accordion__main-heading--large, h2.cmp-accordion__main-heading');
        for (var i = 0; i < headings.length; i++) {
            if (headings[i].textContent.trim() === arguments[0]) {
                var el = headings[i].parentElement;
                while (el && !el.querySelector('.cmp-accordion__button')) {
                    el = el.nextElementSibling || el.parentElement.nextElementSibling;
                    if (!el) break;
                }
                return el;
            }
        }
        return null;
    """, year)

    if not year_section:
        log.warning(f"  Pennsylvania [{year}]: accordion section not found")
        return all_rows

    # Month-level buttons
    month_btns = year_section.find_elements(
        By.CSS_SELECTOR,
        ".cmp-accordion > .cmp-accordion__item > .cmp-accordion__header > .cmp-accordion__button"
    )
    if not month_btns:
        month_btns = year_section.find_elements(
            By.XPATH,
            ".//button[contains(@class,'cmp-accordion__button')]"
            "[.//span[contains(@class,'cmp-accordion__title') and ("
            "normalize-space(text())='January' or normalize-space(text())='February' or "
            "normalize-space(text())='March' or normalize-space(text())='April' or "
            "normalize-space(text())='May' or normalize-space(text())='June' or "
            "normalize-space(text())='July' or normalize-space(text())='August' or "
            "normalize-space(text())='September' or normalize-space(text())='October' or "
            "normalize-space(text())='November' or normalize-space(text())='December'"
            ")]]"
        )

    log.info(f"  Pennsylvania [{year}]: {len(month_btns)} month(s)")

    for month_btn in month_btns:
        try:
            month_name = month_btn.find_element(By.CSS_SELECTOR, ".cmp-accordion__title").text.strip()
        except Exception:
            month_name = "Unknown"

        _pa_expand_if_collapsed(driver, month_btn)
        time.sleep(0.8)

        panel_id = month_btn.get_attribute("aria-controls")
        try:
            month_panel = (driver.find_element(By.ID, panel_id) if panel_id
                           else driver.execute_script(
                               "return arguments[0].closest('.cmp-accordion__item')"
                               ".querySelector('.cmp-accordion__panel');", month_btn))
        except NoSuchElementException:
            continue
        if not month_panel:
            continue

        company_btns = month_panel.find_elements(By.CSS_SELECTOR, ".cmp-accordion__button")

        for comp_btn in company_btns:
            try:
                company_name = comp_btn.find_element(By.CSS_SELECTOR, ".cmp-accordion__title").text.strip()
            except Exception:
                company_name = comp_btn.text.strip().split("\n")[0].strip()
            if not company_name:
                continue

            _pa_expand_if_collapsed(driver, comp_btn)
            time.sleep(0.3)

            comp_panel_id = comp_btn.get_attribute("aria-controls")
            try:
                comp_panel = (driver.find_element(By.ID, comp_panel_id) if comp_panel_id
                              else driver.execute_script(
                                  "return arguments[0].closest('.cmp-accordion__item')"
                                  ".querySelector('.cmp-accordion__panel');", comp_btn))
            except NoSuchElementException:
                comp_panel = None
            if not comp_panel:
                continue

            try:
                raw_text = comp_panel.text.strip()
                if not raw_text:
                    soup = BeautifulSoup(comp_panel.get_attribute("innerHTML"), "html.parser")
                    raw_text = soup.get_text(separator="\n", strip=True)
            except Exception:
                raw_text = ""
            if not raw_text:
                continue

            parsed = _pa_parse_panel_text(raw_text)
            all_rows.append({
                "company":            company_name,
                "city":               parsed["address"],
                "notice_date":        parsed["notice_date"],
                "layoff_date":        parsed["notice_date"],
                "employees_affected": parsed["employees_affected"],
                "closure_type":       parsed["closure_type"],
                "notes":              f"County: {parsed['county']}" if parsed["county"] else "",
            })

    log.info(f"  Pennsylvania [{year}]: {len(all_rows)} rows")
    return all_rows


def scrape_pennsylvania() -> pd.DataFrame:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException

    log.info("Scraping Pennsylvania...")
    driver = _pa_build_driver()
    rows: list[dict] = []
    try:
        driver.get(_PA_URL)
        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".cmp-accordion__button"))
            )
            time.sleep(2)
        except TimeoutException:
            log.error("  Pennsylvania: accordion buttons did not appear")
            return _normalise(pd.DataFrame(), "Pennsylvania")

        for year in _PA_YEARS:
            try:
                rows.extend(_pa_scrape_year(driver, year))
            except Exception as exc:
                log.error(f"  Pennsylvania [{year}]: {exc}")
    finally:
        driver.quit()

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    log.info(f"  Pennsylvania: {len(df)} rows total")
    return _normalise(df, "Pennsylvania")


# ── South Dakota ──────────────────────────────────────────────────────────────
# Pure requests + BeautifulSoup — no Selenium needed.
# URL: https://dlr.sd.gov/workforce_services/businesses/warn_notices.aspx

_SD_URL      = "https://dlr.sd.gov/workforce_services/businesses/warn_notices.aspx"
_SD_BASE_URL = "https://dlr.sd.gov"


def _sd_resolve_pdf(href: str) -> str:
    if not href:
        return ""
    href = href.strip()
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return _SD_BASE_URL + href
    return _SD_BASE_URL + "/workforce_services/businesses/" + href


def scrape_south_dakota() -> pd.DataFrame:
    """
    Scrapes South Dakota WARN notices via requests + BeautifulSoup.
    URL: https://dlr.sd.gov/workforce_services/businesses/warn_notices.aspx
    """
    log.info("Scraping South Dakota...")
    resp = requests.get(_SD_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "lxml")

    table = soup.find("table", {"bordercolor": "003366"}) or soup.find("table")
    if not table:
        log.warning("  South Dakota: WARN table not found")
        return _normalise(pd.DataFrame(), "South Dakota")

    trs = table.find("tbody").find_all("tr") if table.find("tbody") else table.find_all("tr")
    rows = []
    for tr in trs:
        tds = tr.find_all("td")
        if len(tds) < 4:
            continue
        if tds[0].get("bgcolor") == "#e7e7e7":
            continue
        if "tablecolumnheadings" in tds[0].get("class", []):
            continue

        link = tds[0].find("a")
        company = (link or tds[0]).get_text(separator=" ", strip=True)
        pdf_url = _sd_resolve_pdf(link.get("href", "") if link else "")

        if not company:
            continue
        rows.append({
            "company":            company,
            "city":               tds[1].get_text(separator=" ", strip=True),
            "notice_date":        tds[2].get_text(separator=" ", strip=True),
            "employees_affected": tds[3].get_text(separator=" ", strip=True),
            "closure_type":       "",
            "notes":              pdf_url,
        })

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    log.info(f"  South Dakota: {len(df)} rows")
    return _normalise(df, "South Dakota")


# ── Tennessee ─────────────────────────────────────────────────────────────────
# Selenium preferred (JS-rendered accordion tables) with requests fallback.
# URL: https://www.tn.gov/workforce/.../reports.html

_TN_URL = (
    "https://www.tn.gov/workforce/general-resources/major-publications0/"
    "major-publications-redirect/reports.html"
)


def _tn_norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _tn_is_warn_table(table) -> bool:
    headers = [_tn_norm(th.get_text(" ", strip=True)).lower() for th in table.find_all("th")]
    text = " | ".join(headers)
    return bool(headers) and all(k in text for k in ["date", "company", "county"])


def _tn_infer_year(table, default="2026") -> str:
    for node in table.find_all_previous(
        ["h1","h2","h3","h4","h5","h6","button","a","span","div"], limit=30
    ):
        m = re.search(r"\b(20\d{2})\b", _tn_norm(node.get_text(" ", strip=True)))
        if m:
            return m.group(1)
    return default


def _tn_parse_table(table, year: str) -> list[dict]:
    rows = []
    tbody = table.find("tbody")
    if not tbody:
        return rows
    for tr in tbody.find_all("tr", recursive=False):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 6:
            continue
        link    = tds[1].find("a", href=True)
        company = _tn_norm(tds[1].get_text(" ", strip=True))
        cells   = [_tn_norm(td.get_text(" ", strip=True)) for td in tds]
        if not company:
            continue
        from urllib.parse import urljoin
        rows.append({
            "company":            company,
            "city":               cells[2],   # County
            "notice_date":        cells[0],   # Date of Posting
            "layoff_date":        cells[4],   # Closure/Layoff Date
            "employees_affected": cells[3],   # Affected Workers
            "closure_type":       cells[5],   # Notice/Type
            "notes":              urljoin("https://www.tn.gov", link["href"]) if link else "",
        })
    return rows


def _tn_fetch_soup() -> BeautifulSoup:
    try:
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager

        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
        try:
            driver.get(_TN_URL)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            try:
                for btn in driver.find_elements(
                    By.CSS_SELECTOR, "button[data-tn-action='accordion:expandcollapse']"
                ):
                    aria = (btn.get_attribute("aria-label") or "").lower()
                    cls  = (btn.get_attribute("class") or "").lower()
                    if "expand" in aria or "icon-plus" in cls:
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(1.0)
                for btn in driver.find_elements(
                    By.CSS_SELECTOR, "button[aria-expanded='false'][aria-controls*='collapse']"
                ):
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.2)
            except Exception:
                pass
            WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "table tr"))
            )
            time.sleep(1.5)
            html = driver.page_source
        finally:
            driver.quit()
        return BeautifulSoup(html, "html.parser")

    except Exception as exc:
        log.warning(f"  Tennessee: Selenium unavailable ({exc}), falling back to requests")
        resp = requests.get(_TN_URL, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        return BeautifulSoup(resp.text, "html.parser")


def scrape_tennessee() -> pd.DataFrame:
    """
    Scrapes Tennessee WARN notices (all years) from JS-rendered accordion tables.
    URL: https://www.tn.gov/workforce/.../reports.html
    """
    log.info("Scraping Tennessee...")
    try:
        soup = _tn_fetch_soup()
    except Exception as exc:
        log.error(f"  Tennessee: failed to fetch page — {exc}")
        return _normalise(pd.DataFrame(), "Tennessee")

    rows: list[dict] = []
    seen: set[tuple] = set()
    table_count = 0

    for table in soup.find_all("table"):
        if not _tn_is_warn_table(table):
            continue
        year = _tn_infer_year(table, default="2026")
        for r in _tn_parse_table(table, year):
            key = (r["company"], r["notice_date"], r["city"])
            if key not in seen:
                seen.add(key)
                rows.append(r)
        table_count += 1

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    log.info(f"  Tennessee: {len(df)} rows from {table_count} table(s)")
    return _normalise(df, "Tennessee")


# ── California ────────────────────────────────────────────────────────────────
# Downloads the latest XLSX from EDD, reads "Detailed WARN Report" sheet.

def scrape_california() -> pd.DataFrame:
    """
    California WARN: downloads the latest XLSX from EDD and reads
    the 'Detailed WARN Report' sheet.
    URL: https://edd.ca.gov/en/jobs_and_training/Layoff_Services_WARN
    """
    import re as _re
    from io import BytesIO
    from urllib.parse import urljoin

    CA_PAGE        = "https://edd.ca.gov/en/jobs_and_training/Layoff_Services_WARN"
    CA_BASE        = "https://edd.ca.gov"
    TARGET_SHEET   = "Detailed WARN Report"

    log.info("Scraping California...")
    try:
        resp = requests.get(CA_PAGE, headers=HEADERS, timeout=45)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Preferred: explicit "Latest WARN report (XLSX)" anchor
        xlsx_url = None
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            text = " ".join(a.get_text(" ", strip=True).split()).lower()
            if "latest warn report" in text and "xlsx" in text:
                xlsx_url = urljoin(CA_BASE, href)
                break
        # Fallback: any .xlsx link mentioning warn
        if not xlsx_url:
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                text = " ".join(a.get_text(" ", strip=True).split()).lower()
                if ".xlsx" in href.lower() and "warn" in (href.lower() + " " + text):
                    xlsx_url = urljoin(CA_BASE, href)
                    break
        if not xlsx_url:
            log.error("  California: could not find XLSX link on page")
            return _normalise(pd.DataFrame(), "California")

        log.info(f"  California: downloading {xlsx_url}")
        r = requests.get(xlsx_url, headers=HEADERS, timeout=120)
        r.raise_for_status()
        content = r.content

        xls   = pd.ExcelFile(BytesIO(content))
        sheet = next(
            (s for s in xls.sheet_names if s.strip().lower() == TARGET_SHEET.lower()),
            None,
        )
        if not sheet:
            log.error(f"  California: sheet '{TARGET_SHEET}' not found; available={xls.sheet_names}")
            return _normalise(pd.DataFrame(), "California")

        raw = pd.read_excel(BytesIO(content), sheet_name=sheet, header=None, engine="openpyxl")

        # Detect header row (needs county/parish, notice date, company)
        required = {"county/parish", "notice date", "company"}
        hidx = None
        for i in range(min(len(raw), 30)):
            vals = {
                _re.sub(r"\s+", " ", str(c)).strip().lower()
                for c in raw.iloc[i].tolist()
                if c is not None
            }
            if required.issubset(vals):
                hidx = i
                break
        if hidx is None:
            log.error("  California: could not detect header row in sheet")
            return _normalise(pd.DataFrame(), "California")

        data = raw.iloc[hidx + 1:].copy()
        cols = []
        for j, c in enumerate(raw.iloc[hidx].tolist(), start=1):
            s = "" if c is None else _re.sub(r"\s+", " ", str(c)).strip()
            cols.append(s if s and not s.lower().startswith("unnamed") else f"extra_col_{j}")
        data.columns = cols
        data = data.dropna(how="all").reset_index(drop=True)

        if "Company" in data.columns:
            data = data[data["Company"].astype(str).str.strip() != ""].copy()
            data = data[
                ~data["Company"].astype(str).str.contains(
                    r"insert|delete|move|copy", case=False, regex=True
                )
            ].copy()

    except Exception as exc:
        log.error(f"  California scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "California")

    col_map = {
        "Company":          "company",
        "County/Parish":    "city",
        "Notice Date":      "notice_date",
        "Effective Date":   "layoff_date",
        "No. Of Employees": "employees_affected",
        "Layoff/Closure":   "closure_type",
        "Related Industry": "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)

    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  California: {len(data)} rows")
    return _normalise(data, "California")


# ── Colorado ──────────────────────────────────────────────────────────────────
# Exports Google Sheets links from the CDLE WARN page (2025 + 2026).

def scrape_colorado() -> pd.DataFrame:
    """
    Colorado WARN: exports Google Sheets links from the CDLE WARN page.
    URL: https://cdle.colorado.gov/employers/layoff-separations/layoff-warn-list
    Fetches sheets labelled 2025 and 2026 (including real-time links).
    """
    import re as _re
    from io import BytesIO

    CO_PAGE      = "https://cdle.colorado.gov/employers/layoff-separations/layoff-warn-list"
    _SHEET_RE    = _re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")
    _GID_RE      = _re.compile(r"gid=(\d+)")
    WANT_YEARS   = {"2025", "2026"}

    log.info("Scraping Colorado...")
    try:
        resp = requests.get(CO_PAGE, headers=HEADERS, timeout=45)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        seen         = set()
        sheet_links  = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if "docs.google.com/spreadsheets" not in href.lower():
                continue
            m = _SHEET_RE.search(href)
            if not m:
                continue
            sheet_id = m.group(1)
            gm  = _GID_RE.search(href)
            gid = gm.group(1) if gm else None
            key = (sheet_id, gid)
            if key in seen:
                continue
            seen.add(key)
            label       = " ".join(a.get_text(" ", strip=True).split())
            label_lower = label.lower()
            is_wanted   = (
                any(y in label_lower for y in WANT_YEARS)
                or "real-time" in label_lower
                or "real time" in label_lower
            )
            if is_wanted:
                export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
                if gid:
                    export_url += f"&gid={gid}"
                sheet_links.append((label, export_url))

        if not sheet_links:
            log.error("  Colorado: no 2025/2026 Google Sheets links found on page")
            return _normalise(pd.DataFrame(), "Colorado")

        frames = []
        for label, export_url in sheet_links:
            try:
                r = requests.get(export_url, headers=HEADERS, timeout=120, allow_redirects=True)
                r.raise_for_status()
                df = None
                for enc in ("utf-8-sig", "utf-8", "latin-1"):
                    try:
                        df = pd.read_csv(BytesIO(r.content), encoding=enc, low_memory=False)
                        break
                    except Exception:
                        continue
                if df is None:
                    log.warning(f"  Colorado: could not parse sheet '{label}'")
                    continue
                df.columns = [str(c).strip() for c in df.columns]
                df = df.dropna(how="all").reset_index(drop=True)
                frames.append(df)
                log.info(f"  Colorado [{label}]: {len(df)} rows")
            except Exception as exc:
                log.warning(f"  Colorado: sheet '{label}' failed — {exc}")

        if not frames:
            log.error("  Colorado: no data downloaded from any sheet")
            return _normalise(pd.DataFrame(), "Colorado")

        combined = pd.concat(frames, ignore_index=True)

    except Exception as exc:
        log.error(f"  Colorado scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Colorado")

    col_map = {
        "Company Name":        "company",
        "Company":             "company",
        "Employer":            "company",
        "Employer Name":       "company",
        "City":                "city",
        "City/Town":           "city",
        "Notice Date":         "notice_date",
        "WARN Date":           "notice_date",
        "Date of Notice":      "notice_date",
        "Layoff Date":         "layoff_date",
        "Effective Date":      "layoff_date",
        "Date of Layoff":      "layoff_date",
        "No. of Employees":    "employees_affected",
        "Number of Employees": "employees_affected",
        "# Employees":         "employees_affected",
        "Employees Affected":  "employees_affected",
        "Employees":           "employees_affected",
        "Type":                "closure_type",
        "Layoff/Closure":      "closure_type",
        "Action Type":         "closure_type",
        "Layoff Type":         "closure_type",
        "County":              "notes",
        "Comments":            "notes",
        "Notes":               "notes",
    }
    combined.rename(columns={k: v for k, v in col_map.items() if k in combined.columns}, inplace=True)

    for dcol in ("notice_date", "layoff_date"):
        if dcol in combined.columns:
            parsed = pd.to_datetime(combined[dcol], errors="coerce")
            combined[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(combined[dcol].astype(str).str.strip())

    log.info(f"  Colorado: {len(combined)} total rows")
    return _normalise(combined, "Colorado")


# ── Connecticut ───────────────────────────────────────────────────────────────
_CT_GSHEETS_ID = "1OzqiDXufjLdNAANwnTjZ38TsVvw4Q8ffXZKTO6ZF-do"

def scrape_connecticut() -> pd.DataFrame:
    return _scrape_layoffdata_gsheets("Connecticut", _CT_GSHEETS_ID)


# ── Delaware ──────────────────────────────────────────────────────────────────
# JobLink search — GET with pagination (Rails Ransack params).

_DE_RESULTS_URL = (
    "https://joblink.delaware.gov/search/warn_lookups"
    "?commit=Search&q%5Bnotice_on_gteq%5D=2025-01-01&q%5Bnotice_eq%5D="
)

async def _de_scrape_async():
    """Playwright: load DE JobLink results directly, paginate, return row dicts."""
    all_headers, all_rows = [], []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()

        # Navigate directly to filtered results (avoids CSRF/form issues)
        await page.goto(_DE_RESULTS_URL, wait_until="networkidle", timeout=60000)
        try:
            await page.wait_for_selector("table", timeout=20000)
        except Exception:
            pass

        for _ in range(300):                  # up to 300 pages
            html  = await page.content()
            soup  = BeautifulSoup(html, "html.parser")

            table = next(
                (t for t in soup.find_all("table")
                 if t.get("class") and "sortable" in t.get("class")),
                soup.find("table"),
            )
            if not table:
                break

            if not all_headers:
                thead = table.find("thead")
                if thead and thead.find("tr"):
                    for th in thead.find("tr").find_all(["th", "td"]):
                        a   = th.find("a", class_=lambda c: c and "sort_link" in str(c))
                        txt = " ".join((a or th).get_text(" ", strip=True).split())
                        all_headers.append(txt or f"col_{len(all_headers)+1}")

            tbody = table.find("tbody")
            if tbody:
                for tr in tbody.find_all("tr"):
                    cells = tr.find_all("td", recursive=False)
                    if not cells:
                        continue
                    texts = []
                    for i, td in enumerate(cells):
                        a = td.find("a", href=True) if i == 0 else None
                        texts.append(" ".join((a or td).get_text(" ", strip=True).split()))
                    while len(texts) < len(all_headers):
                        texts.append("")
                    row = {all_headers[j]: texts[j] for j in range(len(all_headers))}
                    if any(row.values()):
                        all_rows.append(row)

            if not soup.select_one("a.next_page[href]"):
                break
            try:
                await page.click("a.next_page")
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                break

        await browser.close()
    return all_rows


def scrape_delaware() -> pd.DataFrame:
    """
    Delaware WARN: JobLink search via Playwright (bypasses bot protection).
    URL: https://joblink.delaware.gov/search/warn_lookups
    """
    log.info("Scraping Delaware...")
    try:
        all_rows = _run_async(_de_scrape_async())
        if not all_rows:
            log.error("  Delaware: no rows found")
            return _normalise(pd.DataFrame(), "Delaware")
        data = pd.DataFrame(all_rows).dropna(how="all").reset_index(drop=True)
        log.info(f"  Delaware raw columns: {data.columns.tolist()}")
    except Exception as exc:
        log.error(f"  Delaware scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Delaware")

    col_map = {
        "Employer":            "company",
        "Employer Name":       "company",
        "Company":             "company",
        "Company Name":        "company",
        "City":                "city",
        "Location":            "city",
        "Notice Date":         "notice_date",
        "Date":                "notice_date",
        "Effective Date":      "layoff_date",
        "Layoff Date":         "layoff_date",
        "Workers Affected":    "employees_affected",
        "Employees Affected":  "employees_affected",
        "Number of Employees": "employees_affected",
        "Employees":           "employees_affected",
        "WARN Type":           "closure_type",
        "Type":                "closure_type",
        "Notice Type":         "closure_type",
        "Layoff/Closure":      "closure_type",
        "LWIB Area":           "notes",
        "County":              "notes",
        "Notes":               "notes",
        "Comments":            "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)

    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Delaware: {len(data)} rows")
    return _normalise(data, "Delaware")


# ── Florida ───────────────────────────────────────────────────────────────────
# REACT WARN reports — POST export, returns HTML spreadsheet parsed by pandas.

_FL_REPORTS_URL = "https://reactwarn.floridajobs.org/warnlist/reports"

async def _fl_scrape_async() -> str:
    """Playwright: fill FL REACT export form and return the resulting HTML."""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page    = await browser.new_page()
        await page.goto(_FL_REPORTS_URL, wait_until="networkidle", timeout=60000)

        # Set the start date
        for sel in ["input[name='StateNotificationStartDate']", "#StateNotificationStartDate"]:
            try:
                await page.fill(sel, "2025-01-01")
                break
            except Exception:
                pass

        # Switch hidden appForm field to "export" and submit
        # Use expect_navigation to avoid race condition where page.content()
        # is called while the page is still navigating after submit()
        try:
            async with page.expect_navigation(wait_until="networkidle", timeout=120000):
                await page.evaluate("""
                    var el = document.querySelector('input[name="appForm"]');
                    if (el) el.value = 'export';
                    document.querySelector('form').submit();
                """)
        except Exception:
            # Fallback: click the Export button directly
            try:
                async with page.expect_navigation(wait_until="networkidle", timeout=120000):
                    await page.click("input[value='Export'], button:text('Export')")
            except Exception:
                await page.wait_for_load_state("networkidle", timeout=120000)

        html = await page.content()
        await browser.close()
        return html


def scrape_florida() -> pd.DataFrame:
    """
    Florida WARN: REACT export form via Playwright (bypasses bot protection).
    URL: https://reactwarn.floridajobs.org/warnlist/reports
    """
    from io import StringIO as _SIO

    log.info("Scraping Florida...")
    try:
        html = _run_async(_fl_scrape_async())
        dfs  = pd.read_html(_SIO(html), header=0, flavor="lxml")
        if not dfs:
            log.error("  Florida: no table in export response")
            return _normalise(pd.DataFrame(), "Florida")
        data = dfs[0].copy()
        data.columns = [str(c).strip() for c in data.columns]
        data = data.dropna(how="all").reset_index(drop=True)
        for c in data.columns:
            if data[c].dtype == object:
                data[c] = data[c].astype(str).str.strip().replace({"nan": ""})
        log.info(f"  Florida raw columns: {data.columns.tolist()}")
    except Exception as exc:
        log.error(f"  Florida scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Florida")

    col_map = {
        "Company":                    "company",
        "Employer":                   "company",
        "Employer Name":              "company",
        "Company Name":               "company",
        "City":                       "city",
        "Location":                   "city",
        "State Notification Date":    "notice_date",
        "Notice Date":                "notice_date",
        "Layoff Date":                "layoff_date",
        "Effective Date":             "layoff_date",
        "Number of Employees":        "employees_affected",
        "Employees":                  "employees_affected",
        "Workers Affected":           "employees_affected",
        "Employees Affected":         "employees_affected",
        "Type":                       "closure_type",
        "Layoff/Closure":             "closure_type",
        "Notice Type":                "closure_type",
        "County":                     "notes",
        "LWDB":                       "notes",
        "Notes":                      "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)

    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Florida: {len(data)} rows")
    return _normalise(data, "Florida")


# ── Georgia ───────────────────────────────────────────────────────────────────
# TCSG GravityView + DataTables — requires Selenium (JS-rendered table).

_GA_PAGE_URL = "https://www.tcsg.edu/warn-public-view/"

def _ga_build_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,1000")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    svc = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=svc, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return driver


def scrape_georgia() -> pd.DataFrame:
    """
    Georgia WARN: TCSG GravityView + DataTables (Selenium).
    URL: https://www.tcsg.edu/warn-public-view/
    """
    import re as _re

    log.info("Scraping Georgia...")
    try:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import Select, WebDriverWait

        driver = _ga_build_driver()
        rows_out = []
        try:
            driver.set_page_load_timeout(120)
            driver.get(_GA_PAGE_URL)
            time.sleep(2)

            wait = WebDriverWait(driver, 90)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.gv-datatables")))
            time.sleep(2)

            # Select "All" entries so everything is loaded
            try:
                length_sel = driver.find_element(By.CSS_SELECTOR, ".dataTables_length select")
                Select(length_sel).select_by_visible_text("All")
            except Exception:
                pass

            # Wait for all rows to load
            deadline = time.time() + 240
            while time.time() < deadline:
                time.sleep(1)
                infos = driver.find_elements(By.CSS_SELECTOR, ".dataTables_info")
                if not infos:
                    continue
                info_text = infos[0].text.strip()
                m = _re.search(r"Showing\s+(\d+)\s+to\s+(\d+)\s+of\s+(\d+)\s+entries",
                               info_text.replace(",", ""), _re.I)
                if m:
                    end, total = int(m.group(2)), int(m.group(3))
                    if total > 0 and end >= total:
                        break
                if "No entries" in info_text:
                    break

            time.sleep(1)

            # Read headers
            header_labels = []
            for th in driver.find_elements(By.CSS_SELECTOR, "table.gv-datatables thead tr th"):
                txt = " ".join(th.text.split()).strip()
                if txt and txt.lower() != "entry id":
                    header_labels.append(txt)
            if not header_labels:
                header_labels = ["GA WARN ID", "Company Name", "Submitted Date",
                                  "Total Number of Affected Employees"]

            for tr in driver.find_elements(By.CSS_SELECTOR, "table.gv-datatables tbody tr"):
                tds = tr.find_elements(By.TAG_NAME, "td")
                if not tds:
                    continue
                try:
                    link_el = tds[0].find_element(
                        By.CSS_SELECTOR, 'a[href*="/warn-public-view/entry/"]'
                    )
                except Exception:
                    continue
                href    = (link_el.get_attribute("href") or "").strip()
                warn_id = " ".join(link_el.text.split()).strip()
                if not warn_id or warn_id.lower() in ("ga warn id", "id"):
                    continue
                rest = [" ".join(td.text.split()).strip() for td in tds[1:]]
                while len(rest) < len(header_labels) - 1:
                    rest.append("")
                rest = rest[:len(header_labels) - 1]
                row = {header_labels[0]: warn_id}
                for i, lab in enumerate(header_labels[1:]):
                    row[lab] = rest[i] if i < len(rest) else ""
                row["source_entry_url"] = href
                rows_out.append(row)

        finally:
            try:
                driver.quit()
            except Exception:
                pass

        if not rows_out:
            log.error("  Georgia: no rows scraped")
            return _normalise(pd.DataFrame(), "Georgia")

        data = pd.DataFrame(rows_out)
        data = data.dropna(how="all").reset_index(drop=True)
        for c in data.columns:
            if data[c].dtype == object:
                data[c] = data[c].astype(str).str.strip().replace({"nan": ""})
        log.info(f"  Georgia raw columns: {data.columns.tolist()}")

    except Exception as exc:
        log.error(f"  Georgia scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Georgia")

    col_map = {
        "Company Name":                        "company",
        "Company":                             "company",
        "Employer":                            "company",
        "City":                                "city",
        "Location":                            "city",
        "Submitted Date":                      "notice_date",
        "Notice Date":                         "notice_date",
        "WARN Effective Date":                 "layoff_date",
        "Effective Date":                      "layoff_date",
        "Layoff Date":                         "layoff_date",
        "Total Number of Affected Employees":  "employees_affected",
        "Number of Employees":                 "employees_affected",
        "Employees Affected":                  "employees_affected",
        "Employees":                           "employees_affected",
        "Type":                                "closure_type",
        "Notice Type":                         "closure_type",
        "Layoff/Closure":                      "closure_type",
        "GA WARN ID":                          "notes",
        "County":                              "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)

    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Georgia: {len(data)} rows")
    return _normalise(data, "Georgia")


# ── Hawaii ────────────────────────────────────────────────────────────────────
# labor.hawaii.gov — index page → per-year notice pages → <p> date – company entries

_HI_INDEX_URL = "https://labor.hawaii.gov/wdc/real-time-warn-updates/"
_HI_YEAR_RE   = re.compile(r"/wdc/(\d{4})-warn-notices/?$", re.IGNORECASE)
_HI_DATE_RE   = re.compile(
    r"^\s*(?:[A-Za-z]+,\s*)?(?P<month>January|February|March|April|May|June|July|"
    r"August|September|October|November|December)\s+(?P<day>\d{1,2}),\s*(?P<year>\d{4})\s*",
    re.IGNORECASE,
)
_HI_SEP_RE = re.compile(r"^\s*[–—\-:]\s*")
_HI_MONTH = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
}

def _hi_fetch(url: str) -> str:
    hdrs = {"User-Agent": "Mozilla/5.0","Accept": "text/html"}
    r = requests.get(url, headers=hdrs, timeout=60)
    r.raise_for_status()
    r.encoding = r.apparent_encoding or r.encoding
    return r.text

def _hi_iter_segments(p):
    parts, link = [], ""
    def flush():
        nonlocal parts, link
        text = " ".join("".join(parts).replace("\xa0", " ").split()).strip()
        out  = (text, link) if text else None
        parts, link = [], ""
        return out
    for child in p.children:
        name = getattr(child, "name", None)
        if name == "br":
            seg = flush()
            if seg: yield seg
        elif name == "a":
            href = (child.get("href") or "").strip()
            parts.append(child.get_text(" ", strip=False))
            if href and not link: link = href
        elif name is None:
            parts.append(str(child))
        else:
            parts.append(child.get_text(" ", strip=False))
    seg = flush()
    if seg: yield seg

def scrape_hawaii() -> pd.DataFrame:
    """Hawaii DLIR/WDC — index → per-year pages → <p> notice list."""
    from urllib.parse import urljoin
    log.info("Scraping Hawaii...")
    try:
        idx_html = _hi_fetch(_HI_INDEX_URL)
    except Exception as exc:
        log.error(f"  Hawaii: index fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Hawaii")

    soup      = BeautifulSoup(idx_html, "html.parser")
    year_urls = {}
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        abs_url = urljoin(_HI_INDEX_URL, href)
        m = _HI_YEAR_RE.search(abs_url)
        if m:
            year_urls.setdefault(m.group(1), abs_url)

    now = datetime.now()
    target_years = {str(now.year - 1), str(now.year), str(now.year + 1)}
    rows = []
    for year, url in sorted(year_urls.items()):
        if year not in target_years:
            continue
        try:
            html = _hi_fetch(url)
        except Exception as exc:
            log.warning(f"  Hawaii: skip {year}: {exc}")
            continue
        page_soup = BeautifulSoup(html, "html.parser")
        container = page_soup.select_one(
            "div.elementor-widget-text-editor .elementor-widget-container"
        ) or page_soup.select_one("main .entry-content") or page_soup.find("main")
        if not container:
            continue
        for p in container.find_all("p"):
            for text, _link in _hi_iter_segments(p):
                m2 = _HI_DATE_RE.match(text)
                if not m2:
                    continue
                mon  = m2.group("month").lower()
                day  = int(m2.group("day"))
                yr   = int(m2.group("year"))
                iso  = f"{yr:04d}-{_HI_MONTH[mon]:02d}-{day:02d}"
                rest = _HI_SEP_RE.sub("", text[m2.end():]).strip()
                rows.append({"notice_date": iso, "company": rest})
        log.info(f"  Hawaii [{year}]: {len(rows)} cumulative rows")

    if not rows:
        log.error("  Hawaii: no rows parsed")
        return _normalise(pd.DataFrame(), "Hawaii")

    data = pd.DataFrame(rows)
    data["notice_date"] = pd.to_datetime(data["notice_date"], errors="coerce").dt.strftime("%m/%d/%Y")
    log.info(f"  Hawaii: {len(data)} rows")
    return _normalise(data, "Hawaii")


# ── Idaho ─────────────────────────────────────────────────────────────────────
# labor.idaho.gov — landing page → discover PDF → PyMuPDF table extraction

_ID_PAGE_URL = "https://www.labor.idaho.gov/businesses/layoff-assistance/"

def scrape_idaho() -> pd.DataFrame:
    """Idaho DOL — discover & parse the rolling WARN PDF with PyMuPDF."""
    from io import BytesIO as _BIO
    from urllib.parse import urljoin
    log.info("Scraping Idaho...")
    try:
        import fitz  # PyMuPDF
    except ImportError:
        log.error("  Idaho: pymupdf not installed — pip install pymupdf")
        return _normalise(pd.DataFrame(), "Idaho")

    hdrs = {"User-Agent": "Mozilla/5.0", "Accept": "text/html"}
    try:
        r = requests.get(_ID_PAGE_URL, headers=hdrs, timeout=60)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        log.error(f"  Idaho: landing page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Idaho")

    # Discover the PDF link — score candidates so the Idaho state PDF wins over DOL links
    def _id_score(href: str, text: str) -> int:
        hl, tl = href.lower(), text.lower()
        if not hl.endswith(".pdf"):
            return -1
        score = 0
        if "idaho-warn" in hl:               score += 5
        if "warn" in hl:                     score += 1
        if "labor.idaho.gov" in hl or hl.startswith("/wp-content/"):
            score += 2
        if "idaho warn" in tl or "warn notice" in tl:
            score += 3
        if "/uploads/" in hl:                score += 1
        return score

    best_score, pdf_url = 0, None
    for a in soup.find_all("a", href=True):
        href  = (a.get("href") or "").strip()
        text  = " ".join(a.get_text(" ", strip=True).split())
        abs_url = urljoin(_ID_PAGE_URL, href)
        s = _id_score(abs_url, text)
        if s > best_score:
            best_score, pdf_url = s, abs_url
    if not pdf_url:
        log.error("  Idaho: no WARN PDF link found on landing page")
        return _normalise(pd.DataFrame(), "Idaho")

    log.info(f"  Idaho: downloading {pdf_url}")
    try:
        r2 = requests.get(pdf_url, headers={
            **hdrs,
            "Accept": "application/pdf,*/*;q=0.9",
            "Referer": _ID_PAGE_URL,
        }, timeout=180)
        r2.raise_for_status()
        pdf_bytes = r2.content
    except Exception as exc:
        log.error(f"  Idaho: PDF download failed: {exc}")
        return _normalise(pd.DataFrame(), "Idaho")

    # Parse PDF tables with fitz
    EXPECTED = ["Date of Letter","Updates","Company","Address","City","State","Zip",
                "No. of Employees Affected","Effective or Commencing Date"]
    canonical, table_rows = [], []
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for pi in range(doc.page_count):
            page = doc[pi]
            for tbl in page.find_tables().tables:
                data = tbl.extract()
                if not data:
                    continue
                hdr = [" ".join(str(c).split()) for c in data[0]]
                if not canonical:
                    canonical = hdr
                start = 1 if hdr == canonical else 0
                n = len(canonical) if canonical else len(hdr)
                for raw in data[start:]:
                    cells = [" ".join(str(c).split()) if c else "" for c in raw]
                    if len(cells) < n: cells += [""] * (n - len(cells))
                    elif len(cells) > n: cells = cells[:n]
                    if any(cells):
                        table_rows.append(cells)
        doc.close()
    except Exception as exc:
        log.error(f"  Idaho: PDF parse failed: {exc}")
        return _normalise(pd.DataFrame(), "Idaho")

    if not table_rows:
        log.error("  Idaho: no rows extracted from PDF")
        return _normalise(pd.DataFrame(), "Idaho")

    hdr = canonical if canonical else EXPECTED
    data = pd.DataFrame(table_rows, columns=hdr[:len(table_rows[0])])
    data.columns = [str(c).strip() for c in data.columns]
    data = data.dropna(how="all").reset_index(drop=True)

    # Filter 2025+
    date_col = data.columns[0] if data.columns[0] in data.columns else "Date of Letter"
    if date_col in data.columns:
        years_mask = pd.to_datetime(data[date_col], errors="coerce").dt.year >= 2025
        data = data[years_mask | years_mask.isna()].reset_index(drop=True)

    col_map = {
        "Company":                      "company",
        "City":                         "city",
        "Date of Letter":               "notice_date",
        "Effective or Commencing Date": "layoff_date",
        "No. of Employees Affected":    "employees_affected",
        "Updates":                      "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Idaho: {len(data)} rows")
    return _normalise(data, "Idaho")


# ── Indiana ───────────────────────────────────────────────────────────────────
# in.gov — static HTML table (DataTables JS — data pre-loaded in HTML)

_IN_PAGE_URL  = "https://www.in.gov/dwd/warn-notices/current-warn-notices/"
_IN_TABLE_ID  = "table33066"

def scrape_indiana() -> pd.DataFrame:
    """Indiana DWD — parse the pre-loaded DataTables HTML table."""
    log.info("Scraping Indiana...")
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        r = requests.get(_IN_PAGE_URL, headers=hdrs, timeout=60)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or r.encoding
        html = r.text
    except Exception as exc:
        log.error(f"  Indiana: fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Indiana")

    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table", id=_IN_TABLE_ID)
    if table is None:
        for t in soup.find_all("table"):
            thead = t.find("thead")
            if not thead:
                continue
            cells = " ".join(c.get_text() for c in thead.find_all(["th","td"])).lower()
            if "company" in cells and "notice date" in cells and "naics" in cells:
                table = t
                break
    if table is None:
        log.error("  Indiana: WARN table not found")
        return _normalise(pd.DataFrame(), "Indiana")

    # Parse headers
    thead = table.find("thead")
    raw_hdrs = [" ".join(c.get_text(" ", strip=True).split())
                for c in (thead.find_all(["th","td"]) if thead else [])]
    if not raw_hdrs:
        raw_hdrs = ["Company","City","Affected Workers","Notice Date",
                    "LO/CL Date","NAICS","Description of Work/Industry","Notice Type","Notice PDF"]
    # Last column is the PDF link column
    rows_out = []
    tbody = table.find("tbody")
    for tr in (tbody.find_all("tr") if tbody else table.find_all("tr")[1:]):
        cells = tr.find_all(["td","th"], recursive=False)
        if not cells:
            continue
        vals = []
        for i, td in enumerate(cells):
            if i == len(raw_hdrs) - 1:
                a = td.find("a", href=True)
                vals.append((a.get("href") or "").strip() if a else "")
            else:
                vals.append(" ".join(td.get_text(" ", strip=True).split()))
        while len(vals) < len(raw_hdrs): vals.append("")
        row = {raw_hdrs[j]: vals[j] for j in range(len(raw_hdrs))}
        if any(v for k, v in row.items() if k != raw_hdrs[-1]):
            rows_out.append(row)

    if not rows_out:
        log.error("  Indiana: no rows parsed")
        return _normalise(pd.DataFrame(), "Indiana")

    data = pd.DataFrame(rows_out)
    data.columns = [str(c).strip() for c in data.columns]

    col_map = {
        "Company":                      "company",
        "City":                         "city",
        "Affected Workers":             "employees_affected",
        "Notice Date":                  "notice_date",
        "LO/CL Date":                   "layoff_date",
        "Notice Type":                  "closure_type",
        "Description of Work/Industry": "notes",
        "NAICS":                        "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Indiana: {len(data)} rows")
    return _normalise(data, "Indiana")


# ── Iowa ──────────────────────────────────────────────────────────────────────
# workforce.iowa.gov — landing page → discover XLSX → openpyxl multi-sheet parse

_IA_PAGE_URL = "https://workforce.iowa.gov/employers/resources/warn"
_IA_YEAR_RE  = re.compile(r"^(20\d{2})$")

def scrape_iowa() -> pd.DataFrame:
    """Iowa Workforce Development — download & parse the multi-year WARN XLSX."""
    from io import BytesIO as _BIO
    from urllib.parse import urljoin
    log.info("Scraping Iowa...")
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        r = requests.get(_IA_PAGE_URL, headers=hdrs, timeout=60)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        log.error(f"  Iowa: landing page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Iowa")

    # Discover the XLSX download link
    xlsx_url = None
    best_score = 0
    for a in soup.find_all("a", href=True):
        href  = (a.get("href") or "").strip()
        text  = " ".join(a.get_text(" ", strip=True).split()).lower()
        aria  = (a.get("aria-label") or "").lower()
        combo = f"{text} {aria}"
        score = 0
        if "warn log" in combo or ("warn" in combo and ("excel" in combo or "log" in combo)):
            score += 5
        if href.lower().endswith((".xlsx", ".xls")):
            score += 3
        if "/media/" in href and "/download" in href:
            score += 4
        if score > best_score:
            best_score = score
            xlsx_url = urljoin(_IA_PAGE_URL, href)
    if not xlsx_url:
        log.error("  Iowa: no WARN Log XLSX link found on landing page")
        return _normalise(pd.DataFrame(), "Iowa")

    log.info(f"  Iowa: downloading {xlsx_url}")
    try:
        r2 = requests.get(xlsx_url, headers={**hdrs, "Referer": _IA_PAGE_URL}, timeout=180)
        r2.raise_for_status()
        xlsx_bytes = r2.content
    except Exception as exc:
        log.error(f"  Iowa: XLSX download failed: {exc}")
        return _normalise(pd.DataFrame(), "Iowa")

    # Parse multi-sheet workbook with openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_BIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as exc:
        log.error(f"  Iowa: workbook parse failed: {exc}")
        return _normalise(pd.DataFrame(), "Iowa")

    now = datetime.now()
    target_years = {str(now.year - 1), str(now.year), str(now.year + 1)}
    all_rows = []
    try:
        for name in wb.sheetnames:
            m = _IA_YEAR_RE.match(name.strip())
            if not m or m.group(1) not in target_years:
                continue
            year = m.group(1)
            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            row1 = next(rows_iter, None)
            row2 = next(rows_iter, None)
            if row2 is None:
                continue
            # Row 1 is title; row 2 is header — unless row1 already has column names
            flat1 = [str(c).strip().lower() for c in (row1 or []) if c]
            if "company" in flat1 and any("date" in x for x in flat1):
                hdr_row, extra = row1, [row2]
            else:
                hdr_row, extra = row2, []
            headers = [str(c).strip() if c is not None else "" for c in hdr_row]
            keep_wanted = ["Company","City","County","Notice Type","Emp #",
                           "Notice Date","Layoff Date","Local Workforce Area","Industry"]
            keep_idx = [i for i,h in enumerate(headers) if h in keep_wanted]
            keep_names = [headers[i] for i in keep_idx]
            if "Company" not in keep_names or "Notice Date" not in keep_names:
                continue
            for raw in (extra + list(rows_iter)):
                if raw is None or not any(raw):
                    continue
                d = {}
                has_val = False
                for name2, idx in zip(keep_names, keep_idx):
                    v = raw[idx] if idx < len(raw) else None
                    if v is None or (isinstance(v, str) and not v.strip()):
                        d[name2] = ""
                    else:
                        sv = v.strftime("%m/%d/%Y") if hasattr(v, "strftime") else str(v).strip()
                        d[name2] = sv
                        has_val = True
                if has_val:
                    d["_year"] = year
                    all_rows.append(d)
        wb.close()
    except Exception as exc:
        log.error(f"  Iowa: workbook read error: {exc}")
        return _normalise(pd.DataFrame(), "Iowa")

    if not all_rows:
        log.error("  Iowa: no rows found in target year sheets")
        return _normalise(pd.DataFrame(), "Iowa")

    data = pd.DataFrame(all_rows).drop(columns=["_year"], errors="ignore")
    col_map = {
        "Company":               "company",
        "City":                  "city",
        "Emp #":                 "employees_affected",
        "Notice Date":           "notice_date",
        "Layoff Date":           "layoff_date",
        "Notice Type":           "closure_type",
        "County":                "notes",
        "Local Workforce Area":  "notes",
        "Industry":              "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())

    log.info(f"  Iowa: {len(data)} rows")
    return _normalise(data, "Iowa")


# ── Kansas ────────────────────────────────────────────────────────────────────
# kansasworks.com — paginated search results (requests + BeautifulSoup)

_KS_SEARCH_URL = "https://www.kansasworks.com/search/warn_lookups"
_KS_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

def scrape_kansas() -> pd.DataFrame:
    """Kansas WorkWORKS — paginated WARN search results, 2025-onward."""
    log.info("Scraping Kansas...")
    session = requests.Session()
    session.headers.update(_KS_HEADERS)
    try:
        session.get("https://www.kansasworks.com/search/warn_lookups/new", timeout=30)
    except Exception:
        pass

    rows_out = []
    page = 1
    try:
        while page <= 50:
            params = {
                "commit": "Search",
                "page": str(page),
                "q[employer_name_cont]": "",
                "q[main_contact_contact_info_addresses_full_location_city_matches]": "",
                "q[notice_eq]": "true",
                "q[notice_on_gteq]": "",
                "q[notice_on_lteq]": "",
                "q[service_delivery_area_id_eq]": "",
                "q[zipcode_code_start]": "",
            }
            r = session.get(_KS_SEARCH_URL, params=params, timeout=30)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            table = soup.find("table")
            if not table:
                break
            # Columns: Employer | City | ZIP | LWIB Area | Notice Date | WARN Type
            for tr in table.find_all("tr")[1:]:
                cells = tr.find_all("td")
                if len(cells) < 5:
                    continue
                def _txt(i):
                    return " ".join(cells[i].get_text(" ", strip=True).split()) if len(cells) > i else ""
                rows_out.append({
                    "company":      _txt(0),
                    "city":         _txt(1),
                    "notice_date":  _txt(4),
                    "closure_type": _txt(5),
                    "notes":        _txt(3),
                })
            # Stop when there is no "next page" link
            if not soup.select_one("a.next_page[href], a[rel~='next']"):
                break
            page += 1
            time.sleep(0.3)
    except Exception as exc:
        log.error(f"  Kansas: scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Kansas")

    if not rows_out:
        log.error("  Kansas: no rows found")
        return _normalise(pd.DataFrame(), "Kansas")

    data = pd.DataFrame(rows_out)
    for dcol in ("notice_date",):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())
    log.info(f"  Kansas: {len(data)} rows")
    return _normalise(data, "Kansas")


# ── Kentucky ──────────────────────────────────────────────────────────────────
# layoffdata.com Google Sheet → CSV export (no browser needed)

_KY_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "10y2of44J4Q9wCk78K29yiWkLUPbsC0wEaK1_bZVYiE0/export?format=csv"
)

def scrape_kentucky() -> pd.DataFrame:
    """Kentucky WARN — Google Sheets CSV export via layoffdata.com."""
    from io import StringIO as _SIO
    log.info("Scraping Kentucky...")
    hdrs = {"User-Agent": "Mozilla/5.0", "Accept": "*/*"}
    try:
        r = requests.get(_KY_SHEET_URL, headers=hdrs, timeout=60)
        r.raise_for_status()
        text = r.content.decode("utf-8-sig", errors="replace")
        data = pd.read_csv(_SIO(text), dtype=str).fillna("")
    except Exception as exc:
        log.error(f"  Kentucky: download/parse failed: {exc}")
        return _normalise(pd.DataFrame(), "Kentucky")

    data.columns = [str(c).strip() for c in data.columns]
    col_map = {
        "Employer":               "company",
        "Company":                "company",
        "Employer Name":          "company",
        "City":                   "city",
        "Notice Date":            "notice_date",
        "Warn Date":              "notice_date",
        "Date Received":          "notice_date",
        "Layoff Date":            "layoff_date",
        "Effective Date":         "layoff_date",
        "Warn Type":              "closure_type",
        "Type":                   "closure_type",
        "Employees":              "employees_affected",
        "Number of Employees":    "employees_affected",
        "Workers Affected":       "employees_affected",
        "County":                 "notes",
        "Region":                 "notes",
        "Industry":               "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())
    log.info(f"  Kentucky: {len(data)} rows (pre-filter)")
    return _normalise(data, "Kentucky")


# ── Louisiana ─────────────────────────────────────────────────────────────────
# layoffdata.com Google Sheet → CSV export (no browser needed)

_LA_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1vpc07656NGWlAt9UkeltaHSZSxxkJonSHH-5rTuNpDM/export?format=csv"
)

def scrape_louisiana() -> pd.DataFrame:
    """Louisiana WARN — Google Sheets CSV export via layoffdata.com."""
    from io import StringIO as _SIO
    log.info("Scraping Louisiana...")
    hdrs = {"User-Agent": "Mozilla/5.0", "Accept": "*/*"}
    try:
        r = requests.get(_LA_SHEET_URL, headers=hdrs, timeout=60)
        r.raise_for_status()
        text = r.content.decode("utf-8-sig", errors="replace")
        data = pd.read_csv(_SIO(text), dtype=str).fillna("")
    except Exception as exc:
        log.error(f"  Louisiana: download/parse failed: {exc}")
        return _normalise(pd.DataFrame(), "Louisiana")

    data.columns = [str(c).strip() for c in data.columns]
    col_map = {
        "Employer":               "company",
        "Company":                "company",
        "Employer Name":          "company",
        "City":                   "city",
        "Notice Date":            "notice_date",
        "Warn Date":              "notice_date",
        "Date Received":          "notice_date",
        "Layoff Date":            "layoff_date",
        "Effective Date":         "layoff_date",
        "Warn Type":              "closure_type",
        "Type":                   "closure_type",
        "Employees":              "employees_affected",
        "Number of Employees":    "employees_affected",
        "Workers Affected":       "employees_affected",
        "Parish":                 "notes",
        "County":                 "notes",
        "Region":                 "notes",
        "Industry":               "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date", "layoff_date"):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())
    log.info(f"  Louisiana: {len(data)} rows (pre-filter)")
    return _normalise(data, "Louisiana")


# ── Maine ─────────────────────────────────────────────────────────────────────
# joblink.maine.gov — Rails/Ransack JobLink platform, paginated results table

_ME_SEARCH_NEW = "https://joblink.maine.gov/search/warn_lookups/new"
_ME_SEARCH_URL = "https://joblink.maine.gov/search/warn_lookups"

def scrape_maine() -> pd.DataFrame:
    """Maine WorkSource JobLink — paginated WARN search, listing table only."""
    from urllib.parse import urlencode
    log.info("Scraping Maine...")
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    session = requests.Session()
    session.headers.update(hdrs)
    try:
        session.get(_ME_SEARCH_NEW, timeout=30)  # get session cookie
    except Exception:
        pass

    all_rows = []
    page = 1
    try:
        while page <= 50:
            params = {
                "commit": "Search",
                "page": str(page),
                "q[employer_name_cont]": "",
                "q[main_contact_contact_info_addresses_full_location_city_matches]": "",
                "q[notice_eq]": "true",
                "q[notice_on_gteq]": "",
                "q[notice_on_lteq]": "",
                "q[service_delivery_area_id_eq]": "",
                "q[zipcode_code_start]": "",
            }
            r = session.get(_ME_SEARCH_URL, params=params, timeout=30)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            table = soup.find("table")
            if not table:
                break
            thead = table.find("thead")
            if thead and thead.find("tr"):
                headers = [" ".join(th.get_text(" ", strip=True).split())
                           for th in thead.find("tr").find_all(["th","td"])]
            else:
                headers = ["Employer","City","ZIP","LWIB Area","Notice Date","WARN Type"]

            tbody = table.find("tbody")
            page_rows = 0
            for tr in (tbody.find_all("tr") if tbody else []):
                cells = tr.find_all(["td","th"], recursive=False)
                if not cells:
                    continue
                vals = [" ".join(td.get_text(" ", strip=True).split()) for td in cells]
                while len(vals) < len(headers): vals.append("")
                row = {headers[j]: vals[j] for j in range(len(headers))}
                if any(row.values()):
                    all_rows.append(row)
                    page_rows += 1
            if page_rows == 0:
                break
            nxt = soup.select_one("a.next_page[href], a[rel='next']")
            if not nxt:
                break
            page += 1
            time.sleep(0.2)
    except Exception as exc:
        log.error(f"  Maine: scraping failed: {exc}")
        return _normalise(pd.DataFrame(), "Maine")

    if not all_rows:
        log.error("  Maine: no rows found")
        return _normalise(pd.DataFrame(), "Maine")

    data = pd.DataFrame(all_rows)
    data.columns = [str(c).strip() for c in data.columns]
    col_map = {
        "Employer":    "company",
        "City":        "city",
        "Notice Date": "notice_date",
        "WARN Type":   "closure_type",
        "LWIB Area":   "notes",
    }
    data.rename(columns={k: v for k, v in col_map.items() if k in data.columns}, inplace=True)
    for dcol in ("notice_date",):
        if dcol in data.columns:
            parsed = pd.to_datetime(data[dcol], errors="coerce")
            data[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(data[dcol].astype(str).str.strip())
    log.info(f"  Maine: {len(data)} rows (pre-filter)")
    return _normalise(data, "Maine")


# ── Michigan ───────────────────────────────────────────────────────────────────
_MI_SOURCE_URL = "https://www.michigan.gov/leo/bureaus-agencies/wd/data-public-notices/warn-notices"
_MI_API_PATH   = "/leo/sxa/search/results/"
_MI_API_S      = "{8E97AB1D-D2D4-47F8-8CC4-3F1039C8854F}"
_MI_API_ITEMID = "{BE81F7C2-36A8-4FDE-853C-B05B6E090055}"
_MI_API_V      = "{1FFFCC21-5151-4A2B-ABFC-F7FE4E5C9783}"
_MI_API_SORT   = "Created Date sort,Descending"
_MI_PAGE_SIZE  = 20
_MI_MAX_PAGES  = 200

_MI_XHR_JS = """
var url = arguments[0];
var xhr = new XMLHttpRequest();
xhr.open('GET', url, false);
xhr.setRequestHeader('X-Requested-With', 'XMLHttpRequest');
xhr.send(null);
return [xhr.status, xhr.responseText];
"""


def _mi_build_api_url(page_end):
    sort_enc = _MI_API_SORT.replace(" ", "%20").replace(",", "%2C")
    v_enc    = _MI_API_V.replace("{", "%7B").replace("}", "%7D")
    return (
        f"{_MI_API_PATH}"
        f"?s={_MI_API_S}"
        f"&itemid={_MI_API_ITEMID}"
        f"&sig="
        f"&autoFireSearch=true"
        f"&v={v_enc}"
        f"&p={_MI_PAGE_SIZE}"
        f"&e={page_end}"
        f"&o={sort_enc}"
    )


def _mi_clean(text):
    return re.sub(r"\s+", " ", (text or "").replace("\xa0", " ")).strip()


def _mi_parse_card(card_tag):
    row = {"company": "", "city": "", "notes": "", "closure_type": "", "layoff_date": "", "employees_affected": ""}
    title_a = card_tag.find("a", class_="content-title-link")
    if title_a:
        row["company"] = _mi_clean(title_a.get_text(" ", strip=True))
    else:
        h3 = card_tag.find("h3")
        if h3:
            row["company"] = _mi_clean(h3.get_text(" ", strip=True))
    inner = str(card_tag)
    pattern = re.compile(
        r"<strong>\s*([^<]+?)\s*:?\s*</strong>\s*:?\s*([^<\n\r]+?)(?=\s*<(?:br|strong|/p))",
        re.IGNORECASE,
    )
    for m in pattern.finditer(inner):
        label = _mi_clean(m.group(1)).lower().rstrip(":")
        value = _mi_clean(m.group(2))
        if not value:
            continue
        if label == "city":
            row["city"] = value
        elif label == "county":
            row["notes"] = value
        elif label == "type of company action":
            row["closure_type"] = value
        elif label in ("layoff date", "closure date"):
            row["layoff_date"] = value
        elif label == "number of jobs impacted":
            row["employees_affected"] = value
    for li in card_tag.select("ul.disc > li"):
        text  = _mi_clean(li.get_text(" ", strip=True))
        lower = text.lower()
        if re.match(r"^city", lower) and not row["city"]:
            row["city"] = re.sub(r"^city\s*[:\-]?\s*", "", text, flags=re.I).strip()
        elif re.match(r"^county", lower) and not row["notes"]:
            row["notes"] = re.sub(r"^county\s*[:\-]?\s*", "", text, flags=re.I).strip()
        elif re.match(r"^type\s*of\s*company\s*action", lower) and not row["closure_type"]:
            row["closure_type"] = re.sub(r"^type\s*of\s*company\s*action\s*[:\-]?\s*", "", text, flags=re.I).strip()
        elif re.match(r"^layoff\s*date|^closure\s*date", lower) and not row["layoff_date"]:
            row["layoff_date"] = re.sub(r"^(layoff|closure)\s*date\s*[:\-]?\s*", "", text, flags=re.I).strip()
        elif re.match(r"^number\s*of\s*jobs\s*impacted", lower) and not row["employees_affected"]:
            row["employees_affected"] = re.sub(r"^number\s*of\s*jobs\s*impacted\s*[:\-]?\s*", "", text, flags=re.I).strip()
    return row


def scrape_michigan() -> pd.DataFrame:
    log.info("  Michigan: starting Selenium scrape")
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException
        from webdriver_manager.chrome import ChromeDriverManager
        import json as _json
    except ImportError as exc:
        log.error(f"  Michigan: missing dependency: {exc}")
        return _normalise(pd.DataFrame(), "Michigan")

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1600,1200")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )

    all_rows  = []
    seen_keys = set()
    total     = None

    try:
        driver.get(_MI_SOURCE_URL)
        try:
            WebDriverWait(driver, 45).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div.search-results__section-content")
                )
            )
        except TimeoutException:
            log.error("  Michigan: page did not load results within 45 seconds")
            return _normalise(pd.DataFrame(), "Michigan")
        time.sleep(3)

        try:
            count_el = driver.find_element(By.CSS_SELECTOR, ".results-count")
            m = re.search(r"(\d[\d,]*)", count_el.text)
            if m:
                total = int(m.group(1).replace(",", ""))
        except Exception:
            pass

        def ingest(cards):
            for card in cards:
                row = _mi_parse_card(card)
                key = (row["company"], row.get("layoff_date", ""), row.get("employees_affected", ""), row.get("city", ""))
                if row["company"] and key not in seen_keys:
                    seen_keys.add(key)
                    all_rows.append(row)

        soup0  = BeautifulSoup(driver.page_source, "lxml")
        batch1 = soup0.select("div.search-results__section-content")
        ingest(batch1)

        page_end  = _MI_PAGE_SIZE
        batch_num = 2
        while True:
            if total and len(all_rows) >= total:
                break
            result = driver.execute_script(_MI_XHR_JS, _mi_build_api_url(page_end))
            status, body = result[0], result[1]
            if status != 200 or not body:
                break
            try:
                data = _json.loads(body)
            except Exception:
                break
            results = data.get("Results", [])
            if not results:
                break
            cards = []
            for item in results:
                html = item.get("Html", "")
                if html:
                    fragment = BeautifulSoup(html, "lxml")
                    cards.extend(fragment.select("div.search-results__section-content"))
            prev_len = len(all_rows)
            ingest(cards)
            if not cards or (len(all_rows) == prev_len and len(cards) < _MI_PAGE_SIZE):
                break
            if batch_num >= _MI_MAX_PAGES:
                break
            page_end  += _MI_PAGE_SIZE
            batch_num += 1
            time.sleep(0.5)
    finally:
        driver.quit()

    if not all_rows:
        log.error("  Michigan: no rows collected")
        return _normalise(pd.DataFrame(), "Michigan")

    df = pd.DataFrame(all_rows)
    for dcol in ("layoff_date",):
        if dcol in df.columns:
            parsed = pd.to_datetime(df[dcol], errors="coerce")
            df[dcol] = parsed.dt.strftime("%m/%d/%Y").fillna(df[dcol].astype(str).str.strip())
    log.info(f"  Michigan: {len(df)} rows (pre-filter)")
    return _normalise(df, "Michigan")


# ── Minnesota ──────────────────────────────────────────────────────────────────
# mn.gov is behind Radware Bot Manager; use layoffdata.com → Google Sheets instead
_MN_PAGE_URL = "https://layoffdata.com/minnesota/"
_MN_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_minnesota() -> pd.DataFrame:
    import io
    log.info("  Minnesota: fetching layoffdata.com/minnesota page")
    try:
        r = requests.get(_MN_PAGE_URL, headers=_MN_HDR, timeout=60)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        sheet_id = None
        for a in soup.find_all("a", href=True):
            href = a.get("href", "")
            m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", href)
            if m:
                sheet_id = m.group(1)
                break
        if not sheet_id:
            log.error("  Minnesota: no Google Sheets link found on layoffdata.com")
            return _normalise(pd.DataFrame(), "Minnesota")
    except Exception as exc:
        log.error(f"  Minnesota: failed page fetch: {exc}")
        return _normalise(pd.DataFrame(), "Minnesota")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        dl_hdrs = {**_MN_HDR, "Referer": "https://docs.google.com/"}
        resp = requests.get(export_url, headers=dl_hdrs, timeout=180)
        resp.raise_for_status()
        raw = resp.content
    except Exception as exc:
        log.error(f"  Minnesota: failed XLSX download: {exc}")
        return _normalise(pd.DataFrame(), "Minnesota")

    try:
        bio = io.BytesIO(raw)
        xl  = pd.ExcelFile(bio, engine="openpyxl")
        parts = []
        for sname in xl.sheet_names:
            sheet_df = xl.parse(sname)
            if not sheet_df.empty:
                parts.append(sheet_df)
        if not parts:
            log.error("  Minnesota: empty workbook")
            return _normalise(pd.DataFrame(), "Minnesota")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  Minnesota: failed to parse XLSX: {exc}")
        return _normalise(pd.DataFrame(), "Minnesota")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer", "firm"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "company" not in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "received", "warn date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["employee", "worker", "affected", "impacted"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["county", "region"]) and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  Minnesota: {len(df)} rows (pre-filter)")
    return _normalise(df, "Minnesota")


# ── Nevada ─────────────────────────────────────────────────────────────────────
_NV_PAGE_URL = "https://detr.nv.gov/Page/WARN"
_NV_BASE_URL = "https://detr.nv.gov"
_NV_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def _nv_discover_pdfs(html, years):
    soup = BeautifulSoup(html, "html.parser")
    found = {}
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        text = " ".join(a.get_text(" ", strip=True).split()).lower()
        if ".pdf" not in href.lower():
            continue
        for year in years:
            if re.search(rf"\b{year}\b\s+warn\s+act\s+notices", text, re.IGNORECASE):
                abs_url = href if href.startswith("http") else _NV_BASE_URL + href
                found[year] = abs_url
    return found


def _nv_norm_cell(v):
    s = "" if v is None else str(v)
    s = s.replace("\r", "\n")
    s = "\n".join(p.strip() for p in s.split("\n"))
    s = re.sub(r"\n{2,}", "\n", s)
    s = s.replace("\n", " | ")
    return re.sub(r"\s+", " ", s).strip(" |")


def _nv_looks_data_row(r):
    if len(r) < 8:
        return False
    left = " ".join(r[:3]).lower()
    if any(x in left for x in ["received", "effective", "date"]):
        return False
    if not (r[0].strip() or r[1].strip()):
        return False
    if not any(c.strip() for c in r[4:8]):
        return False
    return True


def _nv_extract_pdf(pdf_path, year, source_url):
    import fitz
    rows = []
    with fitz.open(str(pdf_path)) as doc:
        for page in doc:
            tables = page.find_tables()
            use_text = not tables.tables
            if tables.tables:
                first = tables.tables[0].extract() or []
                if len(first) <= 2:
                    use_text = True
            if use_text:
                tables = page.find_tables(vertical_strategy="text", horizontal_strategy="text")
            for table in tables.tables:
                extracted = table.extract() or []
                if not extracted:
                    continue
                nrows = [[_nv_norm_cell(c) for c in (r or [])] for r in extracted]
                for r in nrows:
                    if not _nv_looks_data_row(r):
                        continue
                    n = len(r)
                    if n >= 11:
                        received = r[0]; effective = r[1]
                        type_lc  = (r[2] + " " + r[3]).strip()
                        affected = r[4]
                        employer = " | ".join(x for x in [r[5], r[6], r[7]] if x)
                        city     = r[8]; county = r[9]
                    elif n >= 10:
                        received = r[0]; effective = r[1]
                        type_lc  = (r[2] + " " + r[3]).strip()
                        affected = r[4]
                        employer = " | ".join(x for x in [r[5], r[6]] if x)
                        city     = r[7]; county = r[8]
                    else:
                        received = r[0]; effective = r[1]
                        type_lc  = r[2]; affected = r[3]
                        employer = r[4]; city = r[5]
                        county   = r[6] if len(r) > 6 else ""
                    if employer.lower() in {"employer", ""}:
                        continue
                    rows.append({
                        "notice_date":        received,
                        "layoff_date":        effective,
                        "closure_type":       type_lc,
                        "employees_affected": affected,
                        "company":            employer,
                        "city":               city,
                        "notes":              county,
                    })
    return rows


def scrape_nevada() -> pd.DataFrame:
    import tempfile
    import uuid as _uuid
    log.info("  Nevada: fetching WARN page")
    cy = date.today().year
    target_years = [str(cy - 1), str(cy), str(cy + 1)]
    try:
        r = requests.get(_NV_PAGE_URL, headers=_NV_HDR, timeout=45)
        r.raise_for_status()
        year_pdfs = _nv_discover_pdfs(r.text, target_years)
    except Exception as exc:
        log.error(f"  Nevada: failed to fetch page: {exc}")
        return _normalise(pd.DataFrame(), "Nevada")

    if not year_pdfs:
        log.error(f"  Nevada: no PDF links found for {target_years}")
        return _normalise(pd.DataFrame(), "Nevada")

    all_rows = []
    tmp_dir = Path(tempfile.gettempdir())
    for year, url in year_pdfs.items():
        tmp_pdf = tmp_dir / f"_nv_warn_{_uuid.uuid4().hex}.pdf"
        try:
            resp = requests.get(url, headers=_NV_HDR, timeout=120, stream=True)
            resp.raise_for_status()
            with open(tmp_pdf, "wb") as f:
                for chunk in resp.iter_content(65536):
                    if chunk:
                        f.write(chunk)
            rows = _nv_extract_pdf(tmp_pdf, year, url)
            all_rows.extend(rows)
            log.info(f"  Nevada: {len(rows)} rows from {year} PDF")
        except Exception as exc:
            log.error(f"  Nevada: failed {year} PDF: {exc}")
        finally:
            if tmp_pdf.exists():
                try:
                    tmp_pdf.unlink()
                except Exception:
                    pass

    if not all_rows:
        log.error("  Nevada: no rows extracted")
        return _normalise(pd.DataFrame(), "Nevada")

    df = pd.DataFrame(all_rows)
    log.info(f"  Nevada: {len(df)} rows (pre-filter)")
    return _normalise(df, "Nevada")


# ── New Hampshire ──────────────────────────────────────────────────────────────
_NH_PAGE_URL = "https://layoffdata.com/new-hampshire/"
_NH_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_new_hampshire() -> pd.DataFrame:
    import io
    log.info("  New Hampshire: fetching layoffdata.com page")
    sheet_id = None
    for attempt in range(3):
        try:
            if attempt > 0:
                time.sleep(attempt * 5)
            r = requests.get(_NH_PAGE_URL, headers=_NH_HDR, timeout=60)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a.get("href", "")
                m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", href)
                if m:
                    sheet_id = m.group(1)
                    break
            if sheet_id:
                break
            log.warning("  New Hampshire: no Google Sheets link found, retrying")
        except Exception as exc:
            log.warning(f"  New Hampshire: page fetch attempt {attempt+1} failed: {exc}")
    if not sheet_id:
        log.error("  New Hampshire: could not find Google Sheets link after retries")
        return _normalise(pd.DataFrame(), "New Hampshire")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        dl_hdrs = {**_NH_HDR, "Referer": "https://docs.google.com/"}
        resp = requests.get(export_url, headers=dl_hdrs, timeout=180)
        resp.raise_for_status()
        raw = resp.content
    except Exception as exc:
        log.error(f"  New Hampshire: failed XLSX download: {exc}")
        return _normalise(pd.DataFrame(), "New Hampshire")

    try:
        bio = io.BytesIO(raw)
        xl  = pd.ExcelFile(bio, engine="openpyxl")
        parts = []
        for sname in xl.sheet_names:
            sheet_df = xl.parse(sname)
            if not sheet_df.empty:
                parts.append(sheet_df)
        if not parts:
            log.error("  New Hampshire: empty workbook")
            return _normalise(pd.DataFrame(), "New Hampshire")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  New Hampshire: failed to parse XLSX: {exc}")
        return _normalise(pd.DataFrame(), "New Hampshire")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer", "firm"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "company" not in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "received", "warn date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["employee", "worker", "affected", "impacted"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["county", "region"]) and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  New Hampshire: {len(df)} rows (pre-filter)")
    return _normalise(df, "New Hampshire")


# ── New Jersey ─────────────────────────────────────────────────────────────────
_NJ_PAGE_URL     = "https://www.nj.gov/labor/business-services/layoffs-and-closing/file-warn-notice/"
_NJ_BASE_URL     = "https://www.nj.gov"
_NJ_TARGET_YEARS = ["2025", "2026"]
_NJ_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def _nj_discover_pdfs(html):
    soup = BeautifulSoup(html, "html.parser")
    found = {}
    for a in soup.find_all("a", href=True):
        text = " ".join(a.get_text(" ", strip=True).split())
        href = a["href"].strip()
        if not href:
            continue
        for year in _NJ_TARGET_YEARS:
            if re.search(rf"\b{year}\b.*warn notice archive", text, re.IGNORECASE):
                if ".pdf" in href.lower() or "pdf" in text.lower():
                    found[year] = href if href.startswith("http") else _NJ_BASE_URL + href
    return found


def _nj_norm_cell(v):
    if v is None:
        return ""
    txt = str(v).replace("\r", "\n")
    txt = "\n".join(p.strip() for p in txt.split("\n"))
    txt = re.sub(r"\n{2,}", "\n", txt)
    txt = txt.replace("\n", " | ")
    return txt.strip(" | ")


def _nj_extract_pdf(pdf_path, year, source_url):
    import fitz
    canonical = ["Company", "City", "Month Posted", "Effective Date", "Workforce Affected"]

    def map_headers(header_row):
        mapped = {}
        for idx, raw in enumerate(header_row):
            h = _nj_norm_cell(raw).lower().strip()
            if "company" in h and "Company" not in mapped:
                mapped["Company"] = [idx]
            elif "city" in h and "City" not in mapped:
                mapped["City"] = [idx]
            elif "month" in h and "Month Posted" not in mapped:
                mapped["Month Posted"] = [idx]
            elif "effective" in h and "Effective Date" not in mapped:
                mapped["Effective Date"] = [idx]
            elif ("workforce" in h or "affected" in h) and "Workforce Affected" not in mapped:
                mapped["Workforce Affected"] = [idx]
        return mapped

    def infer_map(n):
        if n == 7:
            return {"Company": [0, 1], "City": [2], "Month Posted": [4], "Effective Date": [5], "Workforce Affected": [6]}
        if n == 6:
            return {"Company": [0], "City": [1], "Month Posted": [2], "Effective Date": [3], "Workforce Affected": [5]}
        if n >= 10:
            return {"Company": [0, 1], "City": [2, 3, 4], "Month Posted": [5], "Effective Date": [6, 7, 8], "Workforce Affected": [9]}
        return {"Company": [0], "City": [1], "Month Posted": [2], "Effective Date": [3], "Workforce Affected": [4]}

    all_rows = []
    with fitz.open(str(pdf_path)) as doc:
        for page_idx, page in enumerate(doc, start=1):
            tables = page.find_tables()
            if not tables.tables:
                tables = page.find_tables(vertical_strategy="text", horizontal_strategy="text")
            for t_idx, table in enumerate(tables.tables, start=1):
                rows = table.extract() or []
                if not rows:
                    continue
                nrows = [[_nj_norm_cell(v) for v in (row or [])] for row in rows]
                hidx = None
                for ridx, row in enumerate(nrows[:6]):
                    hmap_test = map_headers([_nj_norm_cell(x) for x in row])
                    if "Company" in hmap_test and "City" in hmap_test:
                        hidx = ridx
                        break
                if hidx is not None:
                    header_map = map_headers(nrows[hidx])
                else:
                    sample = next((r for r in nrows if any(c.strip() for c in r)), [])
                    header_map = infer_map(len(sample))
                    hidx = -1
                if "Company" not in header_map:
                    continue
                for r in nrows[hidx + 1:]:
                    if not any(c.strip() for c in r):
                        continue
                    row_dict = {k: "" for k in canonical}
                    for col_name, col_idxs in header_map.items():
                        pieces = [r[i].strip() for i in col_idxs if i < len(r) and r[i].strip()]
                        if pieces:
                            row_dict[col_name] = " | ".join(pieces)
                    comp = row_dict["Company"].strip().lower()
                    if comp in {"company", ""} or "warn notices" in comp:
                        continue
                    if any(x in comp for x in ["month posted", "effective date", "workforce"]):
                        continue
                    all_rows.append({
                        "company":            row_dict["Company"],
                        "city":               row_dict["City"],
                        "notice_date":        row_dict["Month Posted"],
                        "layoff_date":        row_dict["Effective Date"],
                        "employees_affected": row_dict["Workforce Affected"],
                    })

    merged = []
    for row in all_rows:
        if not row["company"].strip() and merged:
            prev = merged[-1]
            for key in ["company", "city", "notice_date", "layoff_date", "employees_affected"]:
                val = row.get(key, "").strip()
                if val and not prev.get(key):
                    prev[key] = val
            continue
        merged.append(row)
    return merged


def scrape_new_jersey() -> pd.DataFrame:
    import tempfile
    import uuid as _uuid
    log.info("  New Jersey: fetching WARN page")
    try:
        r = requests.get(_NJ_PAGE_URL, headers=_NJ_HDR, timeout=45)
        r.raise_for_status()
        year_pdfs = _nj_discover_pdfs(r.text)
    except Exception as exc:
        log.error(f"  New Jersey: failed page fetch: {exc}")
        return _normalise(pd.DataFrame(), "New Jersey")

    if not year_pdfs:
        log.error("  New Jersey: no PDF links found")
        return _normalise(pd.DataFrame(), "New Jersey")

    all_rows = []
    tmp_dir = Path(tempfile.gettempdir())
    for year, url in year_pdfs.items():
        tmp_pdf = tmp_dir / f"_nj_warn_{_uuid.uuid4().hex}.pdf"
        try:
            resp = requests.get(url, headers=_NJ_HDR, timeout=90, stream=True)
            resp.raise_for_status()
            with open(tmp_pdf, "wb") as f:
                for chunk in resp.iter_content(32768):
                    if chunk:
                        f.write(chunk)
            rows = _nj_extract_pdf(tmp_pdf, year, url)
            all_rows.extend(rows)
            log.info(f"  New Jersey: {len(rows)} rows from {year} PDF")
        except Exception as exc:
            log.error(f"  New Jersey: failed {year} PDF: {exc}")
        finally:
            if tmp_pdf.exists():
                try:
                    tmp_pdf.unlink()
                except Exception:
                    pass

    if not all_rows:
        log.error("  New Jersey: no rows extracted")
        return _normalise(pd.DataFrame(), "New Jersey")

    df = pd.DataFrame(all_rows)
    log.info(f"  New Jersey: {len(df)} rows (pre-filter)")
    return _normalise(df, "New Jersey")


# ── New Mexico ─────────────────────────────────────────────────────────────────
_NM_PAGE_URL = "https://www.dws.state.nm.us/Rapid-Response/WARN-Notices"
_NM_BASE_URL = "https://www.dws.state.nm.us"
_NM_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def _nm_discover_pdfs(html):
    from urllib.parse import unquote
    from pathlib import PurePosixPath
    soup = BeautifulSoup(html, "html.parser")
    seen = set()
    found = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if "warn.pdf" not in href.lower():
            continue
        if "/portals/0/dm/business/" not in href.lower():
            continue
        abs_url = href if href.startswith("http") else _NM_BASE_URL + href
        if abs_url in seen:
            continue
        raw_name = unquote(PurePosixPath(href.split("?", 1)[0]).name)
        stem = Path(raw_name).stem
        m = re.match(r"^(\d{4})", stem.replace(" ", "_"))
        if not m:
            continue
        year = m.group(1)
        seen.add(abs_url)
        found.append((year, abs_url))
    found.sort(key=lambda x: int(x[0]), reverse=True)
    return found


def _nm_clean_cell(v):
    if v is None:
        return ""
    s = str(v).replace("\r", " ").replace("\n", " ")
    return " ".join(s.split()).strip()


def _nm_extract_pdf(pdf_path):
    import fitz
    doc = fitz.open(str(pdf_path))
    try:
        best_rows = []
        best_score = 0
        for page in doc:
            tf = page.find_tables()
            if not tf.tables:
                continue
            for t in tf.tables:
                raw = t.extract() or []
                rows = [[_nm_clean_cell(c) for c in row] for row in raw]
                rows = [r for r in rows if any(c for c in r)]
                if len(rows) > best_score:
                    best_score = len(rows)
                    best_rows = rows
        if not best_rows:
            return [], []
        header = [_nm_clean_cell(h) for h in best_rows[0]]
        body   = best_rows[1:]
        while body and not any(body[-1]):
            body.pop()
        date_like = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}$")
        filtered = []
        for row in body:
            if not row:
                continue
            c0 = row[0] if row else ""
            if c0 and not date_like.match(c0) and not filtered:
                continue
            if c0 and date_like.match(c0):
                filtered.append(row)
            elif filtered and any(row[1:]):
                filtered.append(row)
        body = filtered if filtered else body
        return header, body
    finally:
        doc.close()


def scrape_new_mexico() -> pd.DataFrame:
    import tempfile
    import uuid as _uuid
    log.info("  New Mexico: fetching WARN page")
    cy = date.today().year
    target_years = {str(cy - 1), str(cy), str(cy + 1)}

    pdf_list = []
    for attempt in range(3):
        try:
            if attempt > 0:
                time.sleep(attempt * 5)
            r = requests.get(_NM_PAGE_URL, headers=_NM_HDR, timeout=90)
            r.raise_for_status()
            pdf_list = _nm_discover_pdfs(r.text)
            break
        except Exception as exc:
            log.warning(f"  New Mexico: page fetch attempt {attempt+1} failed: {exc}")
            if attempt == 2:
                log.error("  New Mexico: all attempts failed")
                return _normalise(pd.DataFrame(), "New Mexico")

    pdf_list = [(yr, url) for yr, url in pdf_list if yr in target_years]
    if not pdf_list:
        log.error(f"  New Mexico: no PDF links found for {target_years}")
        return _normalise(pd.DataFrame(), "New Mexico")

    all_rows = []
    tmp_dir = Path(tempfile.gettempdir())
    for year, url in pdf_list:
        tmp_pdf = tmp_dir / f"_nm_warn_{_uuid.uuid4().hex}.pdf"
        try:
            resp = requests.get(url, headers=_NM_HDR, timeout=120, allow_redirects=True)
            resp.raise_for_status()
            if len(resp.content) < 500 and b"<html" in resp.content[:200].lower():
                raise RuntimeError(f"Got HTML instead of PDF from {url}")
            with open(tmp_pdf, "wb") as f:
                f.write(resp.content)
            header, body = _nm_extract_pdf(tmp_pdf)
            if not header:
                log.warning(f"  New Mexico: no table found in {year} PDF")
                continue
            # Map columns dynamically
            col_map_idx = {}
            for i, h in enumerate(header):
                hl = h.lower()
                if any(x in hl for x in ["notice", "received"]) and "notice_date" not in col_map_idx:
                    col_map_idx["notice_date"] = i
                elif any(x in hl for x in ["effective", "layoff"]) and "layoff_date" not in col_map_idx:
                    col_map_idx["layoff_date"] = i
                elif any(x in hl for x in ["company", "employer", "firm"]) and "company" not in col_map_idx:
                    col_map_idx["company"] = i
                elif "city" in hl and "company" not in hl and "city" not in col_map_idx:
                    col_map_idx["city"] = i
                elif any(x in hl for x in ["county", "region"]) and "notes" not in col_map_idx:
                    col_map_idx["notes"] = i
                elif any(x in hl for x in ["employ", "worker", "affected", "number"]) and "employees_affected" not in col_map_idx:
                    col_map_idx["employees_affected"] = i
                elif any(x in hl for x in ["type", "action", "closure"]) and "closure_type" not in col_map_idx:
                    col_map_idx["closure_type"] = i
            n = len(header)
            for row in body:
                row_padded = row + [""] * max(0, n - len(row))
                rec = {}
                for field, idx in col_map_idx.items():
                    if idx < len(row_padded):
                        rec[field] = row_padded[idx]
                if rec:
                    all_rows.append(rec)
            log.info(f"  New Mexico: {len(body)} rows from {year} PDF")
        except Exception as exc:
            log.error(f"  New Mexico: failed {year} PDF: {exc}")
        finally:
            if tmp_pdf.exists():
                try:
                    tmp_pdf.unlink()
                except Exception:
                    pass

    if not all_rows:
        log.error("  New Mexico: no rows extracted")
        return _normalise(pd.DataFrame(), "New Mexico")

    df = pd.DataFrame(all_rows)
    log.info(f"  New Mexico: {len(df)} rows (pre-filter)")
    return _normalise(df, "New Mexico")


# ── Mississippi ────────────────────────────────────────────────────────────────
_MS_PAGE_URL = "https://layoffdata.com/mississippi/"
_MS_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_mississippi() -> pd.DataFrame:
    import io
    log.info("  Mississippi: fetching layoffdata.com page")
    sheet_id = None
    for attempt in range(3):
        try:
            if attempt > 0:
                time.sleep(attempt * 5)
            r = requests.get(_MS_PAGE_URL, headers=_MS_HDR, timeout=60)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a.get("href", "")
                m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", href)
                if m:
                    sheet_id = m.group(1)
                    break
            if sheet_id:
                break
            log.warning("  Mississippi: no Google Sheets link found, retrying")
        except Exception as exc:
            log.warning(f"  Mississippi: page fetch attempt {attempt+1} failed: {exc}")
    if not sheet_id:
        log.error("  Mississippi: could not find Google Sheets link after retries")
        return _normalise(pd.DataFrame(), "Mississippi")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        dl_hdrs = {**_MS_HDR, "Referer": "https://docs.google.com/"}
        resp = requests.get(export_url, headers=dl_hdrs, timeout=180)
        resp.raise_for_status()
        raw = resp.content
    except Exception as exc:
        log.error(f"  Mississippi: failed XLSX download: {exc}")
        return _normalise(pd.DataFrame(), "Mississippi")

    try:
        bio = io.BytesIO(raw)
        xl  = pd.ExcelFile(bio, engine="openpyxl")
        parts = []
        for sname in xl.sheet_names:
            sheet_df = xl.parse(sname)
            if not sheet_df.empty:
                parts.append(sheet_df)
        if not parts:
            log.error("  Mississippi: empty workbook")
            return _normalise(pd.DataFrame(), "Mississippi")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  Mississippi: failed to parse XLSX: {exc}")
        return _normalise(pd.DataFrame(), "Mississippi")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer", "firm"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "company" not in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "received", "warn date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["employee", "worker", "affected", "impacted"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["county", "region"]) and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  Mississippi: {len(df)} rows (pre-filter)")
    return _normalise(df, "Mississippi")


# ── Missouri ───────────────────────────────────────────────────────────────────
_MO_PAGE_URL = "https://layoffdata.com/missouri/"
_MO_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_missouri() -> pd.DataFrame:
    import io
    log.info("  Missouri: fetching layoffdata.com page")
    sheet_id = None
    for attempt in range(3):
        try:
            if attempt > 0:
                time.sleep(attempt * 5)
            r = requests.get(_MO_PAGE_URL, headers=_MO_HDR, timeout=60)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a.get("href", "")
                m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", href)
                if m:
                    sheet_id = m.group(1)
                    break
            if sheet_id:
                break
            log.warning("  Missouri: no Google Sheets link found, retrying")
        except Exception as exc:
            log.warning(f"  Missouri: page fetch attempt {attempt+1} failed: {exc}")
    if not sheet_id:
        log.error("  Missouri: could not find Google Sheets link after retries")
        return _normalise(pd.DataFrame(), "Missouri")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        dl_hdrs = {**_MO_HDR, "Referer": "https://docs.google.com/"}
        resp = requests.get(export_url, headers=dl_hdrs, timeout=180)
        resp.raise_for_status()
        raw = resp.content
    except Exception as exc:
        log.error(f"  Missouri: failed XLSX download: {exc}")
        return _normalise(pd.DataFrame(), "Missouri")

    try:
        bio = io.BytesIO(raw)
        xl  = pd.ExcelFile(bio, engine="openpyxl")
        parts = []
        for sname in xl.sheet_names:
            sheet_df = xl.parse(sname)
            if not sheet_df.empty:
                parts.append(sheet_df)
        if not parts:
            log.error("  Missouri: empty workbook")
            return _normalise(pd.DataFrame(), "Missouri")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  Missouri: failed to parse XLSX: {exc}")
        return _normalise(pd.DataFrame(), "Missouri")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer", "firm"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "company" not in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "received", "warn date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["employee", "worker", "affected", "impacted"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["county", "region"]) and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  Missouri: {len(df)} rows (pre-filter)")
    return _normalise(df, "Missouri")


# ── Montana ────────────────────────────────────────────────────────────────────
_MT_PAGE_URL = "https://wsd.dli.mt.gov/wioa/related-links/warn-notice-page"
_MT_BASE_URL = "https://wsd.dli.mt.gov"
_MT_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def _mt_detect_header_row(df_raw) -> int | None:
    required = {"year", "date of notice", "name of company"}
    for idx in range(min(len(df_raw), 40)):
        vals = {re.sub(r"\s+", " ", str(v)).strip().lower() for v in df_raw.iloc[idx].tolist()}
        if required.issubset(vals):
            return idx
    return None


def scrape_montana() -> pd.DataFrame:
    import io
    from urllib.parse import urljoin
    log.info("  Montana: fetching WARN page")
    try:
        r = requests.get(_MT_PAGE_URL, headers=_MT_HDR, timeout=45)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        log.error(f"  Montana: page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Montana")

    wb_url = None
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        text = " ".join(a.get_text(" ", strip=True).split()).lower()
        if "warn notices" in text and ".xlsx" in href.lower():
            wb_url = urljoin(_MT_PAGE_URL, href)
            break
    if not wb_url:
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            combined = (href + " " + a.get_text(" ", strip=True)).lower()
            if ".xlsx" in href.lower() and "warn" in combined:
                wb_url = urljoin(_MT_PAGE_URL, href)
                break
    if not wb_url:
        log.error("  Montana: no XLSX link found on page")
        return _normalise(pd.DataFrame(), "Montana")

    try:
        resp = requests.get(wb_url, headers=_MT_HDR, timeout=120, stream=True)
        resp.raise_for_status()
        data = b"".join(resp.iter_content(65536))
    except Exception as exc:
        log.error(f"  Montana: failed XLSX download: {exc}")
        return _normalise(pd.DataFrame(), "Montana")

    try:
        bio = io.BytesIO(data)
        xl = pd.ExcelFile(bio, engine="openpyxl")
        best_df, best_rows = None, -1
        for sheet_name in xl.sheet_names:
            raw = xl.parse(sheet_name, header=None)
            hidx = _mt_detect_header_row(raw)
            if hidx is None:
                continue
            data_df = raw.iloc[hidx + 1:].copy()
            hdr_vals = raw.iloc[hidx].tolist()
            data_df.columns = [
                re.sub(r"\s+", " ", str(v)).strip() if v is not None else f"col_{j}"
                for j, v in enumerate(hdr_vals)
            ]
            data_df = data_df.dropna(how="all").reset_index(drop=True)
            if len(data_df) > best_rows:
                best_rows, best_df = len(data_df), data_df
    except Exception as exc:
        log.error(f"  Montana: failed to parse XLSX: {exc}")
        return _normalise(pd.DataFrame(), "Montana")

    if best_df is None or best_df.empty:
        log.error("  Montana: no data sheet found")
        return _normalise(pd.DataFrame(), "Montana")

    df = best_df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(
                lambda v: re.sub(r"\s+", " ", str(v)).strip().replace("nan", "") if v is not None else ""
            )

    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if "name of company" in cl and "company" not in col_map.values():
            col_map[col] = "company"
        elif "date of notice" in cl and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif "date of impact" in cl and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif "number of employees" in cl and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif "county" in cl and "city" not in cl and "notes" not in col_map.values():
            col_map[col] = "notes"
        elif "city" in cl and "city" not in col_map.values():
            col_map[col] = "city"
    df.rename(columns=col_map, inplace=True)
    for dcol in ["notice_date", "layoff_date"]:
        if dcol in df.columns:
            parsed = pd.to_datetime(df[dcol], errors="coerce")
            df[dcol] = parsed.dt.strftime("%Y-%m-%d").where(parsed.notna(), df[dcol].astype(str))
    log.info(f"  Montana: {len(df)} rows (pre-filter)")
    return _normalise(df, "Montana")


# ── Nebraska ───────────────────────────────────────────────────────────────────
_NE_PAGE_URL = (
    "https://dol.nebraska.gov/ReemploymentServices/LayoffServices/LayoffsAndDownsizingWARN"
)
_NE_BASE_URL = "https://dol.nebraska.gov"
_NE_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def _ne_clean(s: object) -> str:
    if s is None:
        return ""
    return " ".join(str(s).replace("\r", " ").replace("\n", " ").split()).strip()


def scrape_nebraska() -> pd.DataFrame:
    log.info("  Nebraska: fetching WARN page")
    try:
        r = requests.get(_NE_PAGE_URL, headers=_NE_HDR, timeout=90)
        r.raise_for_status()
        html = r.text
    except Exception as exc:
        log.error(f"  Nebraska: page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Nebraska")

    soup = BeautifulSoup(html, "html.parser")
    table = None
    for t in soup.find_all("table"):
        blob = t.get_text(" ", strip=True).lower()
        if "jobs affected" in blob and "company" in blob and "date" in blob:
            table = t
            break
    if table is None:
        table = soup.find("table")
    if table is None:
        log.error("  Nebraska: no WARN table found")
        return _normalise(pd.DataFrame(), "Nebraska")

    rows = []
    for tr in table.find_all("tr"):
        tds = tr.find_all("td", recursive=False)
        if len(tds) < 2:
            continue
        while len(tds) < 4:
            tds.append(None)
        date_s    = _ne_clean(tds[0].get_text(" ", strip=True)) if tds[0] else ""
        company_s = _ne_clean(tds[1].get_text(" ", strip=True)) if tds[1] else ""
        jobs_s    = _ne_clean(tds[2].get_text(" ", strip=True)) if tds[2] else ""
        loc_s     = _ne_clean(tds[3].get_text(" ", strip=True)) if tds[3] else ""
        if not date_s and not company_s:
            continue
        if company_s.lower() in {"company", "employer", "firm"}:
            continue
        rows.append({
            "notice_date":        date_s,
            "company":            company_s,
            "employees_affected": jobs_s,
            "city":               loc_s,
        })

    if not rows:
        log.error("  Nebraska: no rows parsed")
        return _normalise(pd.DataFrame(), "Nebraska")

    df = pd.DataFrame(rows)
    log.info(f"  Nebraska: {len(df)} rows (pre-filter)")
    return _normalise(df, "Nebraska")


# ── New York ───────────────────────────────────────────────────────────────────
_NY_LISTING_URL  = "https://dol.ny.gov/warn-notices"
_NY_TABLEAU_URL  = (
    "https://public.tableau.com/views/"
    "WorkerAdjustmentRetrainingNotificationWARN/WARN.csv?:showVizHome=no"
)
_NY_BASE_URL = "https://dol.ny.gov"
_NY_NAV_SLUGS = {
    "notices", "dashboard", "regulations", "worker",
    "businesses", "jobseekers", "general-inquiries",
    "2024", "2023", "2022", "2021",
}
_NY_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/pdf,*/*",
}


def _ny_parse_notice_pdf(content: bytes) -> dict:
    """Extract fields from a single NY WARN notice PDF."""
    import fitz
    try:
        doc = fitz.open(stream=content, filetype="pdf")
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
    except Exception:
        return {}

    def _find(pattern, default=""):
        m = re.search(pattern, text, re.IGNORECASE)
        return m.group(1).strip() if m else default

    company      = _find(r"Company:\s*\n(.+)")
    city_line    = _find(r"Company:\s*\n.+\n(.+)")
    city         = re.sub(r",\s*NY.*", "", city_line).strip() if city_line else ""
    affected     = _find(r"Total Number of Affected Workers:\s*(\d[\d,]*)")
    notice_date  = _find(r"Date of Notice:\s*([\w]+ \d+,\s*\d{4}|\d{1,2}/\d{1,2}/\d{4}|\d{4}-\d{2}-\d{2})")
    layoff_date  = _find(r"Layoff Start Date:\s*([\w]+ \d+,\s*\d{4}|\d{1,2}/\d{1,2}/\d{4}|\d{4}-\d{2}-\d{2})")
    closure_type = _find(r"Reason For Layoff:\s*(.+)")
    county       = _find(r"County:\s*(.+)")

    return {
        "company":            company,
        "city":               city,
        "employees_affected": affected.replace(",", ""),
        "notice_date":        notice_date,
        "layoff_date":        layoff_date,
        "closure_type":       closure_type,
        "notes":              county,
    }


def _ny_parse_tableau_csv(content: str) -> list[dict]:
    """Parse the Tableau Public CSV export for NY WARN data."""
    import io as _io
    try:
        df = pd.read_csv(_io.StringIO(content))
        df.columns = [str(c).strip() for c in df.columns]
        col_map = {}
        mapped_vals: set[str] = set()

        def _try_map(col: str, std: str) -> bool:
            if std not in mapped_vals:
                col_map[col] = std
                mapped_vals.add(std)
                return True
            return False

        for col in df.columns:
            cl = col.lower().strip()
            if "business legal name" in cl or ("company" in cl and "name" in cl) or cl == "employer":
                _try_map(col, "company")
            elif cl in ("impacted site county", "county"):
                _try_map(col, "notes")
            elif "date of warn notice" in cl or "notice date" in cl:
                _try_map(col, "notice_date")
            elif "date layoff" in cl or "layoff start" in cl or ("effective" in cl and "date" in cl):
                _try_map(col, "layoff_date")
            elif ("number of affected" in cl or "workers affected" in cl) and "employees_affected" not in mapped_vals:
                _try_map(col, "employees_affected")
            elif cl == "city":
                _try_map(col, "city")
            elif "layoff or closure" in cl:
                _try_map(col, "closure_type")

        df.rename(columns=col_map, inplace=True)
        # Drop columns that were not mapped and keep only mapped ones
        keep = list(mapped_vals)
        df = df[[c for c in keep if c in df.columns]]
        df.fillna("", inplace=True)
        return df.to_dict("records")
    except Exception:
        return []


def scrape_new_york() -> pd.DataFrame:
    import io
    all_rows: list[dict] = []

    # ── Source 1: Tableau Public CSV (current-year view) ──────────────────────
    log.info("  New York: fetching Tableau Public CSV")
    try:
        r = requests.get(_NY_TABLEAU_URL, headers=_NY_HDR, timeout=45)
        r.raise_for_status()
        tableau_rows = _ny_parse_tableau_csv(r.text)
        if tableau_rows:
            log.info(f"  New York: {len(tableau_rows)} rows from Tableau CSV")
            all_rows.extend(tableau_rows)
    except Exception as exc:
        log.warning(f"  New York: Tableau CSV fetch failed: {exc}")

    # ── Source 2: Listing page → individual notice PDFs ───────────────────────
    log.info("  New York: fetching WARN notice listing page")
    try:
        r = requests.get(_NY_LISTING_URL, headers=_NY_HDR, timeout=45)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        notice_links = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if not href.startswith("/warn-"):
                continue
            slug = href[6:]
            if any(slug.startswith(n) for n in _NY_NAV_SLUGS):
                continue
            notice_links.append((_NY_BASE_URL + href, a.get_text(strip=True)))
        log.info(f"  New York: found {len(notice_links)} individual notice links")
    except Exception as exc:
        log.warning(f"  New York: listing page fetch failed: {exc}")
        notice_links = []

    for i, (url, company_text) in enumerate(notice_links):
        try:
            time.sleep(0.3)
            pr = requests.get(url, headers=_NY_HDR, timeout=30)
            pr.raise_for_status()
            rec = _ny_parse_notice_pdf(pr.content)
            if not rec.get("company"):
                rec["company"] = company_text
            # Extract notice date from URL slug if PDF parse missed it
            if not rec.get("notice_date"):
                m = re.search(r"notice-date-(\d{1,2})(\d{2})(\d{4})", url)
                if m:
                    rec["notice_date"] = f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
            if rec.get("company"):
                all_rows.append(rec)
        except Exception as exc:
            # Use URL slug for company name + date extraction as fallback
            m_nd = re.search(r"notice-date-(\d{1,2})(\d{2})(\d{4})", url)
            fallback = {
                "company":    company_text,
                "notice_date": f"{m_nd.group(1)}/{m_nd.group(2)}/{m_nd.group(3)}" if m_nd else "",
            }
            all_rows.append(fallback)

    if not all_rows:
        log.error("  New York: no rows extracted")
        return _normalise(pd.DataFrame(), "New York")

    df = pd.DataFrame(all_rows)
    log.info(f"  New York: {len(df)} rows (pre-filter)")
    return _normalise(df, "New York")


# ── North Carolina ─────────────────────────────────────────────────────────────
_NC_CSV_URL = (
    "https://bi.nc.gov/t/COM-LEAD/views/WARNdashSaleforcedata/WARN/crosstab_download.csv"
)
_NC_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
}


def scrape_north_carolina() -> pd.DataFrame:
    import io
    log.info("  North Carolina: fetching Tableau CSV")
    for attempt in range(1, 4):
        try:
            r = requests.get(_NC_CSV_URL, headers=_NC_HDR, timeout=60, allow_redirects=True)
            if r.status_code == 200 and len(r.content) > 200:
                break
            log.warning(f"  North Carolina: HTTP {r.status_code} on attempt {attempt}")
        except Exception as exc:
            log.warning(f"  North Carolina: attempt {attempt} failed: {exc}")
        time.sleep(2)
    else:
        log.error("  North Carolina: could not download CSV")
        return _normalise(pd.DataFrame(), "North Carolina")

    try:
        df = pd.read_csv(io.StringIO(r.text))
        df.columns = [str(c).strip() for c in df.columns]
    except Exception as exc:
        log.error(f"  North Carolina: CSV parse failed: {exc}")
        return _normalise(pd.DataFrame(), "North Carolina")

    # Map the exact Tableau column names from the downloaded CSV
    col_map = {}
    for col in df.columns:
        cl = col.strip()
        if cl == "Company Name" and "company" not in col_map.values():
            col_map[cl] = "company"
        elif "City__c" in cl and "city" not in col_map.values():
            col_map[cl] = "city"
        elif cl == "Date Received" and "notice_date" not in col_map.values():
            col_map[cl] = "notice_date"
        elif cl == "Effective Date" and "layoff_date" not in col_map.values():
            col_map[cl] = "layoff_date"
        elif "Number affected" in cl and "employees_affected" not in col_map.values():
            col_map[cl] = "employees_affected"
        elif cl in ("TempPermClosLay", "Type") and "closure_type" not in col_map.values():
            col_map[cl] = "closure_type"
        elif cl == "County" and "notes" not in col_map.values():
            col_map[cl] = "notes"

    df.rename(columns=col_map, inplace=True)
    df.fillna("", inplace=True)
    log.info(f"  North Carolina: {len(df)} rows (pre-filter)")
    return _normalise(df, "North Carolina")


# ── North Dakota ───────────────────────────────────────────────────────────────
_ND_PAGE_URL = "https://layoffdata.com/north-dakota/"
_ND_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_north_dakota() -> pd.DataFrame:
    import io
    log.info("  North Dakota: fetching layoffdata.com page")
    sheet_id = None
    for attempt in range(3):
        try:
            if attempt > 0:
                time.sleep(attempt * 5)
            r = requests.get(_ND_PAGE_URL, headers=_ND_HDR, timeout=60)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", a.get("href", ""))
                if m:
                    sheet_id = m.group(1)
                    break
            if sheet_id:
                break
            log.warning("  North Dakota: no Google Sheets link found, retrying")
        except Exception as exc:
            log.warning(f"  North Dakota: attempt {attempt+1} failed: {exc}")
    if not sheet_id:
        log.error("  North Dakota: could not find Google Sheets link")
        return _normalise(pd.DataFrame(), "North Dakota")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        resp = requests.get(export_url, headers={**_ND_HDR, "Referer": "https://docs.google.com/"}, timeout=180)
        resp.raise_for_status()
    except Exception as exc:
        log.error(f"  North Dakota: XLSX download failed: {exc}")
        return _normalise(pd.DataFrame(), "North Dakota")

    try:
        xl = pd.ExcelFile(io.BytesIO(resp.content), engine="openpyxl")
        parts = [xl.parse(s) for s in xl.sheet_names if not xl.parse(s).empty]
        if not parts:
            return _normalise(pd.DataFrame(), "North Dakota")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  North Dakota: XLSX parse failed: {exc}")
        return _normalise(pd.DataFrame(), "North Dakota")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer", "firm"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "company" not in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "received", "warn date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["employee", "worker", "affected", "impacted"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["county", "region"]) and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  North Dakota: {len(df)} rows (pre-filter)")
    return _normalise(df, "North Dakota")


# ── Rhode Island ───────────────────────────────────────────────────────────────
_RI_PAGE_URL     = "https://dlt.ri.gov/employers/worker-adjustment-and-retraining-notification-warn"
_RI_BASE_URL     = "https://dlt.ri.gov"
_RI_TARGET_SHEETS = ["2025", "2026"]
_RI_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}


def _ri_detect_header(raw_df) -> int | None:
    expected = {"warn date", "date received", "company name"}
    for idx in range(min(len(raw_df), 20)):
        vals = {re.sub(r"\s+", " ", str(v)).strip().lower() for v in raw_df.iloc[idx].tolist()}
        if expected.issubset(vals):
            return idx
    return None


def scrape_rhode_island() -> pd.DataFrame:
    import io, re as _re, time
    log.info("  Rhode Island: fetching layoffdata.com page")
    # dlt.ri.gov is behind Cloudflare; use layoffdata.com as data source
    _ld_url = "https://layoffdata.com/rhode-island/"
    for attempt in range(1, 4):
        try:
            r = requests.get(_ld_url, headers=_RI_HDR, timeout=30)
            r.raise_for_status()
            break
        except Exception as exc:
            log.warning(f"  Rhode Island: attempt {attempt} failed: {exc}")
            if attempt < 3:
                time.sleep(30)
            else:
                log.error("  Rhode Island: could not reach layoffdata.com after retries")
                return _normalise(pd.DataFrame(), "Rhode Island")

    soup = BeautifulSoup(r.text, "html.parser")
    sheet_id = None
    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = _re.search(r"docs\.google\.com/spreadsheets/d/([A-Za-z0-9_-]+)", href)
        if m:
            sheet_id = m.group(1)
            break
    if not sheet_id:
        log.error("  Rhode Island: could not find Google Sheets link")
        return _normalise(pd.DataFrame(), "Rhode Island")

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    try:
        resp = requests.get(export_url, headers=_RI_HDR, timeout=60)
        resp.raise_for_status()
        raw_bytes = resp.content
    except Exception as exc:
        log.error(f"  Rhode Island: XLSX download failed: {exc}")
        return _normalise(pd.DataFrame(), "Rhode Island")

    try:
        xl = pd.ExcelFile(io.BytesIO(raw_bytes), engine="openpyxl")
        parts = []
        for sname in xl.sheet_names:
            sheet_df = xl.parse(sname)
            if not sheet_df.empty:
                parts.append(sheet_df)
        if not parts:
            log.error("  Rhode Island: XLSX had no data")
            return _normalise(pd.DataFrame(), "Rhode Island")
        df = pd.concat(parts, ignore_index=True)
    except Exception as exc:
        log.error(f"  Rhode Island: XLSX parse failed: {exc}")
        return _normalise(pd.DataFrame(), "Rhode Island")

    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["company", "employer"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif any(x in cl for x in ["location", "city", "address"]) and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["date received", "warn date", "notice date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["effective", "closing date", "layoff date"]) and "layoff_date" not in col_map.values():
            col_map[col] = "layoff_date"
        elif any(x in cl for x in ["affected", "number", "worker", "employee"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif "county" in cl and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  Rhode Island: {len(df)} rows (pre-filter)")
    return _normalise(df, "Rhode Island")


# ── Arizona ────────────────────────────────────────────────────────────────────
_AZ_BASE    = "https://www.azjobconnection.gov"
_AZ_NEW_URL = _AZ_BASE + "/search/warn_lookups/new"
_AZ_URL     = _AZ_BASE + "/search/warn_lookups"
_AZ_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_arizona() -> pd.DataFrame:
    from urllib.parse import urljoin
    log.info("  Arizona: fetching WARN results")
    s = requests.Session()
    s.headers.update(_AZ_HDR)
    try:
        s.get(_AZ_NEW_URL, timeout=30)
    except Exception as exc:
        log.error(f"  Arizona: session init failed: {exc}")
        return _normalise(pd.DataFrame(), "Arizona")

    params = {
        "commit": "Search",
        "q[notice_on_gteq]": "2025-01-01",
        "q[notice_eq]": "",
        "page": "1",
    }

    all_rows: list[dict] = []
    page = 1
    while page <= 300:
        params["page"] = str(page)
        try:
            r = s.get(_AZ_URL, params=params, timeout=60)
            r.raise_for_status()
        except Exception as exc:
            log.error(f"  Arizona: page {page} failed: {exc}")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        table = None
        for t in soup.find_all("table"):
            if "sortable" in (t.get("class") or []):
                table = t
                break
        if not table:
            break

        thead = table.find("thead")
        if not thead or not thead.find("tr"):
            break
        hdrs = [
            " ".join((th.find("a") or th).get_text(" ", strip=True).split())
            for th in thead.find("tr").find_all(["th", "td"])
        ]

        tbody = table.find("tbody")
        if not tbody:
            break
        trs = tbody.find_all("tr")
        if not trs:
            break

        for tr in trs:
            cells = tr.find_all("td", recursive=False)
            if not cells:
                continue
            texts = []
            for i, td in enumerate(cells):
                a = td.find("a", href=True) if i == 0 else None
                texts.append(" ".join((a or td).get_text(" ", strip=True).split()))
            while len(texts) < len(hdrs):
                texts.append("")
            all_rows.append(dict(zip(hdrs, texts[: len(hdrs)])))

        if not soup.select_one("a.next_page[href]"):
            break
        page += 1

    if not all_rows:
        log.warning("  Arizona: no rows found")
        return _normalise(pd.DataFrame(), "Arizona")

    df = pd.DataFrame(all_rows)
    df.columns = [str(c).strip() for c in df.columns]
    col_map: dict[str, str] = {}
    for col in df.columns:
        cl = col.lower()
        if any(x in cl for x in ["employer", "company"]) and "company" not in col_map.values():
            col_map[col] = "company"
        elif "city" in cl and "city" not in col_map.values():
            col_map[col] = "city"
        elif any(x in cl for x in ["notice", "date"]) and "notice_date" not in col_map.values():
            col_map[col] = "notice_date"
        elif any(x in cl for x in ["employee", "worker"]) and "employees_affected" not in col_map.values():
            col_map[col] = "employees_affected"
        elif any(x in cl for x in ["type", "action"]) and "closure_type" not in col_map.values():
            col_map[col] = "closure_type"
        elif "county" in cl and "notes" not in col_map.values():
            col_map[col] = "notes"
    df.rename(columns=col_map, inplace=True)
    log.info(f"  Arizona: {len(df)} rows (pre-filter)")
    return _normalise(df, "Arizona")


# ── Illinois ───────────────────────────────────────────────────────────────────
_IL_GSHEETS_ID = "1DKdR4I32lvM29o-_RYVOsvyhYAf7cLXpRZ-RfJVaKnc"


def scrape_illinois() -> pd.DataFrame:
    return _scrape_layoffdata_gsheets("Illinois", _IL_GSHEETS_ID)


# ── Massachusetts ──────────────────────────────────────────────────────────────
_MA_GSHEETS_ID = "1-5GreM1TTF7t2BtEGShdCorGCRKnwaAI5ihfXyQZItQ"


def scrape_massachusetts() -> pd.DataFrame:
    return _scrape_layoffdata_gsheets("Massachusetts", _MA_GSHEETS_ID)


# ── Utah ───────────────────────────────────────────────────────────────────────
_UT_URL          = "https://jobs.utah.gov/employer/business/warnnotices.html"
_UT_TARGET_YEARS = {"2025", "2026"}
_UT_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def scrape_utah() -> pd.DataFrame:
    log.info("  Utah: fetching WARN page")
    try:
        r = requests.get(_UT_URL, headers=_UT_HDR, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        log.error(f"  Utah: page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "Utah")

    rows = []
    for h2 in soup.find_all("h2"):
        year = " ".join(h2.get_text().split()).strip()
        if year not in _UT_TARGET_YEARS:
            continue
        table = h2.find_next("table")
        if not table:
            continue
        for tr in table.find_all("tr")[1:]:
            cols = tr.find_all(["td", "th"])
            if len(cols) < 4:
                continue
            c = [" ".join(td.get_text(" ", strip=True).split()) for td in cols[:4]]
            rows.append({
                "notice_date":        c[0],
                "company":            c[1],
                "city":               c[2],
                "employees_affected": c[3],
            })

    if not rows:
        log.warning("  Utah: no rows found")
        return _normalise(pd.DataFrame(), "Utah")

    df = pd.DataFrame(rows)
    log.info(f"  Utah: {len(df)} rows (pre-filter)")
    return _normalise(df, "Utah")


# ── South Carolina ─────────────────────────────────────────────────────────────
_SC_PAGE_URL    = "https://www.scworks.org/layoff-notification-reports"
_SC_BASE_URL    = "https://www.scworks.org"
_SC_TARGET_YEARS = [date.today().year - 1, date.today().year]
_SC_HDR = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
_SC_DATE_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}")
_SC_HEADER_7 = ["Company", "County", "Notice Date", "Layoff/Closure Date",
                "Impacted", "Layoff/Closure", "Address"]


def _sc_revision_key(url: str) -> tuple:
    from urllib.parse import unquote
    name = unquote(url)
    m = re.search(r"(\d{8})\s*\.pdf\s*$", name, re.I)
    if not m:
        return (0, 0, 0)
    d = m.group(1)
    mm, dd, yyyy = int(d[:2]), int(d[2:4]), int(d[4:8])
    if not (1 <= mm <= 12 and 1 <= dd <= 31 and 1990 <= yyyy <= 2100):
        return (0, 0, 0)
    return (yyyy, mm, dd)


def _sc_parse_pdf(pdf_bytes: bytes) -> list[dict]:
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = "\n".join(doc[i].get_text("text") for i in range(doc.page_count))
    doc.close()

    lines = [ln.strip() for ln in text.splitlines()]
    rows = []
    i = 0
    n = len(lines)
    while i < n:
        if i + 6 < n and lines[i:i + 7] == _SC_HEADER_7:
            i += 7
            continue
        if lines[i].startswith("Total WARN"):
            break
        if lines[i] == "County" and i + 1 < n and lines[i + 1] == "Impacted":
            break
        if i + 6 < n:
            c0, c1, c2, c3, c4, c5, c6 = lines[i:i + 7]
            if _SC_DATE_RE.match(c2) and c4.isdigit():
                rows.append({
                    "company":            c0,
                    "notes":              c1,
                    "notice_date":        c2,
                    "layoff_date":        c3,
                    "employees_affected": c4,
                    "closure_type":       c5,
                    "city":               (lambda m: m.group(1).strip() if m else re.sub(r",\s*SC.*", "", c6).rsplit(",", 1)[-1].strip())(re.search(r"([^,]+),\s*SC\b", c6, re.IGNORECASE)),
                })
                i += 7
                continue
        i += 1
    return rows


def scrape_south_carolina() -> pd.DataFrame:
    from urllib.parse import unquote, urljoin
    log.info("  South Carolina: fetching WARN page")
    try:
        r = requests.get(_SC_PAGE_URL, headers=_SC_HDR, timeout=45)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        log.error(f"  South Carolina: page fetch failed: {exc}")
        return _normalise(pd.DataFrame(), "South Carolina")

    # Collect all WARN PDF links
    all_pdfs: list[tuple[str, int]] = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href.lower().endswith(".pdf"):
            continue
        if "/sites/scworks/files" not in href and "warn" not in href.lower():
            if "warn" not in (a.get_text() or "").lower():
                continue
        full = urljoin(_SC_BASE_URL, href)
        name = unquote(full).rsplit("/", 1)[-1]
        m = re.match(r"(20\d{2})", name)
        if m:
            yr = int(m.group(1))
            all_pdfs.append((full, yr))

    # Pick newest revision per target year
    year_pdfs: dict[int, str] = {}
    for yr in _SC_TARGET_YEARS:
        cands = [u for u, y in all_pdfs if y == yr]
        if cands:
            year_pdfs[yr] = max(cands, key=_sc_revision_key)

    if not year_pdfs:
        log.error("  South Carolina: no PDF links found for target years")
        return _normalise(pd.DataFrame(), "South Carolina")

    all_rows = []
    for yr, url in year_pdfs.items():
        try:
            resp = requests.get(url, headers={**_SC_HDR, "Referer": _SC_PAGE_URL}, timeout=180)
            resp.raise_for_status()
            rows = _sc_parse_pdf(resp.content)
            all_rows.extend(rows)
            log.info(f"  South Carolina: {len(rows)} rows from {yr} PDF")
        except Exception as exc:
            log.error(f"  South Carolina: {yr} PDF failed: {exc}")

    if not all_rows:
        log.error("  South Carolina: no rows extracted")
        return _normalise(pd.DataFrame(), "South Carolina")

    df = pd.DataFrame(all_rows)
    log.info(f"  South Carolina: {len(df)} rows (pre-filter)")
    return _normalise(df, "South Carolina")


# ── Registry ──────────────────────────────────────────────────────────────────
# To add a new state: implement scrape_<state>() above and add it here.

SCRAPERS: dict[str, callable] = {
    "Alabama":    scrape_alabama,
    "Alaska":     scrape_alaska,
    "DC":         scrape_dc,
    "Maryland":   scrape_maryland,
    "Ohio":         scrape_ohio,
    "Oklahoma":     scrape_oklahoma,
    "Oregon":         scrape_oregon,
    "Pennsylvania":   scrape_pennsylvania,
    "South Dakota":   scrape_south_dakota,
    "Tennessee":      scrape_tennessee,
    "Texas":          scrape_texas,
    "Vermont":    scrape_vermont,
    "Virginia":   scrape_virginia,
    "Washington":  scrape_washington,
    "California":   scrape_california,
    "Colorado":     scrape_colorado,
    "Connecticut":  scrape_connecticut,
    "Delaware":     scrape_delaware,
    "Florida":      scrape_florida,
    "Georgia":      scrape_georgia,
    "Hawaii":       scrape_hawaii,
    "Idaho":        scrape_idaho,
    "Indiana":      scrape_indiana,
    "Iowa":         scrape_iowa,
    "Kansas":       scrape_kansas,
    "Kentucky":     scrape_kentucky,
    "Louisiana":    scrape_louisiana,
    "Maine":        scrape_maine,
    "Michigan":     scrape_michigan,
    "Minnesota":    scrape_minnesota,
    "Nevada":       scrape_nevada,
    "New Hampshire": scrape_new_hampshire,
    "New Jersey":   scrape_new_jersey,
    "New Mexico":   scrape_new_mexico,
    "Mississippi":  scrape_mississippi,
    "Missouri":     scrape_missouri,
    "Montana":      scrape_montana,
    "Nebraska":     scrape_nebraska,
    "New York":     scrape_new_york,
    "North Carolina": scrape_north_carolina,
    "North Dakota":   scrape_north_dakota,
    "Rhode Island":   scrape_rhode_island,
    "South Carolina": scrape_south_carolina,
    "Utah":           scrape_utah,
    "Arizona":        scrape_arizona,
    "Illinois":       scrape_illinois,
    "Massachusetts":  scrape_massachusetts,
}


# ── Runner ────────────────────────────────────────────────────────────────────

def run_all(
    states: list[str] | None = None,
    output_dir: str = ".",
    combined_filename: str | None = None,
) -> pd.DataFrame:
    """
    Run scrapers for all (or selected) states, save individual CSVs,
    and write a combined CSV ready for website import.

    Args:
        states:            List of state keys to run (default: all).
        output_dir:        Folder to write CSVs into.
        combined_filename: Override the combined CSV filename.

    Returns:
        Combined DataFrame.
    """
    today = date.today().strftime("%Y-%m-%d")
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    targets = states or list(SCRAPERS.keys())
    frames: list[pd.DataFrame] = []

    for name in targets:
        if name not in SCRAPERS:
            log.warning(f"No scraper registered for '{name}' — skipping.")
            continue
        try:
            df = SCRAPERS[name]()
            keep_cols = [c for c in OUTPUT_COLS if c in df.columns]
            frames.append(df[keep_cols].reset_index(drop=True))

            state_slug = name.lower().replace(" ", "_")
            path = out / f"warn_{state_slug}_{today}.csv"
            df.to_csv(path, index=False)
            log.info(f"  Saved {path}")

        except Exception as exc:
            log.error(f"  {name} failed: {exc}")

    if not frames:
        log.warning("No data collected.")
        return pd.DataFrame(columns=OUTPUT_COLS)

    combined = pd.concat(frames, ignore_index=True)
    # Replace NaN/None with empty strings so JSON serialisation produces valid JSON
    combined = combined.fillna("").astype(str).replace({"nan": "", "None": "", "NaT": ""})
    combined_path = out / (combined_filename or f"warn_combined_{today}.csv")
    combined.to_csv(combined_path, index=False)
    log.info(f"\nCombined file saved → {combined_path}  ({len(combined)} total rows)")

    # ── Master file — accumulates all daily results ───────────────────────────
    MASTER_FILE = out / "warn_master.csv"

    df_new = combined.copy()
    df_new['Date_Appended'] = today

    if MASTER_FILE.exists():
        df_master = pd.read_csv(MASTER_FILE, encoding='utf-8')
        df_master = pd.concat([df_master, df_new], ignore_index=True)
    else:
        df_master = df_new

    df_master = df_master.drop_duplicates(subset=['state', 'company', 'notice_date'])
    _sort_dt = pd.to_datetime(df_master['notice_date'], errors='coerce')
    df_master = df_master.iloc[_sort_dt.argsort()[::-1].values]
    df_master.to_csv(MASTER_FILE, index=False, encoding='utf-8')
    log.info(f"Master file updated → {MASTER_FILE}  ({len(df_master)} total rows)")

    # ── Mirror to master_file/ alongside other project master CSVs ───────────
    master_folder = Path("master_file")
    master_folder.mkdir(exist_ok=True)
    df_master.to_csv(master_folder / "warn_master.csv", index=False, encoding='utf-8')
    log.info(f"Master file mirrored → {master_folder / 'warn_master.csv'}")

    return combined


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json as _json, re as _re

    df = run_all(output_dir="warn_output")

    # Write warn_latest.json for the frontend
    payload = {
        "last_updated": date.today().strftime("%Y-%m-%d"),
        "total": len(df),
        "data": df.to_dict("records"),
    }
    raw_json = _json.dumps(payload, ensure_ascii=False)
    # Safety net: replace any bare NaN tokens that json.dumps may produce
    safe_json = _re.sub(r'(?<!["\w])NaN(?!["\w])', "null", raw_json)
    with open("warn_latest.json", "w", encoding="utf-8") as _jf:
        _jf.write(safe_json)
    log.info(f"warn_latest.json written: {len(df)} rows")

    print(f"\n{'='*60}")
    print(f"TOTAL ROWS: {len(df)}")
    print(f"STATES:     {sorted(df['state'].unique().tolist())}")
    print(f"{'='*60}\n")
    print(df.to_string(index=False, max_rows=30))