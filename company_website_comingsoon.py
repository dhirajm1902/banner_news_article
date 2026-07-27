"""
Company Website Grand Openings / Coming Soon Scraper
=====================================================
Scrapers:
  • ALDI           — https://www.aldi.us/grand-openings (Selenium)
  • Dogtopia       — REST API JSON endpoint (requests)
  • Burlington     — https://www.burlington.com/grand-openings (Selenium)
  • Five Below     — https://locations.fivebelow.com/coming-soon/index.html (requests)
  • HomeGoods      — https://www.homegoods.com/grand-openings (Selenium)
  • Homesense      — https://us.homesense.com/grand-openings (Selenium)
  • Jersey Mike's  — https://www.jerseymikes.com/locations/coming-soon (requests)
  • Chick-fil-A    — https://www.chick-fil-a.com/press-room/openings/list-view (requests)
  • LongHorn       — https://www.longhornsteakhouse.com/locations/new-locations (requests)
  • Trader Joe's   — https://www.traderjoes.com/home/announcements?category=store-openings (Playwright)
  • Target         — https://corporate.target.com/press/fact-sheet/2024/04/store-openings (requests)
  • Teso Life      — https://www.tesolife.com/en/stores (Selenium)
  • 7-Eleven       — https://www.7-eleven.com/lp/grand-openings (requests)
  • Marshalls      — https://www.marshalls.com/us/store/jump/topic/Grand-Openings/2600014 (Selenium)
  • Wawa           — https://www.wawa.com/about-us/public-relations/grand-openings (Playwright)
  • Kirkland's     — https://www.kirklands.com/content.jsp?pageName=openingstores (requests)
  • Yard House     — https://www.yardhouse.com/locations/new-locations (Selenium)
  • TJ Maxx        — https://tjmaxx.tjx.com/store/jump/topic/grand-openings/2600014 (Selenium)
  • Natural Grocers — https://www.naturalgrocers.com/new-store-announcements (Selenium)
  • Capital Grille  — https://www.thecapitalgrille.com/locations/new-locations (Selenium)
  • Hobby Lobby     — https://newsroom.hobbylobby.com/new-stores (requests)
  • Ulta Beauty     — https://www.ulta.com/guestservices/ways-to-shop/in-store/grand-openings (requests)
  • Costco          — https://www.costco.ca/f/-/new-locations (Patchright, Cloudflare-protected)

Output:
  docs/company_website_latest.json
  {
    "last_updated": "April 17, 2026 08:33 UTC",
    "data": [
      {
        "company": "ALDI",
        "address": "14600 Palm Beach Blvd., Fort Myers, FL 33905",
        "opening_date": "April 19, 2026",
        "link": "https://www.aldi.us/stores/..."
      }, ...
    ]
  }
"""

import re
import time
import json
import os
import asyncio
import requests
import pandas as pd
import nest_asyncio
from datetime import date, datetime, timezone
from pathlib import Path
from bs4 import BeautifulSoup, NavigableString, Tag
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from playwright.async_api import async_playwright

try:
    from playwright_stealth import stealth_async
    _TJ_STEALTH_MODE = "async"
except ImportError:
    try:
        from playwright_stealth import Stealth
        _TJ_STEALTH_MODE = "new"
    except ImportError:
        _TJ_STEALTH_MODE = "manual"

nest_asyncio.apply()


# ── Master-file helpers ───────────────────────────────────────────────────────

MASTER_PATH = os.path.join("docs", "company_website_master.json")


def _record_key(record: dict) -> str:
    """Stable identity key used to detect duplicates across runs."""
    company = (record.get("company") or "").lower().strip()
    address = re.sub(r"\s+", " ", (record.get("address") or "").lower().strip())
    return f"{company}|{address}"


def update_master(fresh_records: list[dict]) -> tuple[list[dict], int]:
    """
    Merge today's scraped records into the persistent master file.

    Rules:
      • Records already in the master  → is_new = False
      • Brand-new records              → is_new = True, first_seen = today (UTC)
      • If an existing record had no opening_date and we now have one, update it.

    Returns (merged_records, new_count).
    """
    today_str = datetime.now(timezone.utc).strftime("%B %d, %Y")

    existing: dict[str, dict] = {}
    if os.path.exists(MASTER_PATH):
        try:
            with open(MASTER_PATH, "r", encoding="utf-8") as f:
                master_data = json.load(f)
            for rec in master_data.get("data", []):
                rec["is_new"] = False          # clear flag from the previous run
                existing[_record_key(rec)] = rec
            print(f"[Master] Loaded {len(existing)} existing record(s) from {MASTER_PATH}")
        except Exception as e:
            print(f"[Master] Could not read existing master ({e}). Starting fresh.")

    new_count = 0
    for rec in fresh_records:
        key = _record_key(rec)
        if key in existing:
            if rec.get("opening_date") and not existing[key].get("opening_date"):
                existing[key]["opening_date"] = rec["opening_date"]
        else:
            rec = dict(rec)           # don't mutate the original
            rec["is_new"]     = True
            rec["first_seen"] = today_str
            existing[key]     = rec
            new_count        += 1

    merged = list(existing.values())
    return merged, new_count


# ── Shared helpers ────────────────────────────────────────────────────────────

DATE_RE = re.compile(
    r"\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
    r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{1,2},?\s+\d{4}\b"
    r"|\b\d{1,2}[/\-]\d{1,2}[/\-]\d{4}\b",
    re.IGNORECASE,
)


def extract_date(text: str) -> str:
    m = DATE_RE.search(text or "")
    return m.group(0).strip() if m else ""



def make_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
    )
    try:
        driver = webdriver.Chrome(options=options)
    except Exception:
        from selenium.webdriver.chrome.service import Service as ChromeService
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
        except Exception as e2:
            raise RuntimeError(f"Could not create Chrome driver: {e2}") from e2
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


# ── ALDI scraper ──────────────────────────────────────────────────────────────

ALDI_OPENINGS_URL = "https://www.aldi.us/grand-openings"
ALDI_BASE_URL = "https://www.aldi.us"
ALDI_STORE_LINK_RE = re.compile(r"/stores/l/[^/]+/[^/]+/[^/]+/[a-z0-9]+", re.IGNORECASE)


def _aldi_opening_date(driver: webdriver.Chrome, url: str) -> str:
    try:
        driver.get(url)
        time.sleep(3)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        for sel in ["[class*='grand-opening']", "[class*='grandOpening']",
                    "[class*='opening-date']", "[class*='openingDate']",
                    "[class*='store-opening']"]:
            tag = soup.select_one(sel)
            if tag:
                d = extract_date(tag.get_text(" ", strip=True))
                if d:
                    return d

        for tag in soup.find_all(string=re.compile(r"grand\s+opening", re.I)):
            d = extract_date(tag.parent.get_text(" ", strip=True))
            if d:
                return d

        full_text = soup.get_text(" ", strip=True)
        m = re.search(
            r"opening[^.]{0,80}?"
            r"(\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May"
            r"|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?"
            r"|Nov(?:ember)?|Dec(?:ember)?)\s+\d{1,2},?\s+\d{4}\b"
            r"|\b\d{1,2}[/\-]\d{1,2}[/\-]\d{4}\b)",
            full_text, re.IGNORECASE,
        )
        if m:
            return m.group(1).strip()
        return extract_date(full_text)
    except Exception as e:
        print(f"  ALDI date fetch error {url}: {e}")
        return ""


def scrape_aldi(driver: webdriver.Chrome) -> list[dict]:
    print(f"[ALDI] Loading {ALDI_OPENINGS_URL}")
    driver.get(ALDI_OPENINGS_URL)
    time.sleep(5)
    soup = BeautifulSoup(driver.page_source, "html.parser")

    seen, stores = set(), []
    for a in soup.find_all("a", href=ALDI_STORE_LINK_RE):
        href = a["href"].strip()
        full_link = href if href.startswith("http") else ALDI_BASE_URL + href
        if full_link in seen:
            continue
        seen.add(full_link)

        address_text = a.get_text(separator=" ", strip=True)
        m = re.match(r"^(.*?),\s*(.*?),\s*([A-Z]{2})\s*(\d{5}(?:-\d{4})?)$", address_text)
        if m:
            street, city, state, zip_code = m.groups()
            full_address = f"{street.strip(', ')}, {city.strip()}, {state.strip()} {zip_code.strip()}"
        else:
            full_address = address_text

        stores.append({"full_link": full_link, "address": full_address})

    print(f"[ALDI] Found {len(stores)} store(s). Fetching opening dates…")
    results = []
    for i, store in enumerate(stores, 1):
        print(f"  [{i}/{len(stores)}] {store['full_link']}")
        opening_date = _aldi_opening_date(driver, store["full_link"])
        print(f"         → {opening_date or '(not found)'}")
        results.append({
            "company":      "ALDI",
            "address":      store["address"],
            "opening_date": opening_date.strip(),
            "link":         store["full_link"],
        })
        time.sleep(1.0)
    return results


# ── Dogtopia scraper ──────────────────────────────────────────────────────────

DOGTOPIA_API = "https://www.dogtopia.com/wp-json/store-locator/v1/locations.json"


def scrape_dogtopia() -> list[dict]:
    print(f"[Dogtopia] Fetching {DOGTOPIA_API}")
    try:
        response = requests.get(DOGTOPIA_API, timeout=30)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        print(f"[Dogtopia] Error: {e}")
        return []

    results = []
    for loc in data:
        if not loc.get("opening_soon"):
            continue
        try:
            addr_info  = loc["store_info"]["location_address_info"][0]
            hours_info = loc["store_info"]["location_hours_info"][0]

            street   = addr_info.get("location_street_address", "") or ""
            city     = addr_info.get("location_city", "") or ""
            state    = addr_info.get("location_state_prov", "") or ""
            zipcode  = addr_info.get("location_zip_postal", "") or ""
            country  = addr_info.get("location_country", "") or ""
            address  = f"{street}, {city}, {state} {zipcode}, {country}".strip(", ")

            opening_date = hours_info.get("coming_soon_header_text", "") or ""
            link         = loc.get("link", "") or ""

            results.append({
                "company":      "Dogtopia",
                "address":      address,
                "opening_date": opening_date.strip(),
                "link":         link,
            })
        except (KeyError, IndexError, TypeError) as e:
            print(f"  [Dogtopia] Skipping malformed record: {e}")

    print(f"[Dogtopia] Found {len(results)} coming-soon location(s).")
    return results


# ── Burlington scraper ────────────────────────────────────────────────────────
#
# burlington.com is a Next.js app. The grand-openings list is not rendered as
# literal HTML (no state-name headings or <li>/<a> rows to scrape) — it's
# shipped as a JSON blob (`"storeOpenings": {...}`) inside a React Server
# Components flight payload (a `self.__next_f.push([...])` script). We pull
# that JSON out directly instead of walking the DOM.

BURLINGTON_URL = "https://www.burlington.com/grand-openings"


def _burlington_extract_store_openings(html: str) -> dict:
    idx = html.find('\\"storeOpenings\\"')
    if idx == -1:
        return {}
    brace_start = html.find("{", idx)
    if brace_start == -1:
        return {}

    depth = 0
    for i in range(brace_start, len(html)):
        c = html[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                raw = html[brace_start:i + 1]
                break
    else:
        return {}

    try:
        return json.loads(raw.replace('\\"', '"'))
    except json.JSONDecodeError as e:
        print(f"[Burlington] Failed to parse storeOpenings JSON: {e}")
        return {}


def _burlington_parse(html: str) -> list[dict]:
    store_openings = _burlington_extract_store_openings(html)
    print(f"[Burlington] Found {len(store_openings)} state(s) in storeOpenings data.")

    stores = []
    for state_name, entries in store_openings.items():
        for entry in entries:
            street = (entry.get("streetAddress") or "").strip()
            city   = (entry.get("city") or "").strip()
            date   = (entry.get("date") or "").strip()
            address_parts = [p for p in (street, city, state_name) if p]
            stores.append({
                "address":      ", ".join(address_parts),
                "opening_date": date,
            })
    return stores


def scrape_burlington(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Burlington] Loading {BURLINGTON_URL}")

    try:
        resp = requests.get(
            BURLINGTON_URL,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            timeout=20,
        )
        resp.raise_for_status()
        stores = _burlington_parse(resp.text)
    except Exception as e:
        print(f"[Burlington] Static fetch failed: {e}; falling back to Selenium.")
        driver.get(BURLINGTON_URL)
        time.sleep(8)
        stores = _burlington_parse(driver.page_source)

    results = [
        {
            "company":      "Burlington",
            "address":      s["address"],
            "opening_date": s["opening_date"],
            "link":         BURLINGTON_URL,
        }
        for s in stores
    ]
    print(f"[Burlington] {len(results)} store(s) parsed.")
    return results


# ── Five Below scraper ───────────────────────────────────────────────────────

FIVE_BELOW_INDEX_URL = "https://locations.fivebelow.com/coming-soon/index.html"
FIVE_BELOW_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36"
    )
}


def _five_below_get_json_ld(url: str) -> dict:
    resp = requests.get(url, headers=FIVE_BELOW_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    script = soup.find("script", type="application/ld+json")
    return json.loads(script.string)


def _five_below_format_address(address: dict) -> str:
    if not address:
        return ""
    parts = [
        address.get("streetAddress", "").strip(),
        address.get("addressLocality", "").strip(),
        address.get("addressRegion", "").strip(),
        address.get("postalCode", "").strip(),
    ]
    return ", ".join(p for p in parts if p)


def scrape_five_below() -> list[dict]:
    print(f"[Five Below] Fetching {FIVE_BELOW_INDEX_URL}")
    try:
        index_data = _five_below_get_json_ld(FIVE_BELOW_INDEX_URL)
        state_urls = [
            item["item"]["url"]
            for item in index_data["@graph"][0]["mainEntity"]["itemListElement"]
        ]
    except Exception as e:
        print(f"[Five Below] Error fetching index: {e}")
        return []

    results = []
    for state_url in state_urls:
        try:
            state_data = _five_below_get_json_ld(state_url)
            city_urls = [
                item["item"]["url"]
                for item in state_data["@graph"][0]["mainEntity"]["itemListElement"]
            ]
        except Exception as e:
            print(f"[Five Below] Error fetching state {state_url}: {e}")
            continue

        for city_url in city_urls:
            try:
                city_data = _five_below_get_json_ld(city_url)
                for entry in city_data["@graph"][0]["mainEntity"]["itemListElement"]:
                    store = entry["item"]
                    results.append({
                        "company":      "Five Below",
                        "address":      _five_below_format_address(store.get("address", {})),
                        "opening_date": "",
                        "link":         store.get("url", "") or "",
                    })
            except Exception as e:
                print(f"[Five Below] Error fetching city {city_url}: {e}")
                continue

    print(f"[Five Below] Found {len(results)} coming-soon location(s).")
    return results


# ── HomeGoods scraper ────────────────────────────────────────────────────────

HOMEGOODS_URL = "https://www.homegoods.com/grand-openings"
HOMEGOODS_BASE_URL = "https://www.homegoods.com"


def scrape_homegoods(driver: webdriver.Chrome) -> list[dict]:
    print(f"[HomeGoods] Loading {HOMEGOODS_URL}")
    driver.get(HOMEGOODS_URL)
    time.sleep(5)
    soup = BeautifulSoup(driver.page_source, "html.parser")

    results = []
    for state_li in soup.find_all("li", class_="state-dropdown"):
        state = state_li.find("h4").get_text(strip=True) if state_li.find("h4") else ""

        for store_li in state_li.find_all("li"):
            address_tag  = store_li.find("address")
            opening_tag  = store_li.find("h5")
            link_tag     = store_li.find("a", class_="arrow-link")

            address      = address_tag.get_text(separator=", ", strip=True) if address_tag else ""
            opening_date = opening_tag.get_text(strip=True) if opening_tag else ""
            href         = link_tag["href"] if link_tag and link_tag.get("href") else ""
            link         = (HOMEGOODS_BASE_URL + href) if href.startswith("/") else href

            if not address:
                continue

            results.append({
                "company":      "HomeGoods",
                "address":      f"{address}, {state}" if state else address,
                "opening_date": opening_date.strip(),
                "link":         link,
            })

    print(f"[HomeGoods] {len(results)} store(s) parsed.")
    return results


# ── Homesense scraper ────────────────────────────────────────────────────────

HOMESENSE_URL = "https://us.homesense.com/grand-openings"


def scrape_homesense(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Homesense] Loading {HOMESENSE_URL}")
    driver.get(HOMESENSE_URL)
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "locator-txt"))
        )
    except Exception:
        print("[Homesense] Timed out waiting for cards; trying anyway…")
        time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    cards = soup.find_all("div", class_="locator-txt")
    print(f"[Homesense] Found {len(cards)} store card(s).")

    results = []
    for card in cards:
        city_tag     = card.find("h2")
        date_tag     = card.find(class_="lime-txt")
        city         = city_tag.get_text(strip=True) if city_tag else ""
        opening_date = date_tag.get_text(strip=True) if date_tag else ""

        address_parts = []
        directions_url = ""
        for p in card.find_all("p"):
            if "lime-txt" in (p.get("class") or []):
                continue
            hit_hours = False
            for elem in p.children:
                if isinstance(elem, Tag):
                    if elem.name == "img":
                        hit_hours = True
                    elif elem.name == "a":
                        href = elem.get("href", "")
                        text = elem.get_text(strip=True)
                        if href.startswith("tel:"):
                            continue
                        elif "maps" in href or "Get Directions" in text:
                            directions_url = href
                elif isinstance(elem, NavigableString) and not hit_hours:
                    text = str(elem).strip().strip(",").strip()
                    if text:
                        address_parts.append(text)

        street  = ", ".join(p for p in address_parts if p)
        address = f"{street}, {city}" if street and city else street or city

        results.append({
            "company":      "Homesense",
            "address":      address,
            "opening_date": extract_date(opening_date),
            "link":         directions_url,
        })

    print(f"[Homesense] {len(results)} store(s) parsed.")
    return results


# ── Jersey Mike's scraper ────────────────────────────────────────────────────

JERSEY_MIKES_BASE_URL = "https://www.jerseymikes.com/locations/coming-soon"

JM_STATES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia",
}

JM_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
    )
}


def _jm_resolve(data: list, index):
    """Resolve a Vue.js serialised index back to its actual value."""
    if isinstance(index, int) and 0 <= index < len(data):
        return data[index]
    return index


def _jm_parse_payload(raw_data: list) -> list[dict]:
    """Parse store records from the Vue.js JSON payload embedded in the page."""
    stores = []
    try:
        if len(raw_data) <= 6:
            return stores
        for store_idx in raw_data[6]:
            if store_idx >= len(raw_data):
                continue
            store_dict = _jm_resolve(raw_data, store_idx)
            if not isinstance(store_dict, dict):
                continue
            store = {}
            for key in ("id", "number", "name", "openDate"):
                if key in store_dict:
                    store[key] = _jm_resolve(raw_data, store_dict[key])
            if "address" in store_dict:
                addr_raw = _jm_resolve(raw_data, store_dict["address"])
                if isinstance(addr_raw, dict):
                    store["address"] = {
                        k: _jm_resolve(raw_data, addr_raw[k])
                        for k in ("street1", "city", "subdivisionCode", "postalCode")
                        if k in addr_raw
                    }
                else:
                    store["address"] = {}
            stores.append(store)
    except Exception as e:
        print(f"  [Jersey Mike's] Payload parse error: {e}")
    return stores


def scrape_jersey_mikes() -> list[dict]:
    print(f"[Jersey Mike's] Scraping {JERSEY_MIKES_BASE_URL} …")
    results = []

    for state_code, state_name in JM_STATES.items():
        url = f"{JERSEY_MIKES_BASE_URL}/US-{state_code}"
        try:
            resp = requests.get(url, headers=JM_HEADERS, timeout=15)
            resp.raise_for_status()
        except Exception:
            continue

        soup = BeautifulSoup(resp.text, "html.parser")
        script = soup.find("script", type="application/json")
        if not (script and script.string):
            continue

        try:
            raw_data = json.loads(script.string)
        except Exception:
            continue

        stores = _jm_parse_payload(raw_data)
        for s in stores:
            addr     = s.get("address") or {}
            street   = addr.get("street1", "") or ""
            city     = addr.get("city", "") or ""
            zip_code = addr.get("postalCode", "") or ""
            address_parts = [p for p in [street, city, f"{state_code} {zip_code}".strip()] if p]
            address  = ", ".join(address_parts) if address_parts else state_name

            open_date = s.get("openDate", "") or ""

            results.append({
                "company":      "Jersey Mike's",
                "address":      address,
                "opening_date": extract_date(open_date) or "In Development",
                "link":         url,
            })

        time.sleep(0.5)

    print(f"[Jersey Mike's] Found {len(results)} coming-soon location(s).")
    return results


# ── Chick-fil-A scraper ──────────────────────────────────────────────────────

CHICK_FIL_A_BASE_URL = "https://www.chick-fil-a.com/press-room/openings/list-view"
CHICK_FIL_A_PAGES    = 5
CHICK_FIL_A_HEADERS  = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}


def _cfa_fetch_page(page: int) -> str:
    url = CHICK_FIL_A_BASE_URL if page == 1 else f"{CHICK_FIL_A_BASE_URL}?query-0-page={page}"
    resp = requests.get(url, headers=CHICK_FIL_A_HEADERS, timeout=15)
    resp.raise_for_status()
    return resp.text


def _cfa_parse_list(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    rows = []
    for h3 in soup.find_all("h3"):
        a = h3.find("a", href=True)
        if not a:
            continue
        href = a["href"]
        if "/press-room/" not in href or "openings" in href:
            continue
        rows.append({"url": href})
    return rows


def _cfa_fetch_article(url: str) -> tuple[str, str]:
    try:
        resp = requests.get(url, headers=CHICK_FIL_A_HEADERS, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [Chick-fil-A] Could not fetch {url}: {e}")
        return "", ""

    soup = BeautifulSoup(resp.text, "html.parser")
    body_text = " ".join(p.get_text(" ", strip=True) for p in soup.find_all("p"))

    address = ""
    addr_m = re.search(
        r"[Ll]ocated at\s+([\d]+[^,\.]+(?:Ave|Blvd|Dr|Rd|St|Way|Ln|Pkwy|Hwy|"
        r"Pike|Plaza|Circle|Cir|Court|Ct|Trail|Trl|Route|Rt|Square|Sq|"
        r"Drive|Street|Road|Lane|Parkway|Highway)[^,\.]*)",
        body_text,
    )
    if addr_m:
        address = addr_m.group(1).strip()

    opening_date = ""
    date_m = re.search(
        r"(?:begin serving|opening|opens?|open(?:ing)? its doors)[^.]{0,80}"
        r"on\s+(?:\w+day,\s*)?"
        r"((?:January|February|March|April|May|June|July|August|September|"
        r"October|November|December)\s+\d{1,2},?\s+\d{4})",
        body_text,
        re.IGNORECASE,
    )
    if date_m:
        opening_date = date_m.group(1).strip()
    else:
        h1 = soup.find("h1")
        if h1:
            container = h1.find_parent()
            if container:
                pub_m = re.search(
                    r"((?:January|February|March|April|May|June|July|August|"
                    r"September|October|November|December)\s+\d{1,2},?\s+\d{4})",
                    container.get_text(" ", strip=True),
                )
                if pub_m:
                    opening_date = pub_m.group(1).strip()

    return address, opening_date


def scrape_chick_fil_a() -> list[dict]:
    print(f"[Chick-fil-A] Scraping {CHICK_FIL_A_BASE_URL} ({CHICK_FIL_A_PAGES} pages)…")
    listings = []
    for page in range(1, CHICK_FIL_A_PAGES + 1):
        try:
            html = _cfa_fetch_page(page)
            rows = _cfa_parse_list(html)
            listings.extend(rows)
            print(f"  Page {page}: {len(rows)} record(s) (total: {len(listings)})")
        except Exception as e:
            print(f"  [Chick-fil-A] Page {page} error: {e}")
        time.sleep(0.5)

    results = []
    for i, row in enumerate(listings, 1):
        print(f"  [{i}/{len(listings)}] {row['url']}")
        address, opening_date = _cfa_fetch_article(row["url"])
        results.append({
            "company":      "Chick-fil-A",
            "address":      address,
            "opening_date": opening_date,
            "link":         row["url"],
        })
        time.sleep(0.5)

    print(f"[Chick-fil-A] Found {len(results)} opening(s).")
    return results


# ── LongHorn Steakhouse scraper ──────────────────────────────────────────────

LONGHORN_URL = "https://www.longhornsteakhouse.com/locations/new-locations"
LONGHORN_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
    )
}


def scrape_longhorn() -> list[dict]:
    print(f"[LongHorn] Fetching {LONGHORN_URL}")
    try:
        resp = requests.get(LONGHORN_URL, headers=LONGHORN_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"[LongHorn] Error: {e}")
        return []

    soup = BeautifulSoup(resp.content, "html.parser")
    content = soup.find_all("div", class_="newloc row")

    results = []
    for data in content:
        address_div = data.find("div", class_="span7 margtop3")
        if not address_div:
            continue

        parts = address_div.get_text(separator="\n", strip=True).split("\n")
        street     = parts[1] if len(parts) > 1 else ""
        city_state = parts[2] if len(parts) > 2 else ""
        field3     = parts[3] if len(parts) > 3 else ""
        field4     = parts[4] if len(parts) > 4 else ""

        # field3 is a phone number when it starts with '(' or a digit;
        # for "OPENING SPRING 2026" rows there is no phone — field3 is the date.
        is_phone     = field3.startswith("(") or (bool(field3) and field3[0].isdigit())
        opening_date = field4 if is_phone else field3

        address = f"{street}, {city_state}".strip(", ")
        if not address:
            continue

        results.append({
            "company":      "LongHorn Steakhouse",
            "address":      address,
            "opening_date": extract_date(opening_date) or opening_date,
            "link":         LONGHORN_URL,
        })

    print(f"[LongHorn] {len(results)} location(s) parsed.")
    return results


# ── Trader Joe's scraper ─────────────────────────────────────────────────────

TJ_BASE_URL      = "https://www.traderjoes.com/home/announcements?category=store-openings"
TJ_PAGES         = 5
TJ_PAGE_LOAD_MS  = 4000
TJ_CLICK_WAIT_MS = 3500


def _tj_parse_opening_date(text: str) -> str:
    # TJ articles use "Date & Time of Opening:" (the "of" was missing before)
    m = re.search(
        r"Date\s*&\s*Time\s*(?:of\s*)?Opening\s*[:\-]?\s*\n?\s*([^\n]{3,80})",
        text, re.IGNORECASE,
    )
    if m:
        val = m.group(1).strip()
        return val if val else "TBD"
    # Fallback: any full date near an "open" keyword
    m2 = re.search(
        r"open(?:ing|s)?[^.]{0,80}?"
        r"((?:January|February|March|April|May|June|July|August|September|"
        r"October|November|December)\s+\d{1,2},?\s+\d{4})",
        text, re.IGNORECASE,
    )
    if m2:
        return m2.group(1).strip()
    return extract_date(text)


def _tj_parse_address(text: str) -> str:
    # "Store Location:\n123 Main St\nCity, State ZIP"
    m = re.search(
        r"Store\s*Location\s*[:\-]?\s*\n([^\n]+)\n([^\n]+)",
        text, re.IGNORECASE,
    )
    if m:
        return f"{m.group(1).strip()}, {m.group(2).strip()}"
    m2 = re.search(
        r"Store\s*Location\s*[:\-]?\s*\n?\s*([^\n]{5,120})",
        text, re.IGNORECASE,
    )
    return m2.group(1).strip() if m2 else ""


async def _tj_apply_stealth(page) -> None:
    if _TJ_STEALTH_MODE == "async":
        await stealth_async(page)
    elif _TJ_STEALTH_MODE == "new":
        stealth = Stealth()
        await stealth.apply_stealth_async(page)
    else:
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1, 2, 3] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
            window.chrome = { runtime: {} };
        """)


async def _tj_dismiss_overlays(page) -> None:
    for sel in ["button:has-text('GOT IT')", "button:has-text('Got it')",
                "button[aria-label='Close dialog']", ".klaviyo-close-form"]:
        try:
            btn = await page.query_selector(sel)
            if btn:
                await btn.click()
                await page.wait_for_timeout(800)
        except Exception:
            pass


async def _tj_collect_article_urls(page) -> list[str]:
    """Phase 1: walk listing pages and return all article URLs."""
    seen, urls = set(), []
    current_page = 1
    while True:
        if current_page == 1:
            await page.goto(TJ_BASE_URL, wait_until="load")
        await page.wait_for_timeout(TJ_PAGE_LOAD_MS)
        await _tj_dismiss_overlays(page)

        # Wait for React to render the article cards before querying
        try:
            await page.wait_for_selector("a[href*='/home/announcements']", timeout=12000)
        except Exception:
            print(f"[Trader Joe's] Page {current_page}: timed out waiting for article cards.")

        # Broaden to href*='/home/announcements' (no trailing slash required)
        anchors = await page.query_selector_all("a[href*='/home/announcements']")
        page_count = 0
        for a in anchors:
            href = (await a.get_attribute("href") or "").strip()
            # Skip the category listing URL and bare /home/announcements
            if not href or "?category" in href:
                continue
            bare = href.rstrip("/").split("?")[0]
            if bare in ("/home/announcements", "https://www.traderjoes.com/home/announcements"):
                continue
            full = href if href.startswith("http") else f"https://www.traderjoes.com{href}"
            if full not in seen:
                seen.add(full)
                urls.append(full)
                page_count += 1

        print(f"[Trader Joe's] Listing page {current_page}: {page_count} new link(s) (total: {len(urls)})")

        if TJ_PAGES and current_page >= TJ_PAGES:
            print(f"[Trader Joe's] Reached page limit ({TJ_PAGES}).")
            break

        # Try each next-page selector separately (comma-groups can silently miss)
        next_btn = None
        for sel in [
            f"li[aria-label='Go to page {current_page + 1}']",
            f"button[aria-label='Go to page {current_page + 1}']",
            "button[class*='arrow'][class*='right']:not([disabled])",
            "button[class*='next']:not([disabled])",
            "a[aria-label='Next page']",
            "button[aria-label='Next page']",
        ]:
            next_btn = await page.query_selector(sel)
            if next_btn:
                break

        if not next_btn:
            print("[Trader Joe's] No more listing pages.")
            break
        await next_btn.click()
        await page.wait_for_timeout(TJ_PAGE_LOAD_MS)
        current_page += 1
    return urls


async def _tj_scrape() -> list[dict]:
    results = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="America/New_York",
        )
        page = await context.new_page()
        await _tj_apply_stealth(page)

        # Phase 1: collect all article URLs from listing pages
        article_urls = await _tj_collect_article_urls(page)

        # Phase 2: visit each article and extract details
        for i, url in enumerate(article_urls, 1):
            try:
                print(f"  [{i}/{len(article_urls)}] {url}")
                await page.goto(url, wait_until="load")
                await page.wait_for_timeout(3000)
                await _tj_dismiss_overlays(page)
                body = await page.inner_text("body")
                opening = _tj_parse_opening_date(body)
                address = _tj_parse_address(body)
                print(f"    → {opening or '(no date)'} | {address or '(no address)'}")
                results.append({
                    "company":      "Trader Joe's",
                    "address":      address,
                    "opening_date": opening,
                    "link":         url,
                })
            except Exception as e:
                print(f"    ERROR: {e}")

        await browser.close()
    return results


def _tj_extract_slugs(obj, depth: int = 0) -> list[str]:
    """Recursively pull article slugs out of __NEXT_DATA__ JSON."""
    if depth > 12:
        return []
    slugs: list[str] = []
    if isinstance(obj, dict):
        slug = obj.get("slug") or obj.get("url") or obj.get("path")
        t    = (obj.get("__typename") or obj.get("type") or "").lower()
        # Accept any object that has a slug and looks like content (article/post/announcement/etc.)
        CONTENT_TYPES = ("announcement", "article", "post", "story", "entry", "content", "page")
        if slug and isinstance(slug, str) and any(ct in t for ct in CONTENT_TYPES):
            slugs.append(slug.lstrip("/"))
        # Also capture if the slug appears under an "announcements"-keyed list
        for k, v in obj.items():
            if k.lower() in ("announcements", "articles", "posts", "items", "entries", "nodes"):
                if isinstance(v, list):
                    for item in v:
                        if isinstance(item, dict):
                            s = item.get("slug") or item.get("url") or item.get("path")
                            if s and isinstance(s, str):
                                slugs.append(s.lstrip("/"))
            slugs.extend(_tj_extract_slugs(v, depth + 1))
    elif isinstance(obj, list):
        for item in obj:
            slugs.extend(_tj_extract_slugs(item, depth + 1))
    return list(dict.fromkeys(slugs))  # dedupe, preserve order


def _tj_requests_scrape() -> list[dict]:
    """Primary: plain requests + __NEXT_DATA__ extraction (no JS engine needed)."""
    _hdrs = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Cache-Control":   "no-cache",
    }
    session = requests.Session()
    session.headers.update(_hdrs)

    seen: set[str]    = set()
    article_urls: list[str] = []

    for pg in range(1, TJ_PAGES + 1):
        url = TJ_BASE_URL if pg == 1 else f"{TJ_BASE_URL}&page={pg}"
        try:
            resp = session.get(url, timeout=30)
            print(f"[Trader Joe's] Page {pg} HTTP {resp.status_code}, {len(resp.text)} chars")
            soup = BeautifulSoup(resp.text, "html.parser")

            # 1. Try __NEXT_DATA__ (most reliable for Next.js)
            script_tag = soup.find("script", {"id": "__NEXT_DATA__"})
            if script_tag and script_tag.string:
                nd = json.loads(script_tag.string)
                slugs = _tj_extract_slugs(nd)
                print(f"[Trader Joe's] __NEXT_DATA__ slugs found: {len(slugs)}")
                for slug in slugs:
                    full = f"https://www.traderjoes.com/home/announcements/{slug}"
                    if full not in seen:
                        seen.add(full); article_urls.append(full)
            else:
                print(f"[Trader Joe's] No __NEXT_DATA__ script on page {pg}")

            # 2. Fallback: raw <a> link scan
            before = len(article_urls)
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if "/home/announcements/" not in href or "?category" in href:
                    continue
                full = href if href.startswith("http") else f"https://www.traderjoes.com{href}"
                if full not in seen:
                    seen.add(full); article_urls.append(full)
            print(f"[Trader Joe's] <a> scan added {len(article_urls) - before} link(s)")

        except Exception as e:
            print(f"[Trader Joe's] Requests listing page {pg}: {e}")

    print(f"[Trader Joe's] Requests found {len(article_urls)} article URL(s).")
    results: list[dict] = []
    for url in article_urls:
        try:
            resp   = session.get(url, timeout=20)
            soup   = BeautifulSoup(resp.text, "html.parser")
            body   = soup.get_text("\n")
            opening = _tj_parse_opening_date(body)
            address = _tj_parse_address(body)
            print(f"  → {opening or '(no date)'} | {address or '(no address)'}")
            results.append({
                "company":      "Trader Joe's",
                "address":      address,
                "opening_date": opening,
                "link":         url,
            })
        except Exception as e:
            print(f"[Trader Joe's] Article {url}: {e}")
    return results


def scrape_trader_joes() -> list[dict]:
    print(f"[Trader Joe's] Scraping {TJ_BASE_URL} ({TJ_PAGES} pages)…")
    # Try lightweight requests approach first (reliable in CI/CD)
    try:
        results = _tj_requests_scrape()
        if results:
            print(f"[Trader Joe's] Found {len(results)} opening(s) via requests.")
            return results
        print("[Trader Joe's] Requests returned 0 — falling back to Playwright…")
    except Exception as e:
        print(f"[Trader Joe's] Requests failed ({e}) — falling back to Playwright…")
    # Playwright fallback
    try:
        results = asyncio.run(_tj_scrape())
    except Exception as e:
        print(f"[Trader Joe's] Playwright error: {e}")
        return []
    print(f"[Trader Joe's] Found {len(results)} opening(s) via Playwright.")
    return results


# ── Wawa scraper ─────────────────────────────────────────────────────────────

WAWA_BASE_URL     = "https://www.wawa.com/about-us/public-relations/grand-openings"
WAWA_PAGE_LOAD_MS = 4000


async def _wawa_dismiss_popups(page) -> None:
    for sel in ["button:has-text('Accept')", "button:has-text('Accept All')",
                "button:has-text('Got it')", "[aria-label='Close']"]:
        try:
            btn = await page.query_selector(sel)
            if btn:
                await btn.click()
                await page.wait_for_timeout(800)
                return
        except Exception:
            pass


async def _wawa_scrape() -> list[dict]:
    results = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="America/New_York",
        )
        page = await context.new_page()
        await _tj_apply_stealth(page)  # reuse shared stealth helper

        await page.goto(WAWA_BASE_URL, wait_until="networkidle")
        await page.wait_for_timeout(WAWA_PAGE_LOAD_MS)
        await _wawa_dismiss_popups(page)

        current_page = 1
        while True:
            print(f"[Wawa] Page {current_page} …")
            try:
                await page.wait_for_selector("p.md\\:text-lg", timeout=8000)
            except Exception:
                print("[Wawa] Timed out waiting for store entries.")

            store_paras = await page.query_selector_all("p.md\\:text-lg")
            print(f"  {len(store_paras)} entry/entries found.")

            for para in store_paras:
                try:
                    full_text = (await para.inner_text()).strip()
                    if "Opening Date:" not in full_text:
                        continue

                    addr_m = re.search(
                        r"Store\s+\d+\s*\|\s*(.+?)(?:\n|Opening Date:)",
                        full_text, re.IGNORECASE | re.DOTALL,
                    )
                    address = addr_m.group(1).strip() if addr_m else ""

                    date_m = re.search(
                        r"Opening\s+Date\s*[:\-]?\s*([^\n]+)",
                        full_text, re.IGNORECASE,
                    )
                    opening_date = date_m.group(1).strip() if date_m else ""

                    link_el = await para.query_selector("a")
                    href = ""
                    if link_el:
                        href = await link_el.get_attribute("href") or ""
                        if href.startswith("/"):
                            href = f"https://www.wawa.com{href}"

                    results.append({
                        "company":      "Wawa",
                        "address":      address,
                        "opening_date": extract_date(opening_date) or opening_date,
                        "link":         href,
                    })
                except Exception as e:
                    print(f"  [Wawa] Entry error: {e}")

            next_btn = (
                await page.query_selector("a[aria-label='Next page']")
                or await page.query_selector("button[aria-label='Next page']")
                or await page.query_selector("a:has-text('Next')")
                or await page.query_selector("button:has-text('Next')")
                or await page.query_selector(f"a[aria-label='page {current_page + 1}']")
                or await page.query_selector(f"li[aria-label='Go to page {current_page + 1}']")
            )
            if not next_btn:
                print("[Wawa] No more pages.")
                break
            await next_btn.click()
            await page.wait_for_timeout(WAWA_PAGE_LOAD_MS)
            await _wawa_dismiss_popups(page)
            current_page += 1

        await browser.close()
    return results


def scrape_wawa() -> list[dict]:
    print(f"[Wawa] Scraping {WAWA_BASE_URL} …")
    try:
        results = asyncio.run(_wawa_scrape())
    except Exception as e:
        print(f"[Wawa] Error: {e}")
        return []
    print(f"[Wawa] Found {len(results)} opening(s).")
    return results


# ── Target scraper ────────────────────────────────────────────────────────────

TARGET_URL = "https://corporate.target.com/press/fact-sheet/2024/04/store-openings"
TARGET_BASE_URL = "https://corporate.target.com"
TARGET_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}
TARGET_JUNK_RE = re.compile(
    r"(click to collapse|click to expand|Target in \w+|Opened in \w+)",
    re.IGNORECASE,
)


def _target_extract_date(sentence: str) -> str:
    m = re.search(
        r"(January|February|March|April|May|June|July|August|"
        r"September|October|November|December)\.?\s+\d{1,2}",
        sentence, re.IGNORECASE,
    )
    return m.group(0).replace(".", "") if m else sentence


def scrape_target() -> list[dict]:
    print(f"[Target] Fetching {TARGET_URL}")
    try:
        resp = requests.get(TARGET_URL, headers=TARGET_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"[Target] Error: {e}")
        return []

    soup = BeautifulSoup(resp.content, "html.parser")
    content_area = (
        soup.find("div", class_="field-items")
        or soup.find("article")
        or soup.body
    )

    results = []
    current_opening_date = None

    for tag in content_area.find_all(["p", "h3"]):
        text = tag.get_text(" ", strip=True)

        if "officially open" in text.lower():
            current_opening_date = _target_extract_date(text)
            continue

        if tag.name == "h3":
            if TARGET_JUNK_RE.search(text):
                continue

            store_name = text
            address_tag = tag.find_next_sibling("p")
            address = address_tag.get_text(" ", strip=True) if address_tag else ""
            if not address:
                continue

            anchor_id = tag.get("id") or (
                store_name.lower()
                .replace(" ", "-")
                .replace(":", "")
                .replace(",", "")
                .replace("(", "")
                .replace(")", "")
            )
            results.append({
                "company":      "Target",
                "address":      f"{store_name}, {address}",
                "opening_date": current_opening_date or "",
                "link":         f"{TARGET_URL}#{anchor_id}",
            })

    print(f"[Target] {len(results)} store(s) parsed.")
    return results


# ── Teso Life scraper ────────────────────────────────────────────────────────

TESO_URL = "https://www.tesolife.com/en/stores"
TESO_BASE_URL = "https://www.tesolife.com"


def _teso_extract_opening_date(store_tag) -> str:
    full_text = store_tag.get_text(" ", strip=True)
    date_tag = store_tag.find(class_=lambda c: c and any(
        x in c.lower() for x in ["date", "opening", "coming", "soon", "badge", "label"]
    ))
    if date_tag:
        return date_tag.get_text(strip=True)
    if "coming soon" in full_text.lower():
        return "Coming Soon"
    date_match = re.search(
        r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4}'
        r'|\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4}'
        r'|\d{1,2}[\/\-]\d{4}',
        full_text, re.IGNORECASE,
    )
    if date_match:
        return date_match.group()
    return "Open"


def scrape_teso_life(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Teso Life] Loading {TESO_URL}")
    driver.get(TESO_URL)
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, ".store-list-item, .sl-state-group, [class*='store']")
            )
        )
    except Exception:
        print("[Teso Life] Timed out waiting for store elements; trying anyway…")
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    results = []

    state_groups = soup.find_all("div", class_="sl-state-group")
    if state_groups:
        print(f"[Teso Life] Found {len(state_groups)} state group(s).")
        for group in state_groups:
            state_tag = group.find("h5") or group.find("h4") or group.find("h3")
            state_name = state_tag.get_text(strip=True) if state_tag else ""

            store_items = group.find_all("a", class_="store-list-item-action")
            if not store_items:
                store_items = group.find_all(
                    ["a", "div"], class_=lambda c: c and "store" in c.lower()
                )

            for store in store_items:
                opening_date = _teso_extract_opening_date(store)
                if opening_date == "Open":
                    continue

                addr_tag = store.find("p") or store.find(
                    "span", class_=lambda c: c and "address" in c.lower()
                )
                address = addr_tag.get_text(strip=True) if addr_tag else ""
                if state_name:
                    address = f"{address}, {state_name}" if address else state_name

                href = store.get("href", "")
                if href and not href.startswith("http"):
                    href = TESO_BASE_URL + href

                results.append({
                    "company":      "Teso Life",
                    "address":      address,
                    "opening_date": opening_date,
                    "link":         href,
                })
    else:
        print("[Teso Life] No state groups found, trying flat store list…")
        store_items = soup.find_all(
            lambda tag: tag.name in ["a", "div", "li"]
            and any("store" in c.lower() for c in tag.get("class", []))
        )
        print(f"[Teso Life] Found {len(store_items)} store item(s).")

        for store in store_items:
            opening_date = _teso_extract_opening_date(store)
            if opening_date == "Open":
                continue

            href = store.get("href", "")
            if href and not href.startswith("http"):
                href = TESO_BASE_URL + href

            results.append({
                "company":      "Teso Life",
                "address":      store.get_text(" ", strip=True)[:200],
                "opening_date": opening_date,
                "link":         href,
            })

    print(f"[Teso Life] {len(results)} coming-soon/upcoming store(s) found.")
    return results


# ── 7-Eleven scraper ─────────────────────────────────────────────────────────

SEVEN_ELEVEN_URL = "https://www.7-eleven.com/lp/grand-openings"
SEVEN_ELEVEN_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
}


def scrape_seven_eleven() -> list[dict]:
    print(f"[7-Eleven] Fetching {SEVEN_ELEVEN_URL}")
    try:
        resp = requests.get(SEVEN_ELEVEN_URL, headers=SEVEN_ELEVEN_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"[7-Eleven] Error: {e}")
        return []

    soup = BeautifulSoup(resp.content, "html.parser")
    items = soup.find_all("li", class_="item")
    print(f"[7-Eleven] Found {len(items)} item(s).")

    results = []
    for item in items:
        try:
            # Partial class match guards against hashed Next.js suffixes (e.g. Heading_h6__OzLeX)
            h2 = item.find("h2", class_=lambda c: c and any("Heading_h6" in v for v in c))
            h4 = item.find("h4", class_=lambda c: c and any("Heading_h4" in v for v in c))
            h2 = h2 or item.find("h2")
            h4 = h4 or item.find("h4")

            address1 = h2.get_text(strip=True) if h2 else ""
            address2 = h4.get_text(strip=True) if h4 else ""
            address = f"{address1} {address2}".strip()
            if not address:
                continue

            p_tag = item.find("p")
            opening_date_raw = p_tag.get_text(strip=True) if p_tag else ""

            cta = item.find("div", class_="cta")
            link = ""
            if cta:
                a_tag = cta.find("a")
                if a_tag:
                    link = a_tag.get("href", "")

            results.append({
                "company":      "7-Eleven",
                "address":      address,
                "opening_date": extract_date(opening_date_raw) or opening_date_raw,
                "link":         link,
            })
        except Exception as e:
            print(f"  [7-Eleven] Error parsing item: {e}")

    print(f"[7-Eleven] {len(results)} store(s) parsed.")
    return results


# ── Marshalls scraper ────────────────────────────────────────────────────────

MARSHALLS_URL = "https://www.marshalls.com/us/store/jump/topic/Grand-Openings/2600014"
MARSHALLS_BASE_URL = "https://www.marshalls.com"


def scrape_marshalls(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Marshalls] Loading {MARSHALLS_URL}")
    driver.get(MARSHALLS_URL)
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    results = []

    for box in soup.find_all("section", class_="subsection"):
        h2 = box.find("h2")
        if not h2:
            continue
        state = h2.get_text(strip=True)

        for item in box.find_all("li", class_=lambda c: c and "store-list-item" in c):
            h3 = item.find("h3", class_=lambda c: c and "address-heading" in c)
            city = h3.get_text(strip=True) if h3 else ""

            street_div = item.find("div", class_="street-address")
            street = street_div.get_text(strip=True) if street_div else ""

            address_parts = [p for p in [street, city, state] if p]
            full_address = ", ".join(address_parts)

            strong = item.find("strong")
            opening_date = strong.get_text(strip=True) if strong else ""
            if not opening_date:
                continue

            link_tag = item.find("a", href=lambda h: h and "directions.jsp" in h)
            href = ""
            if link_tag:
                href = link_tag["href"]
                if not href.startswith("http"):
                    href = MARSHALLS_BASE_URL + href

            results.append({
                "company":      "Marshalls",
                "address":      full_address,
                "opening_date": extract_date(opening_date) or opening_date,
                "link":         href,
            })

    print(f"[Marshalls] {len(results)} store(s) parsed.")
    return results


# ── TJ Maxx scraper ─────────────────────────────────────────────────────────

TJMAXX_URL      = "https://tjmaxx.tjx.com/store/jump/topic/grand-openings/2600014"
TJMAXX_BASE_URL = "https://tjmaxx.tjx.com"


def scrape_tjmaxx(driver: webdriver.Chrome) -> list[dict]:
    print(f"[TJ Maxx] Loading {TJMAXX_URL}")
    driver.get(TJMAXX_URL)
    time.sleep(8)  # JS-heavy TJX page needs extra render time

    soup = BeautifulSoup(driver.page_source, "html.parser")
    results = []

    # Primary: TJX grand-openings list items
    items = soup.select("li.grandopenings-store")

    # Fallback: same subsection/store-list-item structure used by Marshalls
    if not items:
        print("[TJ Maxx] Primary selector found 0 items — trying subsection fallback.")
        for box in soup.find_all("section", class_="subsection"):
            h2 = box.find("h2")
            state = h2.get_text(strip=True) if h2 else ""
            for item in box.find_all("li", class_=lambda c: c and "store-list-item" in c):
                h3 = item.find("h3", class_=lambda c: c and "address-heading" in c)
                city = h3.get_text(strip=True) if h3 else ""
                street_div = item.find("div", class_="street-address")
                street = street_div.get_text(strip=True) if street_div else ""
                address_parts = [p for p in [street, city, state] if p]
                strong = item.find("strong")
                opening_date = strong.get_text(strip=True) if strong else ""
                if not opening_date:
                    continue
                link_tag = item.find("a", href=lambda h: h and "directions.jsp" in h)
                href = ""
                if link_tag:
                    href = link_tag["href"]
                    if not href.startswith("http"):
                        href = TJMAXX_BASE_URL + href
                results.append({
                    "company":      "TJ Maxx",
                    "address":      ", ".join(address_parts),
                    "opening_date": extract_date(opening_date) or opening_date,
                    "link":         href or TJMAXX_URL,
                })
        print(f"[TJ Maxx] {len(results)} store(s) parsed (fallback).")
        return results

    print(f"[TJ Maxx] Found {len(items)} store item(s).")
    for item in items:
        address_tag = item.select_one("div.street-address")
        city_tag    = item.find("h3", class_=lambda c: c and "address-heading" in c)
        strong      = item.find("strong")

        street       = address_tag.get_text(strip=True) if address_tag else ""
        city         = city_tag.get_text(strip=True) if city_tag else ""
        opening_date = strong.get_text(strip=True) if strong else ""

        address_parts = [p for p in [street, city] if p]
        if not address_parts:
            continue

        link_tag = item.find("a", href=lambda h: h and "directions.jsp" in h)
        href = ""
        if link_tag:
            href = link_tag["href"]
            if not href.startswith("http"):
                href = TJMAXX_BASE_URL + href

        results.append({
            "company":      "TJ Maxx",
            "address":      ", ".join(address_parts),
            "opening_date": extract_date(opening_date) or opening_date,
            "link":         href or TJMAXX_URL,
        })

    print(f"[TJ Maxx] {len(results)} store(s) parsed.")
    return results


# ── Kirkland's scraper ───────────────────────────────────────────────────────

KIRKLANDS_URL = "https://www.kirklands.com/content.jsp?pageName=openingstores&icid=storelocator_new"
KIRKLANDS_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,"
        "image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.kirklands.com/",
    "Connection": "keep-alive",
}


def scrape_kirklands() -> list[dict]:
    print(f"[Kirkland's] Fetching {KIRKLANDS_URL}")
    try:
        resp = requests.get(KIRKLANDS_URL, headers=KIRKLANDS_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"[Kirkland's] Error: {e}")
        return []

    soup = BeautifulSoup(resp.content, "html.parser")
    results = []

    # Try multiple selector strategies (site HTML may vary)
    SELECTOR_CANDIDATES = [
        "div.storeInfo",
        "div[class*='storeInfo']",
        "div[class*='store-info']",
        "div[class*='store_info']",
        "li[class*='store']",
        "div[class*='opening']",
        "td[class*='store']",
    ]
    boxes = []
    for sel in SELECTOR_CANDIDATES:
        boxes = soup.select(sel)
        if boxes:
            print(f"[Kirkland's] Matched selector: {sel} ({len(boxes)} items)")
            break

    if not boxes:
        # Last resort: scan all <li> / <p> elements near an "opening" heading
        headings = soup.find_all(
            lambda t: t.name in ("h1", "h2", "h3", "h4") and
                      ("opening" in t.get_text(strip=True).lower() or
                       "coming soon" in t.get_text(strip=True).lower())
        )
        for h in headings:
            parent = h.find_parent(["section", "div", "table"]) or h.parent
            boxes.extend(parent.find_all(["li", "p", "tr"]))
        print(f"[Kirkland's] Fallback heading scan: {len(boxes)} candidate element(s)")

    for box in boxes:
        address = box.get_text(separator=" ", strip=True)
        if not address or len(address) < 8:
            continue
        opening_date = extract_date(address) or "Coming Soon"
        results.append({
            "company":      "Kirkland's",
            "address":      address,
            "opening_date": opening_date,
            "link":         KIRKLANDS_URL,
        })

    print(f"[Kirkland's] {len(results)} store(s) parsed.")
    return results


# ── Yard House scraper ───────────────────────────────────────────────────────

YARDHOUSE_URL = "https://www.yardhouse.com/locations/new-locations"


def scrape_yardhouse(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Yard House] Loading {YARDHOUSE_URL}")
    driver.get(YARDHOUSE_URL)
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.new-locations__item"))
        )
    except Exception:
        print("[Yard House] Timed out waiting for location cards; trying anyway…")
    time.sleep(3)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    boxes = soup.select("div.new-locations__item")
    print(f"[Yard House] Found {len(boxes)} location card(s).")

    results = []
    for box in boxes:
        address_tag = box.select_one("div.new-locations__address")
        date_tag    = box.select_one("div.new-locations__date, p.new-locations__date")

        address      = address_tag.get_text(" ", strip=True) if address_tag else ""
        opening_date = date_tag.get_text(strip=True) if date_tag else ""

        if not address:
            continue

        results.append({
            "company":      "Yard House",
            "address":      address,
            "opening_date": extract_date(opening_date) or opening_date,
            "link":         YARDHOUSE_URL,
        })

    print(f"[Yard House] {len(results)} location(s) parsed.")
    return results


# ── Natural Grocers scraper ──────────────────────────────────────────────────

NATURAL_GROCERS_URL = "https://www.naturalgrocers.com/new-store-announcements"


def scrape_natural_grocers(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Natural Grocers] Loading {NATURAL_GROCERS_URL}")
    driver.get(NATURAL_GROCERS_URL)
    time.sleep(8)  # JS-rendered page needs time to settle

    soup = BeautifulSoup(driver.page_source, "html.parser")
    results = []

    # Page is structured as <h2> section heading → <ul> of <li> store entries
    for h2 in soup.find_all("h2"):
        ul = h2.find_next_sibling("ul")
        if not ul:
            continue

        for li in ul.find_all("li", attrs={"data-list-item-id": True}):
            a_tag = li.find("a", href=True)
            if a_tag:
                raw_text = a_tag.get_text(strip=True)
                href     = a_tag["href"]
                link     = href if href.startswith("http") else f"https://www.naturalgrocers.com{href}"
            else:
                raw_text = li.get_text(strip=True)
                link     = NATURAL_GROCERS_URL

            # Entries are formatted as "City, ST - MM/DD/YYYY" or "City, ST - Season YYYY"
            if " - " in raw_text:
                address, opening_date = raw_text.split(" - ", 1)
                address      = address.strip()
                opening_date = opening_date.strip()
            else:
                address      = raw_text
                opening_date = ""

            if not address:
                continue

            results.append({
                "company":      "Natural Grocers",
                "address":      address,
                "opening_date": extract_date(opening_date) or opening_date,
                "link":         link,
            })

    print(f"[Natural Grocers] {len(results)} location(s) parsed.")
    return results


# ── Capital Grille scraper ───────────────────────────────────────────────────

CAPITAL_GRILLE_URL = "https://www.thecapitalgrille.com/locations/new-locations"


def scrape_capital_grille(driver: webdriver.Chrome) -> list[dict]:
    print(f"[Capital Grille] Loading {CAPITAL_GRILLE_URL}")
    driver.get(CAPITAL_GRILLE_URL)
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li.location-card, div.new-locations__item, div[class*='location']"))
        )
    except Exception:
        print("[Capital Grille] Timed out waiting for location cards; trying anyway…")
    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, "html.parser")

    # Primary selector
    cards = soup.find_all("li", class_="location-card")

    # Fallback: LongHorn/Yard House style structure used by other Darden brands
    if not cards:
        print("[Capital Grille] Primary selector found 0 cards — trying fallback selectors.")
        cards = soup.select("div.new-locations__item") or soup.find_all("div", class_=lambda c: c and "newloc" in c.lower())

    print(f"[Capital Grille] Found {len(cards)} card(s).")

    results = []
    for card in cards:
        address_tag  = card.find("div", class_="card-text") or card.find("div", class_=lambda c: c and "address" in c.lower())
        date_tag     = card.find("div", class_="card-status") or card.find("div", class_=lambda c: c and ("date" in c.lower() or "status" in c.lower()))

        address      = address_tag.get_text(strip=True) if address_tag else card.get_text(" ", strip=True)[:150]
        opening_date = date_tag.get_text(strip=True) if date_tag else ""

        if not address:
            continue
        results.append({
            "company":      "Capital Grille",
            "address":      address,
            "opening_date": extract_date(opening_date) or opening_date,
            "link":         CAPITAL_GRILLE_URL,
        })

    print(f"[Capital Grille] {len(results)} location(s) parsed.")
    return results


# ── Hobby Lobby scraper ──────────────────────────────────────────────────────

HOBBY_LOBBY_BASE_URL = "https://newsroom.hobbylobby.com"
HOBBY_LOBBY_LIST_URL = HOBBY_LOBBY_BASE_URL + "/new-stores"
HOBBY_LOBBY_HEADERS  = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}
HOBBY_LOBBY_DATE_RE = re.compile(
    r"\b(?:January|February|March|April|May|June|July|August|September|"
    r"October|November|December)\s+\d{1,2},?\s+\d{4}\b"
)
HOBBY_LOBBY_ADDR_LOCATED = re.compile(
    r"(?:located at|at)\s+([\d]+[^.,]+(?:,\s*[A-Za-z ]+)?)",
    re.IGNORECASE,
)
HOBBY_LOBBY_ADDR_STREET = re.compile(
    r"\b(\d{2,5}\s+[A-Z][a-z]+(?:\s+[A-Za-z]+){1,5}"
    r"\s+(?:St|Ave|Blvd|Dr|Rd|Ln|Way|Pkwy|Hwy|Court|Ct|Place|Pl|Square|Sq|"
    r"Loop|Trail|Trl|Circle|Cir|Row|Mall|Drive|Street|Avenue|Road|Lane|Parkway|Highway))\.?",
    re.IGNORECASE,
)


def _hl_listing_page(page_num: int) -> list[dict]:
    url = f"{HOBBY_LOBBY_LIST_URL}?page={page_num}"
    resp = requests.get(url, headers=HOBBY_LOBBY_HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")
    results = []
    for art in soup.select("article.card"):
        title_tag = art.select_one("h3")
        date_tag  = art.select_one(".text-dark.opacity-75.small, .mb-3.small, time")
        link_tag  = art.select_one("a[href*='/new-stores/']")
        title = title_tag.get_text(strip=True) if title_tag else ""
        date  = date_tag.get_text(strip=True)  if date_tag  else ""
        href  = link_tag["href"]               if link_tag  else ""
        link  = (HOBBY_LOBBY_BASE_URL + href) if href.startswith("/") else href
        results.append({"title": title, "listing_date": date, "link": link})
    return results


def _hl_article(url: str) -> tuple[str, str]:
    try:
        resp = requests.get(url, headers=HOBBY_LOBBY_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [Hobby Lobby] Could not fetch {url}: {e}")
        return "", ""

    soup = BeautifulSoup(resp.text, "html.parser")

    time_tag = soup.select_one("time")
    if time_tag:
        opening_date = time_tag.get_text(strip=True)
    else:
        body_text = soup.get_text(" ", strip=True)
        dates = HOBBY_LOBBY_DATE_RE.findall(body_text)
        opening_date = dates[0] if dates else ""

    paras = soup.select("article p, .entry-content p, main p, .post-content p") or soup.find_all("p")
    body_text = " ".join(p.get_text(" ", strip=True) for p in paras)

    address = ""
    m = HOBBY_LOBBY_ADDR_LOCATED.search(body_text)
    if m:
        address = m.group(1).strip().rstrip(",")
    else:
        m2 = HOBBY_LOBBY_ADDR_STREET.search(body_text)
        if m2:
            address = m2.group(1).strip()

    return address, opening_date


def scrape_hobby_lobby(max_pages: int = 2) -> list[dict]:
    print(f"[Hobby Lobby] Scraping {HOBBY_LOBBY_LIST_URL} ({max_pages} page(s))…")
    cards: list[dict] = []
    for page in range(1, max_pages + 1):
        try:
            page_cards = _hl_listing_page(page)
            if not page_cards:
                print(f"[Hobby Lobby] Page {page}: no articles — stopping.")
                break
            cards.extend(page_cards)
            print(f"[Hobby Lobby] Page {page}: {len(page_cards)} article(s) (total: {len(cards)})")
        except Exception as e:
            print(f"[Hobby Lobby] Page {page} error: {e}")
        time.sleep(0.5)

    results = []
    for i, card in enumerate(cards, 1):
        print(f"  [{i}/{len(cards)}] {card['link']}")
        address, opening_date = _hl_article(card["link"])
        # Fall back to listing date if article page had no date
        if not opening_date:
            opening_date = card.get("listing_date", "")
        # Strip time component from listing date strings like "May 12, 2026 10:30 AM"
        opening_date = re.sub(r"\s+\d{1,2}:\d{2}\s*[AP]M.*", "", opening_date).strip()
        print(f"    → {opening_date or '(no date)'} | {address or '(no address)'}")
        results.append({
            "company":      "Hobby Lobby",
            "address":      address,
            "opening_date": extract_date(opening_date) or opening_date,
            "link":         card["link"],
        })
        time.sleep(0.4)

    print(f"[Hobby Lobby] {len(results)} new store article(s) parsed.")
    return results


# ── Ulta Beauty scraper ──────────────────────────────────────────────────────

ULTA_URL = "https://www.ulta.com/guestservices/ways-to-shop/in-store/grand-openings"
ULTA_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}


def scrape_ulta() -> list[dict]:
    print(f"[Ulta Beauty] Fetching {ULTA_URL}")
    try:
        resp = requests.get(ULTA_URL, headers=ULTA_HEADERS, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"[Ulta Beauty] Error: {e}")
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    cards = soup.find_all("div", class_="StoreListItem__content")
    print(f"[Ulta Beauty] Found {len(cards)} card(s).")

    results = []
    for card in cards:
        try:
            paras = card.find_all_next("p")
            address2 = paras[1].get_text(strip=True) if len(paras) > 1 else ""
            address3 = paras[2].get_text(strip=True) if len(paras) > 2 else ""
            address  = ", ".join(p for p in [address2, address3] if p)

            status_tag   = card.find("div", class_="StoreListItem__statusText")
            opening_date = status_tag.get_text(strip=True) if status_tag else ""

            if not address:
                continue

            results.append({
                "company":      "Ulta Beauty",
                "address":      address,
                "opening_date": extract_date(opening_date) or opening_date,
                "link":         ULTA_URL,
            })
        except Exception as e:
            print(f"  [Ulta Beauty] Error parsing card: {e}")

    print(f"[Ulta Beauty] {len(results)} store(s) parsed.")
    return results


# ── Costco scraper ───────────────────────────────────────────────────────────
#
# costco.ca sits behind a Cloudflare bot check ("Just a moment…" interstitial),
# so this uses Patchright (a Playwright fork with anti-detection patches)
# instead of the regular playwright/Selenium drivers used above. It needs a
# real (non-headless) browser window to clear the Cloudflare challenge — see
# the CI workflow, which runs it under xvfb for that reason.

COSTCO_URL = "https://www.costco.ca/f/-/new-locations"


def scrape_costco() -> list[dict]:
    print(f"[Costco] Loading {COSTCO_URL}")
    try:
        from patchright.sync_api import sync_playwright
    except ImportError:
        print("[Costco] patchright not installed — run: pip install patchright && patchright install chromium")
        return []

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=False,
                args=["--no-sandbox", "--disable-setuid-sandbox"],
            )
            context = browser.new_context(
                viewport={"width": 1280, "height": 800},
                locale="en-US",
            )
            page = context.new_page()
            page.goto(COSTCO_URL, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_function("document.title !== 'Just a moment...'", timeout=20000)
            # Wait until the markdown block that holds the location list is rendered
            page.wait_for_selector('div[data-testid="MarkdownRenderer"]', timeout=15000)
            html = page.content()
            browser.close()
    except Exception as e:
        print(f"[Costco] Error: {e}")
        return []

    soup = BeautifulSoup(html, "html.parser")
    # Each location lives in its own separate MarkdownRenderer div (along with
    # unrelated ones like the membership banner and footer), so scan all of them.
    containers = soup.find_all("div", attrs={"data-testid": "MarkdownRenderer"})
    print(f"[Costco] Found {len(containers)} MarkdownRenderer container(s).")

    results = []
    for container in containers:
        for strong in container.find_all("strong"):
            address = strong.get_text(strip=True)
            if not address:
                continue

            # The opening date sits as plain text right after the <strong> tag,
            # e.g. "<strong>W Roseville, CA</strong> - January 2026"
            date_text = strong.next_sibling
            date_text = str(date_text).strip(" -\n\r\t") if date_text else ""

            results.append({
                "company":      "Costco",
                "address":      address,
                "opening_date": extract_date(date_text) or date_text,
                "link":         COSTCO_URL,
            })

    print(f"[Costco] {len(results)} location(s) parsed.")
    return results


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    all_stores = []

    # ── ALDI ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_aldi(driver))
    except Exception as e:
        print(f"[ALDI] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Dogtopia ──
    try:
        all_stores.extend(scrape_dogtopia())
    except Exception as e:
        print(f"[Dogtopia] Scraping failed: {e}")

    # ── Burlington ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_burlington(driver))
    except Exception as e:
        print(f"[Burlington] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Five Below ──
    try:
        all_stores.extend(scrape_five_below())
    except Exception as e:
        print(f"[Five Below] Scraping failed: {e}")

    # ── HomeGoods ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_homegoods(driver))
    except Exception as e:
        print(f"[HomeGoods] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Homesense ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_homesense(driver))
    except Exception as e:
        print(f"[Homesense] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Jersey Mike's ──
    try:
        all_stores.extend(scrape_jersey_mikes())
    except Exception as e:
        print(f"[Jersey Mike's] Scraping failed: {e}")

    # ── Chick-fil-A ──
    try:
        all_stores.extend(scrape_chick_fil_a())
    except Exception as e:
        print(f"[Chick-fil-A] Scraping failed: {e}")

    # ── LongHorn Steakhouse ──
    try:
        all_stores.extend(scrape_longhorn())
    except Exception as e:
        print(f"[LongHorn] Scraping failed: {e}")

    # ── Trader Joe's ──
    try:
        all_stores.extend(scrape_trader_joes())
    except Exception as e:
        print(f"[Trader Joe's] Scraping failed: {e}")

    # ── Wawa ──
    try:
        all_stores.extend(scrape_wawa())
    except Exception as e:
        print(f"[Wawa] Scraping failed: {e}")

    # ── Target ──
    try:
        all_stores.extend(scrape_target())
    except Exception as e:
        print(f"[Target] Scraping failed: {e}")

    # ── Teso Life ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_teso_life(driver))
    except Exception as e:
        print(f"[Teso Life] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── 7-Eleven ──
    try:
        all_stores.extend(scrape_seven_eleven())
    except Exception as e:
        print(f"[7-Eleven] Scraping failed: {e}")

    # ── Marshalls ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_marshalls(driver))
    except Exception as e:
        print(f"[Marshalls] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── TJ Maxx ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_tjmaxx(driver))
    except Exception as e:
        print(f"[TJ Maxx] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Kirkland's ──
    try:
        all_stores.extend(scrape_kirklands())
    except Exception as e:
        print(f"[Kirkland's] Scraping failed: {e}")

    # ── Yard House ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_yardhouse(driver))
    except Exception as e:
        print(f"[Yard House] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Natural Grocers ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_natural_grocers(driver))
    except Exception as e:
        print(f"[Natural Grocers] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Capital Grille ──
    driver = None
    try:
        driver = make_driver()
        all_stores.extend(scrape_capital_grille(driver))
    except Exception as e:
        print(f"[Capital Grille] Scraping failed: {e}")
    finally:
        if driver:
            driver.quit()

    # ── Hobby Lobby ──
    try:
        all_stores.extend(scrape_hobby_lobby())
    except Exception as e:
        print(f"[Hobby Lobby] Scraping failed: {e}")

    # ── Ulta Beauty ──
    try:
        all_stores.extend(scrape_ulta())
    except Exception as e:
        print(f"[Ulta Beauty] Scraping failed: {e}")

    # ── Costco ──
    try:
        all_stores.extend(scrape_costco())
    except Exception as e:
        print(f"[Costco] Scraping failed: {e}")

    print(f"\nTotal records collected this run: {len(all_stores)}")

    # ── Merge into master + save JSON ──
    os.makedirs("docs", exist_ok=True)
    now_str = datetime.now(timezone.utc).strftime("%B %d, %Y %H:%M UTC")

    merged_records, new_count = update_master(all_stores)

    master_output = {
        "last_updated":   now_str,
        "total_records":  len(merged_records),
        "new_this_run":   new_count,
        "data":           merged_records,
    }
    with open(MASTER_PATH, "w", encoding="utf-8") as f:
        json.dump(master_output, f, indent=2, ensure_ascii=False)
    print(f"Master  → {MASTER_PATH}  ({len(merged_records)} total, {new_count} NEW)")

    # ── Also save current-run snapshot (no is_new / first_seen clutter) ──
    snapshot_output = {
        "last_updated": now_str,
        "data": all_stores,
    }
    snapshot_path = os.path.join("docs", "company_website_latest.json")
    with open(snapshot_path, "w", encoding="utf-8") as f:
        json.dump(snapshot_output, f, indent=2, ensure_ascii=False)
    print(f"Snapshot → {snapshot_path}")

    # ── Save Excel (master data, new rows highlighted yellow) ──
    if merged_records:
        from openpyxl.styles import PatternFill
        NEW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        cols = ["company", "address", "opening_date", "link", "is_new", "first_seen"]
        # Ensure every record has all columns (backfill missing keys)
        for rec in merged_records:
            rec.setdefault("is_new",     False)
            rec.setdefault("first_seen", "")

        df = pd.DataFrame(merged_records, columns=cols)
        today_str  = date.today().strftime("%B %d, %Y")
        excel_path = f"company_website_openings_{today_str}.xlsx"

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            # Sheet 1 – all records, new rows highlighted
            df.to_excel(writer, index=False, sheet_name="All Openings")
            ws_all = writer.sheets["All Openings"]
            for row_idx, is_new in enumerate(df["is_new"], start=2):
                if is_new:
                    for col_idx in range(1, len(cols) + 1):
                        ws_all.cell(row=row_idx, column=col_idx).fill = NEW_FILL
            for col in ws_all.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws_all.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            # Sheet 2 – new records only
            new_df = df[df["is_new"] == True]
            if not new_df.empty:
                new_df.to_excel(writer, index=False, sheet_name="New This Run")
                ws_new = writer.sheets["New This Run"]
                for col in ws_new.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    ws_new.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        new_label = f"{new_count} new row(s) highlighted in yellow" if new_count else "no new rows"
        print(f"Excel   → {excel_path}  ({new_label})")

    # ── CSV master — source of truth for NEW detection ───────────────────────
    # Load existing CSV to find which scraped records are genuinely new.
    # is_new = True only when (company, address) pair is absent from the CSV.
    CW_MASTER_CSV = Path("master_file") / "company_website_master.csv"
    today_str = date.today().strftime("%Y-%m-%d")

    # Build a dict of existing records keyed by _record_key, reset is_new flags
    existing = {}
    if CW_MASTER_CSV.exists():
        df_existing = pd.read_csv(CW_MASTER_CSV, encoding='utf-8').fillna('')
        for _, row in df_existing.iterrows():
            k = _record_key(row.to_dict())
            d = row.to_dict()
            d['is_new'] = False          # clear flag from previous run
            existing[k] = d

    # Overlay fresh scraped records; new = not in existing CSV
    csv_new_count = 0
    for rec in merged_records:
        k = _record_key(rec)
        if k in existing:
            # Update opening_date if we now have one and didn't before
            if rec.get('opening_date') and not existing[k].get('opening_date'):
                existing[k]['opening_date'] = rec['opening_date']
        else:
            existing[k] = {**rec, 'is_new': True,
                           'first_seen': today_str, 'Date_Appended': today_str}
            csv_new_count += 1

    df_csv_master = pd.DataFrame(list(existing.values()))
    df_csv_master = df_csv_master.sort_values(['company', 'address'])
    df_csv_master.to_csv(CW_MASTER_CSV, index=False, encoding='utf-8')
    print(f"CSV Master → {CW_MASTER_CSV}  ({len(df_csv_master)} total rows, {csv_new_count} new)")


if __name__ == "__main__":
    main()
